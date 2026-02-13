from __future__ import annotations
from time import sleep

from nicegui import ui, app
import json
import asyncio
import webbrowser
import os
import html
import json
import re
import datetime
import copy
from io import BytesIO
try:
    from PIL import Image
except ImportError:
    Image = None
try:
    import win32clipboard
except ImportError:
    win32clipboard = None
from server import get_server
from auth_logic import auth_client
import config as cfg
from about_info import ABOUT_INFO
from local_config import local_config

# 确保临时输出目录存在并执行清理
TEMP_PREVIEW_DIR = 'temp_previews'
if not os.path.exists(TEMP_PREVIEW_DIR):
    os.makedirs(TEMP_PREVIEW_DIR)

def cleanup_temp_previews():
    """清理所有临时预览文件"""
    try:
        import shutil
        if os.path.exists(TEMP_PREVIEW_DIR):
            for filename in os.listdir(TEMP_PREVIEW_DIR):
                file_path = os.path.join(TEMP_PREVIEW_DIR, filename)
                try:
                    if os.path.isfile(file_path) or os.path.islink(file_path):
                        os.unlink(file_path)
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)
                except Exception as e:
                    print(f'清理文件失败 {file_path}: {e}')
    except Exception as e:
        print(f'清理目录失败: {e}')

# 启动时清理一次
cleanup_temp_previews()

app.add_static_files('/test_output', 'test_output')
app.add_static_files(f'/{TEMP_PREVIEW_DIR}', TEMP_PREVIEW_DIR)

# 1. 获取服务器单例
ps_server = get_server()

# --- 登录状态 ---
# 用于控制“未登录时不触发 PS/策略/剪贴板相关提示”
auth_logged_in: bool = False

# --- UI 交互逻辑 ---
status_dot = None
status_text = None
error_text = None

def check_connection_status():
    """定时检测连接状态"""
   

# --- 状态标志 ---
initialized = False
consecutive_heartbeat_failures = 0

def format_multiline_text(text: str) -> str:
    """
    预处理服务器返回的长文本，支持 \n 换行并兼容 Markdown
    """
    if not text or str(text).lower() == "null":
        return ""
    # 将单换行符替换为 Markdown 的硬换行（两个空格+换行）
    return text.replace('\n', '  \n')

async def show_message_dialog(
    title: str = '系统公告',
    content: str = '',
    confirm_text: str = '我知道了'
):
    """
    通用信息展示对话框
    """
    formatted_content = format_multiline_text(content)
    with ui.dialog() as dialog, ui.card().classes('p-8 rounded-3xl shadow-2xl').style('min-width: 360px; max-width: 500px;'):
        # 标题
        ui.label(title).classes('text-xl font-bold text-gray-800 mb-4 tracking-tight')
        
        # 内容区域 (支持多行并自适应高度)
        with ui.scroll_area().style('max-height: 300px;'):
            ui.markdown(formatted_content).classes('text-sm text-gray-600 leading-relaxed')
        
        # 确认按钮
        with ui.row().classes('w-full justify-end mt-6'):
            ui.button(confirm_text, on_click=dialog.close).classes(
                'px-8 py-2 rounded-xl bg-cyan-600 text-white font-medium shadow-md'
            )
    
    await dialog

async def show_heartbeat_error_dialog():
    """显示心跳包失败对话框（强力拦截，必须退出）"""
    with ui.dialog().props('persistent') as dialog, ui.card().classes('p-8 rounded-3xl shadow-2xl').style('min-width: 400px;'):
        ui.label('与服务器失去连接').classes('text-2xl font-bold text-red-600 mb-2')
        ui.label('连续多次尝试连接验证服务器失败，为了保障您的账号安全，请检查网络连接并重新登录。').classes('text-sm text-gray-500 mb-8 leading-relaxed')
        
        async def do_exit():
            app.shutdown()

        ui.button('确认并退出', on_click=do_exit).classes(
            'w-full py-4 rounded-2xl bg-red-600 text-white font-bold shadow-lg'
        )
    await dialog

def show_about_dialog():
    """显示关于对话框"""
    with ui.dialog() as dialog, ui.card().classes('p-10 rounded-[32px] ice-card bg-white items-center gap-6').style('min-width: 360px;'):
        # Logo - 蓝底圆角矩形包装
        with ui.element('div').classes('ice-logo-badge').style('width: 80px; height: 80px; border-radius: 20px;'):
            ui.html(ABOUT_INFO['logo_svg'], sanitize=False)
        
        # 标题与名称
        with ui.column().classes('items-center gap-1'):
            ui.label('ice美化助手').classes('text-2xl ice-title text-cyan-700 font-bold')
            ui.label('ICE Tools').classes('text-[10px] text-gray-400 tracking-widest uppercase font-medium')
        
        # 描述
        ui.label(ABOUT_INFO['description']).classes('text-sm text-gray-500 text-center leading-relaxed')
        
        # 分隔线
        ui.element('div').classes('w-12 h-1 bg-cyan-100 rounded-full my-2')
        
        # 版本与版权
        with ui.column().classes('items-center gap-1'):
            ui.label(f"版本 {ABOUT_INFO['version']}").classes('text-xs text-gray-400 font-mono')
            ui.label(ABOUT_INFO['copyright']).classes('text-[10px] text-gray-300 uppercase tracking-tight')
    
    dialog.open()

async def do_heartbeat():
    """执行单次心跳检测"""
    global consecutive_heartbeat_failures
    print(f">>> [DEBUG] 正在发送心跳请求... (当前连续失败次数: {consecutive_heartbeat_failures})")
    
    res = await auth_client.heartbeat()
    if res.get("status") == "success":
        consecutive_heartbeat_failures = 0
    else:
        consecutive_heartbeat_failures += 1
        print(f"心跳包失败 ({consecutive_heartbeat_failures}/{cfg.HEARTBEAT_MAX_RETRIES}): {res.get('message')}")
        
        if consecutive_heartbeat_failures >= cfg.HEARTBEAT_MAX_RETRIES:
            # 达到失败上限，停止心跳并弹出报错
            heartbeat_timer.deactivate()
            await show_heartbeat_error_dialog()

# 心跳定时器 (初始化时不启动)
heartbeat_timer = ui.timer(cfg.HEARTBEAT_INTERVAL, do_heartbeat, active=False)

async def init():
    """初始化函数：拉取公告并检查版本"""
    global initialized
    if initialized:
        return
    initialized = True
    
    try:
        config_res = await auth_client.get_config()
        if not isinstance(config_res, dict) or config_res.get("status") != "success":
            print(f"拉取配置失败: {config_res}")
            return

        public_cfg = config_res.get("public", {})
        
        # 1. 检查公告 (Notice)
        notice = public_cfg.get("notice")
        if notice and str(notice).strip().lower() != "null":
            await show_message_dialog(
                title="系统公告",
                content=notice,
                confirm_text="确定"
            )

        # 2. 检查版本
        latest_version = public_cfg.get("version")
        if latest_version and str(latest_version) != str(cfg.APP_VERSION):
            update_url = public_cfg.get("update_url", cfg.OFFICIAL_WEBSITE)
            update_notes = public_cfg.get("update_notes", "有新版本可用，请前往官网下载。")
            
            # 弹出专用升级对话框 (增加 persistent 属性，防止点击背景取消)
            with ui.dialog().props('persistent') as upgrade_dialog, ui.card().classes('p-8 rounded-3xl shadow-2xl').style('min-width: 400px;'):
                ui.label('发现新版本').classes('text-2xl font-bold text-cyan-700 mb-2')
                
                with ui.row().classes('items-center gap-4 mb-6'):
                    with ui.column().classes('gap-0'):
                        ui.label('当前版本').classes('text-[10px] text-gray-400 tracking-widest uppercase')
                        ui.label(cfg.APP_VERSION).classes('text-sm font-mono text-gray-500')
                    ui.icon('east', color='gray-300')
                    with ui.column().classes('gap-0'):
                        ui.label('最新版本').classes('text-[10px] text-cyan-400 tracking-widest uppercase')
                        ui.label(latest_version).classes('text-sm font-mono text-cyan-700 font-bold')
                
                ui.label('更新公告').classes('text-xs font-bold text-gray-400 uppercase tracking-widest mb-2')
                with ui.scroll_area().style('height: 120px;').classes('bg-gray-50 rounded-xl p-4 mb-8 border border-gray-100'):
                    ui.markdown(format_multiline_text(update_notes)).classes('text-sm text-gray-600')
                
                async def do_update():
                    webbrowser.open(update_url)
                    app.shutdown()

                ui.button('立即前往下载', on_click=do_update).classes(
                    'w-full py-4 rounded-2xl bg-gradient-to-r from-cyan-500 to-cyan-600 text-white font-bold shadow-lg'
                )
            
            await upgrade_dialog
            return

    except Exception as e:
        print(f"初始化检查失败: {e}")

    # 注：不再在启动时自动 start ps_server，保持静默。
    # await ps_server.start()
    print("应用初始化完成，待用户登录...")

# --- 样式 (保持不变) ---
ui.add_head_html('<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">')
ui.colors(primary='#06b6d4', secondary='#0891b2', positive='#10b981', negative='#ef4444', warning='#f59e0b', info='#3b82f6')
ui.add_css('''
  /* ... 你的 CSS 代码保持不变 ... */
  :root { --ice-primary: #06b6d4; --ice-primary-dark: #0891b2; }
  .ice-card { border-radius: 32px; box-shadow: 0 12px 40px rgba(0,0,0,0.12), 0 4px 12px rgba(0,0,0,0.06); }
  .ice-auth-card { border-radius: 32px; box-shadow: 0 12px 40px rgba(0,0,0,0.12), 0 4px 12px rgba(0,0,0,0.06); background: rgba(255, 255, 255, 0.98); }
  .ice-title { font-weight: 700; letter-spacing: -0.01em; }
  html, body { height: 100%; overflow: hidden; margin: 0; padding: 0; }
  /* 强制隐藏所有滚动条，增强原生感 */
  ::-webkit-scrollbar { display: none; }
  * { scrollbar-width: none; -ms-overflow-style: none; }
  
  body { background: radial-gradient(circle at 50% 50%, #164e63 0%, #083344 100%); }
  .ice-shell { position: fixed; top: 0; left: 0; right: 0; bottom: 0; display: flex; align-items: center; justify-content: center; padding: 24px; overflow: hidden; gap: 24px; }
  .ice-logo-wrap { display: flex; justify-content: center; margin-bottom: 18px; }
  .ice-logo-badge { width: 86px; height: 86px; border-radius: 22px; display: inline-flex; align-items: center; justify-content: center; color: #ffffff; background: linear-gradient(135deg, #22d3ee 0%, var(--ice-primary-dark) 100%); box-shadow: 0 4px 16px rgba(6, 182, 212, 0.25); }
  .ice-win-controls { position: fixed; top: 12px; right: 12px; z-index: 9999; display: flex; gap: 8px; }
  .ice-win-btn-minimal { width: 20px !important; height: 20px !important; min-width: 20px !important; min-height: 20px !important; background: transparent !important; border: none !important; display: inline-flex !important; align-items: center !important; justify-content: center !important; cursor: pointer !important; transition: opacity 0.2s ease !important; padding: 0 !important; margin: 0 !important; opacity: 0.85; box-shadow: none !important; }
  .ice-win-btn-minimal:hover { opacity: 1 !important; background: transparent !important; }
  .ice-win-btn-minimal .q-btn__content { padding: 0 !important; }
  .ice-win-btn-minimal svg { width: 20px !important; height: 20px !important; color: #ffffff !important; filter: drop-shadow(0 1px 2px rgba(0,0,0,0.2)); }
  .ice-footer { margin-top: auto; padding-top: 20px; padding-bottom: 4px; display: flex; flex-direction: column; align-items: center; gap: 6px; width: 100%; text-align: center; }
  .ice-footer-links { display: flex; gap: 24px; justify-content: center; width: 100%; }
  .ice-footer-link { color: var(--ice-primary); text-decoration: none; font-size: 12px; cursor: pointer; transition: color 0.2s; font-weight: 500; }
  .ice-footer-link:hover { color: var(--ice-primary-dark); }
  .ice-footer-text { color: #9ca3af; font-size: 11px; text-align: center; width: 100%; letter-spacing: 0.02em; }
  .q-tab--active { color: var(--ice-primary) !important; }
  .q-tabs__indicator { background: var(--ice-primary) !important; display: none !important; }

  /* 优化通知弹窗样式 */
  .q-notification { border-radius: 16px !important; padding: 12px 20px !important; box-shadow: 0 8px 32px rgba(0,0,0,0.12) !important; }
  
  /* 登录注册核销页面输入框圆角矩形样式 */
  .ice-auth-card .q-field--outlined .q-field__control {
    border-radius: 12px !important;
    background: rgba(255, 255, 255, 0.5) !important;
    transition: all 0.3s ease;
  }
  .ice-auth-card .q-field--outlined .q-field__control:hover {
    background: rgba(255, 255, 255, 0.8) !important;
  }
  .ice-auth-card .q-field--focused .q-field__control {
    background: white !important;
    box-shadow: 0 4px 12px rgba(6, 182, 212, 0.1) !important;
  }
  /* 修正 Quasar 内部 border-radius */
  .ice-auth-card .q-field--outlined .q-field__control:before {
    border-radius: 12px !important;
    border-color: rgba(6, 182, 212, 0.2) !important;
  }
  .ice-auth-card .q-field--outlined .q-field__control:after {
    border-radius: 12px !important;
  }

  /* 状态组件动画 */
  .status-spin { animation: spin 1.2s linear infinite; }
  @keyframes spin { from { transform: rotate(0deg); } to { transform: rotate(360deg); } }
  
  .checkmark-draw {
    stroke-dasharray: 48;
    stroke-dashoffset: 48;
    animation: draw 0.6s cubic-bezier(0.65, 0, 0.45, 1) forwards;
  }
  @keyframes draw { to { stroke-dashoffset: 0; } }
  
  .scale-in { animation: scaleIn 0.4s cubic-bezier(0.175, 0.885, 0.32, 1.275) forwards; }
  @keyframes scaleIn { from { transform: scale(0); opacity: 0; } to { transform: scale(1); opacity: 1; } }
  
  /* 页面过渡动画 */
  .page-fade-in { animation: pageFadeIn 0.6s cubic-bezier(0.22, 1, 0.36, 1) forwards; }
  @keyframes pageFadeIn { from { opacity: 0; transform: scale(0.98); filter: blur(10px); } to { opacity: 1; transform: scale(1); filter: blur(0); } }
  
  .page-fade-out { animation: pageFadeOut 0.4s ease forwards; pointer-events: none; }
  @keyframes pageFadeOut { from { opacity: 1; transform: scale(1); } to { opacity: 0; transform: scale(1.05); filter: blur(5px); } }

  /* 极速出图面板样式 */
  .ice-rapid-panel {
    position: absolute;
    top: 0;
    right: 0;
    width: 280px;
    height: 100%;
    background: rgba(255, 255, 255, 0.95);
    backdrop-filter: blur(20px);
    border-left: 1px solid rgba(0, 0, 0, 0.05);
    box-shadow: -10px 0 30px rgba(0, 0, 0, 0.05);
    z-index: 50;
    transition: all 0.5s cubic-bezier(0.34, 1.56, 0.64, 1);
    display: flex;
    flex-direction: column;
    padding: 24px;
  }
  .ice-rapid-collapsed {
    transform: translateX(100%);
    box-shadow: none;
  }
  .ice-rapid-handle {
    position: absolute;
    left: -28px;
    top: 50%;
    transform: translateY(-50%);
    width: 28px;
    height: 64px;
    background: rgba(255, 255, 255, 0.9);
    backdrop-filter: blur(10px);
    border-radius: 12px 0 0 12px;
    border: 1px solid rgba(0, 0, 0, 0.05);
    border-right: none;
    display: flex;
    align-items: center;
    justify-content: center;
    cursor: pointer;
    color: var(--ice-primary);
    box-shadow: -4px 0 10px rgba(0, 0, 0, 0.03);
    transition: all 0.2s;
  }
  .ice-rapid-handle:hover {
    color: var(--ice-primary-dark);
    padding-right: 4px;
  }
  .ice-green-groove {
    background: #f1f5f9;
    border-radius: 16px;
    padding: 8px 12px;
    box-shadow: inset 0 2px 4px rgba(0, 0, 0, 0.05);
    display: flex;
    flex-wrap: wrap;
    gap: 6px;
    align-items: center;
    min-height: 40px;
    border: 1px solid rgba(0, 0, 0, 0.02);
  }
  .ice-capsule {
    padding: 3px 10px;
    border-radius: 8px;
    font-size: 10.5px;
    font-weight: 600;
    display: flex;
    align-items: center;
    gap: 4px;
    box-shadow: 0 1px 2px rgba(0, 0, 0, 0.05);
  }
  .ice-capsule-text {
    background: white;
    color: #3b82f6;
    border: 1px solid #dbeafe;
  }
  .ice-capsule-image {
    background: white;
    color: #8b5cf6;
    border: 1px solid #ede9fe;
  }
  .ice-rapid-space {
    background: #ecfdf5;
    color: #10b981;
    border: 1px solid #d1fae5;
    padding: 2px 6px;
    border-radius: 6px;
    font-size: 10px;
    font-weight: bold;
    display: flex;
    align-items: center;
    justify-content: center;
    box-shadow: 0 1px 2px rgba(16, 185, 129, 0.05);
  }
  .ice-rapid-textarea .q-textarea {
    background: white !important;
  }
  .ice-rapid-textarea .q-field__control {
    border-radius: 12px !important;
    padding: 8px 12px !important;
  }
  .ice-rapid-textarea textarea {
    font-size: 12px !important;
    line-height: 1.5 !important;
  }
  .ice-batch-uploader .q-uploader {
    width: 100% !important;
    border: 1px solid #e2e8f0 !important;
    border-radius: 12px !important;
    background: #f8fafc !important;
    min-height: 92px !important;
    transition: all 0.22s ease !important;
    box-shadow: inset 0 1px 0 rgba(255, 255, 255, 0.95), inset 0 0 0 1px rgba(255, 255, 255, 0.32), inset 0 6px 14px rgba(15, 23, 42, 0.03);
    cursor: pointer;
    overflow: hidden;
  }
  .ice-batch-uploader .q-uploader::before {
    content: "";
    position: absolute;
    inset: 0;
    border-radius: 12px;
    border: 1.2px dashed rgba(148, 163, 184, 0.35);
    pointer-events: none;
  }
  .ice-batch-uploader .q-uploader:hover {
    border-color: #22d3ee !important;
    background: #f0f9ff !important;
    box-shadow: inset 0 1px 0 rgba(255, 255, 255, 0.95), inset 0 0 0 1px rgba(34, 211, 238, 0.16), inset 0 8px 16px rgba(6, 182, 212, 0.06);
    transform: none;
  }
  .ice-batch-uploader .q-uploader__header {
    min-height: 92px !important;
    background: transparent !important;
    color: #64748b !important;
    padding: 10px 12px !important;
  }
  .ice-batch-uploader .q-uploader__header-content {
    width: 100%;
    justify-content: center;
    text-align: center;
    gap: 4px;
  }
  .ice-batch-uploader .q-uploader__header-content::before {
    content: "table_view";
    font-family: "Material Icons";
    font-size: 20px;
    line-height: 1;
    color: #06b6d4;
    opacity: 0.9;
    display: block;
    margin-bottom: 2px;
  }
  .ice-batch-uploader .q-uploader__title {
    font-size: 12px !important;
    font-weight: 700 !important;
    color: #0f172a !important;
    letter-spacing: 0.01em;
  }
  .ice-batch-uploader .q-uploader__subtitle {
    display: none !important;
  }
  .ice-batch-uploader .q-uploader__list,
  .ice-batch-uploader .q-uploader__header .q-btn,
  .ice-batch-uploader .q-linear-progress,
  .ice-batch-uploader .q-uploader__header-content .q-chip,
  .ice-batch-uploader .q-uploader__header-content .q-uploader__counter {
    display: none !important;
  }
  .ice-batch-uploader-note {
    font-size: 9px;
    color: #94a3b8;
    padding-left: 6px;
    letter-spacing: 0.02em;
  }
  .ice-preview-table-wrap {
    border: 1px solid #e2e8f0;
    border-radius: 14px;
    background: #f8fafc;
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.95), inset 0 10px 20px rgba(15, 23, 42, 0.03);
    overflow: hidden;
  }
  .ice-preview-table-header {
    background: #f1f5f9;
    border-bottom: 1px solid #dbe5ef;
  }
  .ice-preview-table-row {
    width: 100%;
    margin: 0 !important;
    border-bottom: 1px solid #eef2f7;
    background: #ffffff;
    transition: background 0.18s ease;
  }
  .ice-preview-table-row:hover {
    background: #f8fbff;
  }
  .ice-preview-cell {
    min-height: 34px;
    display: flex;
    align-items: center;
    padding: 0 10px;
    font-size: 11px;
    color: #334155;
    border-right: 1px solid #eef2f7;
  }
  .ice-preview-cell:last-child {
    border-right: none;
  }
  .ice-preview-cell-head {
    min-height: 34px;
    font-size: 10px;
    font-weight: 700;
    color: #64748b;
    text-transform: uppercase;
    letter-spacing: 0.04em;
  }
  .ice-preview-col-check-head, .ice-preview-col-serial-head {
    background: #eef2f7;
    color: #64748b;
  }
  .ice-preview-col-var-head {
    background: #ecfeff;
    color: #0e7490;
  }
  .ice-preview-col-check, .ice-preview-col-serial {
    background: #f1f5f9 !important;
  }
  .ice-preview-col-var {
    background: #ffffff;
    color: #1e293b;
  }
  .ice-preview-scroll .q-scrollarea__content {
    padding: 0 !important;
  }
  .ice-preview-grid-row {
    display: grid !important;
    width: 100%;
    align-items: stretch;
  }
  .ice-preview-grid-row.ice-preview-table-header {
    position: sticky;
    top: 0;
    z-index: 3;
  }
  .ice-preview-grid-row .ice-preview-cell {
    min-width: 0;
  }
  .ice-btn-terminate {
    padding: 4px 12px;
    border-radius: 8px;
    background: #fef2f2;
    color: #ef4444;
    border: 1px solid #fee2e2;
    font-size: 10px;
    font-weight: bold;
    cursor: pointer;
    transition: all 0.2s;
    display: flex;
    align-items: center;
    gap: 4px;
  }
  .ice-btn-terminate:hover {
    background: #fee2e2;
    transform: translateY(-1px);
  }
  
  .ice-shimmer-bar {
    position: relative;
    overflow: hidden;
  }
  .ice-shimmer-bar::after {
    content: '';
    position: absolute;
    top: 0;
    left: -100%;
    width: 50%;
    height: 100%;
    background: linear-gradient(90deg, transparent, rgba(255,255,255,0.4), transparent);
    animation: shimmer 1.5s infinite;
  }
  @keyframes shimmer {
    100% { left: 200%; }
  }

  /* 工作台容器样式 */
  .ice-workspace-container { position: fixed; top: 0; left: 0; right: 0; bottom: 0; display: flex; align-items: center; justify-content: center; width: 100%; height: 100vh; background: transparent; z-index: 1; padding: 20px; opacity: 0; visibility: hidden; }
  .ice-workspace-container.show { visibility: visible; }
  .ice-workspace-card { display: flex; width: 100%; max-width: 860px; height: 560px; background: rgba(255, 255, 255, 0.98); backdrop-filter: blur(25px); border-radius: 24px; box-shadow: 0 30px 60px rgba(0,0,0,0.25); overflow: hidden; position: relative; border: 1px solid rgba(255, 255, 255, 0.4); }
  
  .ice-sidebar { width: 72px; display: flex; flex-direction: column; align-items: center; padding: 24px 0; justify-content: space-between; position: relative; z-index: 10; background: #F8FAFC; border-right: 1px solid #e2e8f0; height: 100%; }
  
  /* 模式按钮与分体式 Dock */
  .ice-dock-top, .ice-dock-bottom { display: flex; flex-direction: column; align-items: center; width: 100%; }
  .ice-dock-top { gap: 12px; }
  .ice-dock-bottom { gap: 12px; }

  .ice-mode-btn { width: 42px; height: 42px; border-radius: 12px; background: #ecfeff; border: none; display: flex; align-items: center; justify-content: center; cursor: pointer; transition: all 0.25s cubic-bezier(0.34, 1.56, 0.64, 1); position: relative; padding: 0 !important; color: #06b6d4; }
  .ice-mode-btn:hover { background: #cffafe; transform: scale(1.05); }
  .ice-mode-btn.active { background: linear-gradient(135deg, #06b6d4 0%, #0891b2 100%); box-shadow: 0 4px 12px rgba(6, 182, 212, 0.4); color: white; }
  .ice-mode-btn.active svg { color: #ffffff !important; }
  .ice-mode-btn svg { width: 20px; height: 20px; color: inherit; transition: none; }
  
  /* 精美悬浮气泡 - 参考 test.html */
  .ice-mode-info-bar { 
    position: absolute; 
    left: 56px; 
    top: 50%; 
    transform: translateY(-50%) translateX(-4px); 
    background: linear-gradient(to right, #a5f3fc, #67e8f9); 
    color: #0e7490; 
    display: flex; 
    align-items: center; 
    padding: 6px 12px; 
    border-radius: 8px; 
    white-space: nowrap; 
    z-index: 100; 
    transition: all 0.2s ease; 
    opacity: 0; 
    pointer-events: none; 
    box-shadow: 0 4px 6px rgba(0,0,0,0.05); 
    font-size: 12px; 
    font-weight: 600;
  }
  .ice-mode-info-bar::before { 
    content: ""; 
    position: absolute; 
    left: -4px; 
    top: 50%; 
    transform: translateY(-50%); 
    border-width: 5px 5px 5px 0; 
    border-style: solid; 
    border-color: transparent #a5f3fc transparent transparent; 
  }
  .ice-mode-btn:hover .ice-mode-info-bar { opacity: 1; transform: translateY(-50%) translateX(4px); }
  
  /* 激活状态气泡微调 */
  .ice-mode-btn.active .ice-mode-info-bar { background: #a5f3fc; }
  .ice-mode-info-text { color: inherit; }

  /* PS连接状态指示点 */
  .ice-ps-indicator { display: flex; flex-direction: column; align-items: center; position: relative; cursor: pointer; padding: 4px 0; }
  .ice-ps-dot { width: 8px; height: 8px; border-radius: 50%; transition: all 0.3s ease; }
  .ice-ps-dot.connected { background: #10b981; box-shadow: 0 0 0 3px rgba(16, 185, 129, 0.2); animation: pulse-green 2s infinite; }
  .ice-ps-dot.disconnected { background: #9ca3af; }
  .ice-ps-dot.waiting { background: #06b6d4; animation: spin 1.2s linear infinite; }
  
  /* 底部图标按钮风格统一 */
  .ice-bottom-btn { width: 42px; height: 42px; display: flex; align-items: center; justify-content: center; cursor: pointer; color: #94a3b8; transition: all 0.2s ease; position: relative; background: transparent; }
  .ice-bottom-btn:hover { color: #06b6d4; transform: scale(1.05); }
  .ice-bottom-btn svg { width: 18px; height: 18px; }

  /* 底部通用 Tooltip */
  .ice-bottom-tooltip { 
    position: absolute; 
    left: 56px; 
    top: 50%; 
    transform: translateY(-50%); 
    background: #f1f5f9; 
    color: #64748b; 
    padding: 6px 12px; 
    border-radius: 8px; 
    font-size: 11px; 
    white-space: nowrap; 
    opacity: 0; 
    transition: all 0.2s ease; 
    pointer-events: none; 
    box-shadow: 0 4px 6px rgba(0,0,0,0.05); 
    font-weight: 500;
  }
  .ice-bottom-tooltip::before { 
    content: ""; 
    position: absolute; 
    left: -4px; 
    top: 50%; 
    transform: translateY(-50%); 
    border-width: 5px 5px 5px 0; 
    border-style: solid; 
    border-color: transparent #f1f5f9 transparent transparent; 
  }
  .ice-ps-indicator:hover .ice-bottom-tooltip, .ice-bottom-btn:hover .ice-bottom-tooltip { opacity: 1; left: 60px; }
  
  /* 侧边栏底部 Logo */
  .ice-sidebar-logo { width: 42px; height: 42px; border-radius: 14px; background: linear-gradient(135deg, #22d3ee 0%, #0891b2 100%); display: flex; align-items: center; justify-content: center; color: white; box-shadow: 0 4px 12px rgba(6, 182, 212, 0.25); cursor: default; transition: transform 0.3s ease; }
  .ice-sidebar-logo:hover { transform: rotate(10deg) scale(1.05); }
  .ice-sidebar-logo svg { width: 22px; height: 22px; }
  
  /* --- 制作模板工作台样式 --- */
  
  .ice-top-bar { height: 42px; width: 100%; padding: 0 16px; display: flex; align-items: flex-end; padding-bottom: 2px; flex-shrink: 0; background: transparent; z-index: 5; }
  .ice-doc-selector { background: white; padding: 3px 10px; border-radius: 12px; border: 1px solid transparent; display: flex; align-items: center; gap: 6px; cursor: pointer; box-shadow: 0 1px 2px rgba(0,0,0,0.05); transition: all 0.2s; }
  .ice-doc-selector:hover { border-color: var(--ice-primary); transform: translateY(-1px); }
  .ice-doc-name { font-size: 12px; font-weight: 600; color: #334155; }

  .ice-workspace-inner { display: flex; gap: 10px; padding: 0 14px 10px 14px; width: 100%; flex: 1 1 0%; min-height: 0; overflow: hidden; align-items: stretch; }
  
  .ice-left-panel { width: 200px; height: 100%; min-height: 0; background: white; border-radius: 20px; padding: 10px; display: flex; flex-direction: column; box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.05); flex-shrink: 0; overflow: hidden; }
  .ice-right-panel { flex: 1 1 0%; height: 100%; min-height: 0; background: white; border-radius: 20px; box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.05); padding: 14px 16px 80px 16px; overflow-y: auto; display: flex; flex-direction: column; gap: 10px; }

  .ice-search-box { background: #f8fafc; border-radius: 10px; padding: 4px 10px; font-size: 11px; color: #64748b; margin-bottom: 6px; display: flex; align-items: center; gap: 6px; border: 1px solid #e2e8f0; height: 28px; }
  
  .ice-layer-item { display: flex; align-items: center; justify-content: space-between; padding: 8px; border-radius: 8px; cursor: pointer; font-size: 12px; color: #334155; transition: background 0.1s; position: relative; }
  .ice-layer-item:hover { background-color: #ecfeff; }
  .ice-layer-info { display: flex; align-items: center; gap: 8px; overflow: hidden; }
  .ice-btn-add-layer { width: 20px; height: 20px; border-radius: 50%; background: var(--ice-primary); color: white; display: flex; align-items: center; justify-content: center; font-size: 10px; opacity: 0; transform: scale(0.5); transition: all 0.2s; cursor: pointer; }
  .ice-layer-item:hover .ice-btn-add-layer { opacity: 1; transform: scale(1); }
  
  .ice-section-header { font-size: 11px; font-weight: 700; color: #64748b; margin: 12px 0 4px 0; text-transform: uppercase; letter-spacing: 0.5px; display: flex; align-items: center; gap: 6px; }
  .ice-config-card { background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 12px; padding: 10px; margin-bottom: 4px; display: flex; flex-direction: column; gap: 6px; animation: cardIn 0.3s ease-out; }
  @keyframes cardIn { from { opacity: 0; transform: translateY(10px); } to { opacity: 1; transform: translateY(0); } }
  
  .ice-card-header { display: flex; justify-content: space-between; align-items: center; }
  .ice-tag-type { font-size: 9px; font-weight: 800; padding: 2px 5px; border-radius: 4px; text-transform: uppercase; }
  .ice-tag-text { color: #3b82f6; background: #eff6ff; border: 1px solid #dbeafe; }
  .ice-tag-img { color: #8b5cf6; background: #f5f3ff; border: 1px solid #ede9fe; }
  .ice-card-title { font-size: 12px; font-weight: 600; color: #334155; max-width: 150px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
  
  .ice-regex-item { display: flex; align-items: center; gap: 8px; background: white; padding: 6px 10px; border-radius: 8px; border: 1px solid #e2e8f0; margin-top: 4px; }
  
  .ice-float-dock { position: absolute; bottom: 24px; left: 50%; transform: translateX(-50%); background: rgba(255, 255, 255, 0.9); backdrop-filter: blur(12px); padding: 8px 16px; border-radius: 24px; box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.1); border: 1px solid rgba(255,255,255,0.6); display: flex; gap: 12px; align-items: center; z-index: 100; }
  .ice-dock-sep { width: 1px; height: 20px; background: #e2e8f0; }
  .ice-float-btn { width: 36px; height: 36px; border-radius: 10px; display: flex; align-items: center; justify-content: center; cursor: pointer; transition: all 0.2s; color: #64748b; background: transparent; }
  .ice-float-btn:hover { background: #f1f5f9; color: var(--ice-primary); transform: translateY(-2px); }
  .ice-float-btn.primary { background: linear-gradient(135deg, var(--ice-primary) 0%, var(--ice-primary-dark) 100%); color: white; box-shadow: 0 4px 12px rgba(6, 182, 212, 0.3); }
  
  /* 工作台主内容区 */
  .ice-workspace-content { flex: 1; display: flex; flex-direction: column; padding: 0; background-color: #f1f5f9; position: relative; overflow: hidden; box-shadow: inset 0 2px 8px rgba(0,0,0,0.04); height: 100%; min-height: 0; }

  /* 下拉菜单美化 */
  .ice-dropdown-menu { border-radius: 16px !important; box-shadow: 0 12px 32px rgba(0,0,0,0.12) !important; padding: 8px !important; background: white !important; border: 1px solid #f1f5f9 !important; min-width: 120px; width: auto; overflow: visible !important; }
    .ice-dropdown-menu::before { content: ""; position: absolute; top: -5px; left: 20px; width: 10px; height: 10px; background: white; transform: rotate(45deg); border-left: 1px solid #f1f5f9; border-top: 1px solid #f1f5f9; border-top-left-radius: 3px; z-index: 0; }
  .ice-email-menu::before, .ice-filter-menu::before { left: auto !important; right: 20px !important; }
  .ice-dropdown-item { border-radius: 10px !important; margin: 2px 0 !important; transition: all 0.2s !important; min-height: 38px !important; padding: 0 12px !important; color: #64748b !important; font-size: 13px !important; display: flex; align-items: center; cursor: pointer; position: relative; z-index: 1; }
  .ice-dropdown-item:hover { background: #f8fafc !important; color: #06b6d4 !important; }
  .ice-dropdown-item.active { background: #06b6d4 !important; color: white !important; font-weight: 600 !important; }
  .ice-dropdown-sep { height: 1px; background: #f1f5f9; margin: 6px 4px; }
  .ice-dropdown-add-btn { border: 1.5px dashed #e2e8f0 !important; border-radius: 12px !important; margin-top: 4px !important; color: #94a3b8 !important; font-size: 12px !important; cursor: pointer !important; transition: all 0.2s !important; height: 36px !important; display: flex !important; align-items: center !important; justify-content: center !important; gap: 6px !important; }
  .ice-dropdown-add-btn:hover { border-color: #06b6d4 !important; color: #06b6d4 !important; background: #ecfeff !important; }
  .ice-btn-delete { color: #94a3b8 !important; transition: all 0.2s ease !important; cursor: pointer; }
  .ice-btn-delete:hover { color: #ef4444 !important; transform: scale(1.1); }

  /* 正则编辑器弹窗样式 */
  .regex-dialog-card { width: 480px; max-width: 95vw; max-height: 90vh; border-radius: 20px !important; overflow: hidden; background: white; display: flex; flex-direction: column; }
  .regex-section-label { font-size: 12px; font-weight: 600; color: #334155; margin-bottom: 4px; }
  .regex-input-group { background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 10px; padding: 2px 10px; transition: all 0.2s; }
  .regex-input-group:focus-within { border-color: #06b6d4; box-shadow: 0 0 0 3px rgba(6, 182, 212, 0.1); background: white; }
  .regex-test-area { background: #f1f5f9; border-radius: 14px; padding: 12px; margin-top: 8px; }
  .regex-test-box { background: white; border: 1px solid #e2e8f0; border-radius: 10px; padding: 8px; height: 80px; font-family: monospace; font-size: 11px; }
  .regex-ai-card { background: linear-gradient(135deg, #f5f3ff 0%, #ede9fe 100%); border: 1px solid #ddd6fe; border-radius: 14px; padding: 12px; display: flex; gap: 12px; align-items: flex-start; }
  .regex-ai-icon { width: 32px; height: 32px; background: white; border-radius: 10px; display: flex; align-items: center; justify-content: center; box-shadow: 0 4px 6px rgba(139, 92, 246, 0.1); flex-shrink: 0; }
  .regex-save-btn { background: #06b6d4; color: white; border-radius: 10px; padding: 0 20px; height: 38px; font-weight: 600; font-size: 13px; }
  .ice-rounded-input .q-field__control { border-radius: 12px !important; }
''')

# --- 策略解析与加载逻辑 ---

class StrategyParser:
    """
    负责将 TemplateState 转换为符合《编辑策略规范 v1.1.0》的 JSON 数据
    """
    @staticmethod
    def serialize(state, current_layer_tree):
        # 1. 基础校验：必须有渲染方案
        if not state.render_presets:
            return None, [("system", "至少需要一个渲染方案")]

        # 2. 变量组标号映射 (文字层优先)
        group_mapping = {}
        next_group_idx = 1
        
        # 文字组标号
        for key in state.text_groups:
            if any(r['mapping_key'] == key for r in state.text_rules):
                group_mapping[key] = next_group_idx
                next_group_idx += 1
        
        # 图片组标号
        for key in state.image_groups:
            if any(r['mapping_key'] == key for r in state.image_rules):
                group_mapping[key] = next_group_idx
                next_group_idx += 1

        # 3. 构建 operations
        operations = []
        invalid_rules = [] # 记录失效的规则路径
        
        # 辅助函数：校验路径并提取 ID
        def resolve(path):
            return ps_server._resolve_layer_path(current_layer_tree, path)

        # 文本操作
        for r in state.text_rules:
            layer_id, parent_chain, _ = resolve(r['path'])
            if layer_id is None:
                invalid_rules.append(('text', r['path']))
                continue
            
            op = {
                "type": "update_text_layer",
                "layer_id": layer_id,
                "parent_chain": parent_chain,
                "target_path": r['path'],
                "group": group_mapping.get(r['mapping_key']),
                "regex_steps": r.get('regex_steps', [])
            }
            operations.append(op)
            
        # 图片操作
        for r in state.image_rules:
            layer_id, parent_chain, _ = resolve(r['path'])
            if layer_id is None:
                invalid_rules.append(('image', r['path']))
                continue
            
            op = {
                "type": "replace_image",
                "layer_id": layer_id,
                "parent_chain": parent_chain,
                "target_path": r['path'],
                "group": group_mapping.get(r['mapping_key'])
            }
            operations.append(op)
            
        # 滤镜操作
        for r in state.filter_rules:
            layer_id, parent_chain, kind = resolve(r['path'])
            if layer_id is None:
                invalid_rules.append(('filter', r['path']))
                continue
            
            # 安全性校验：局部滤镜仅支持智能对象
            if kind != "SMARTOBJECT":
                # 虽然 UI 层会限制只有 SO 才有滤镜按钮，但 serialize 阶段还是校验一下更稳
                continue
            
            for step in r.get('filter_steps', []):
                op = {
                    "type": "apply_filter",
                    "layer_id": layer_id,
                    "parent_chain": parent_chain,
                    "target_path": r['path'],
                    "filter_type": step['type'],
                    "params": step['params']
                }
                operations.append(op)

        # 4. 构建 renders (不再对 root_layers 进行阻塞式校验)
        renders = []
        filename_set = set()
        duplicate_filenames = []
        
        for p in state.render_presets:
            full_filename = f"{p['filename']}.{p['format']}"
            if full_filename in filename_set:
                duplicate_filenames.append(full_filename)
            else:
                filename_set.add(full_filename)

            # 解析 root_layers 路径为 root_ids
            root_ids = []
            for path in p.get('root_layers', []):
                rid, _, _ = resolve(path)
                if rid: root_ids.append(rid)

            render_config = {
                "name": p['name'],
                "description": p.get('description', ''),
                "output_path": os.path.abspath(p.get('output_path', './output')),
                "filename": p['filename'],
                "format": p['format'],
                "quality": p.get('quality', 100),
                "root_ids": root_ids, # 插件端需要数字 ID
                "tiling": p.get('tiling', {"enabled": False}),
                "filters": state.global_filter_steps if state.global_filter_active else []
            }
            renders.append(render_config)

        if duplicate_filenames:
            # 这里的 info 是给用户看的重复文件名列表
            return None, [("system", f"检测到重复的渲染输出路径: {', '.join(duplicate_filenames)}。请修改渲染方案的命名模板或格式。")]

        if invalid_rules:
            return None, invalid_rules

        return {
            "version": "1.1.0",
            "created": datetime.datetime.now().isoformat(),
            "modified": datetime.datetime.now().isoformat(),
            "operations": operations,
            "renders": renders
        }, None

    @staticmethod
    def _resolve_path_exists(tree, path_str):
        if not tree: return False
        # 统一格式化：去除两端空格
        parts = [p.strip() for p in path_str.split('>')]
        if not parts or parts[0] != '主文档':
            return False
        
        if len(parts) == 1: 
            return True
            
        current_level = tree
        for i, part in enumerate(parts[1:]):
            found_node = None
            # 严格匹配每一级名称
            for node in current_level:
                if node['name'].strip() == part:
                    found_node = node
                    break
            
            if found_node:
                # 如果已经是路径最后一级，匹配成功
                if i == len(parts) - 2:
                    return True
                # 否则继续向子节点查找
                current_level = found_node.get('children', [])
            else:
                return False
        return False

class StrategyLoader:
    """
    负责将 JSON 策略数据还原为 TemplateState
    """
    @staticmethod
    def deserialize(data, state, current_layer_tree):
        state.reset()
        
        ops = data.get('operations', [])
        text_ops = [op for op in ops if op['type'] == 'update_text_layer']
        img_ops = [op for op in ops if op['type'] == 'replace_image']
        filter_ops = [op for op in ops if op['type'] == 'apply_filter']
        
        # 1. 确定文字组分界
        all_text_groups = sorted(list(set(op.get('group') for op in text_ops if op.get('group'))))
        max_text_group = all_text_groups[-1] if all_text_groups else 0
        
        # 2. 还原文字规则
        for op in text_ops:
            path = op['target_path']
            group_idx = op.get('group', 1)
            group_name = f"文字组 {group_idx}"
            if group_name not in state.text_groups:
                while len(state.text_groups) < group_idx:
                    state.text_groups.append(f"文字组 {len(state.text_groups) + 1}")
            
            rule = {
                "name": path.split(' > ')[-1],
                "path": path,
                "mapping_key": group_name,
                "regex_steps": op.get('regex_steps', [])
            }
            state.text_rules.append(rule)

        # 3. 还原图片规则
        for op in img_ops:
            path = op['target_path']
            group_idx = op.get('group', 1)
            img_group_idx = max(1, group_idx - max_text_group)
            group_name = f"图片组 {img_group_idx}"
            
            if group_name not in state.image_groups:
                while len(state.image_groups) < img_group_idx:
                    state.image_groups.append(f"图片组 {len(state.image_groups) + 1}")

            rule = {
                "name": path.split(' > ')[-1],
                "path": path,
                "mapping_key": group_name
            }
            state.image_rules.append(rule)

        # 4. 还原滤镜规则
        path_to_filter_rule = {}
        for op in filter_ops:
            path = op['target_path']
            if path not in path_to_filter_rule:
                path_to_filter_rule[path] = {
                    "name": path.split(' > ')[-1],
                    "path": path,
                    "filter_steps": []
                }
            
            path_to_filter_rule[path]['filter_steps'].append({
                "type": op['filter_type'],
                "params": op['params']
            })
        state.filter_rules = list(path_to_filter_rule.values())

        # 5. 还原渲染方案
        if data.get('renders'):
            first_render = data['renders'][0]
            if first_render.get('filters'):
                state.global_filter_steps = first_render['filters']
                state.global_filter_active = True
            
            for r in data['renders']:
                preset = {
                    "name": r['name'],
                    "filename": r['filename'],
                    "format": r['format'],
                    "root_layers": r.get('root_layers', []),
                    "tiling": r.get('tiling', {"enabled": False}),
                    "output_path": r.get('output_path', './output'),
                    "quality": r.get('quality', 100)
                }
                state.render_presets.append(preset)
        
        # 6. 核心改进：确保文档加载后始终拥有至少一个渲染方案
        state.ensure_render_preset()

# --- 制作模板模式状态管理 ---

class TemplateState:
    """
    制作模板模式的全局状态
    """
    def __init__(self):
        self.current_doc = None
        self.layer_tree = []
        self.expanded_nodes = set() # 存储展开的节点 ID
        self.text_rules = [] # List[Dict] -> {id, name, path, mapping_key, regex_steps}
        self.image_rules = [] # List[Dict] -> {id, name, path, mapping_key}
        self.filter_rules = [] # List[Dict] -> {id, name, path, filter_steps}
        self.global_filter_steps = [] # 全局滤镜步骤
        self.global_filter_active = False # 全局滤镜容器是否激活显示
        self.render_presets = [] # 渲染方案列表
        self.open_docs = []
        # 变量组管理
        self.text_groups = ["文字组 1"]
        self.image_groups = ["图片组 1"]
        self.is_loading = False
        self.is_previewing = False
        self.system_filters = [] # 缓存从 system_filter.json 加载的配置
        self.load_system_filters()

    def load_system_filters(self):
        """加载滤镜系统配置"""
        try:
            config_path = os.path.join(os.path.dirname(__file__), 'config', 'system_filter.json')
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    self.system_filters = json.load(f)
        except Exception as e:
            print(f"加载系统滤镜配置失败: {e}")
            self.system_filters = []

    def reset(self):
        self.text_rules = []
        self.image_rules = []
        self.filter_rules = []
        self.global_filter_steps = []
        self.global_filter_active = False
        self.render_presets = []
        self.expanded_nodes = set()
        self.text_groups = ["文字组 1"]
        self.image_groups = ["图片组 1"]
        self.is_loading = False
        self.is_previewing = False

    def add_text_rule(self, layer):
        # 防止重复添加
        path = layer.get('path', '主文档')
        if any(r['path'] == path for r in self.text_rules):
            ui.notify(f"图层 {layer['name']} 已存在于规则中", type='warning')
            return False
        
        self.text_rules.append({
            "name": layer['name'],
            "path": path,
            "mapping_key": template_state.text_groups[0], # 默认选第一组
            "regex_steps": []
        })
        return True

    def add_image_rule(self, layer):
        path = layer.get('path', '主文档')
        if any(r['path'] == path for r in self.image_rules):
            ui.notify(f"图层 {layer['name']} 已存在于规则中", type='warning')
            return False
            
        self.image_rules.append({
            "name": layer['name'],
            "path": path,
            "mapping_key": template_state.image_groups[0] # 默认选第一组
        })
        return True

    def add_filter_rule(self, layer):
        path = layer.get('path', '主文档')
        if any(r['path'] == path for r in self.filter_rules):
            ui.notify(f"图层 {layer['name']} 已存在于滤镜规则中", type='warning')
            return False
            
        self.filter_rules.append({
            "name": layer['name'],
            "path": path,
            "filter_steps": []
        })
        return True

    def remove_rule(self, rule_type, path):
        if rule_type == 'text':
            self.text_rules = [r for r in self.text_rules if r['path'] != path]
        elif rule_type == 'image':
            self.image_rules = [r for r in self.image_rules if r['path'] != path]
        elif rule_type == 'filter':
            self.filter_rules = [r for r in self.filter_rules if r['path'] != path]

    def ensure_render_preset(self):
        """确保容器中至少有一个渲染方案，如果没有则根据预设自动添加一个"""
        if self.render_presets:
            return
            
        # 1. 加载本地持久化的默认设置
        defaults = local_config.data.get('settings', {}).get('render_defaults', {})
        
        # 2. 构造新方案 (逻辑与 add_render_preset 保持高度一致)
        new_preset = {
            "name": "默认方案",
            "description": "自动生成的默认方案",
            "output_path": "./output",
            "filename": defaults.get('filename_template', "{文字组 1}_{模板名}_{时间}"),
            "format": defaults.get('format', "jpg"),
            "quality": 85,
            "root_layers": [],
            "tiling": copy.deepcopy(defaults.get('tiling', {
                "enabled": False,
                "width": 1920,
                "height": 1080,
                "ppi": 300
            }))
        }
        
        # 3. 核心：根据图层树当前的可见性自动勾选根图层
        if self.layer_tree:
            visible_paths = [
                node.get('path', f"主文档 > {node['name']}") 
                for node in self.layer_tree 
                if node.get('visible', True)
            ]
            new_preset['root_layers'] = visible_paths
            
        self.render_presets.append(new_preset)
        return True

# 全局状态单例
template_state = TemplateState()

class FilterEditor:
    """
    通用滤镜参数编辑器弹窗
    """
    def __init__(self, filter_type_config, initial_params=None, on_save=None, container=None):
        self.config = filter_type_config # 包含 type, name, params 定义
        self.params = initial_params or self._get_default_params()
        self.on_save = on_save
        self.container = container
        self.dialog = None

    def _get_default_params(self):
        """从配置中获取默认参数字典"""
        return {p['internal_name']: p['default'] for p in self.config.get('params', [])}

    def open(self):
        context = self.container if self.container else ui.element('div').classes('hidden')
        with context:
            # 弹窗宽度锁定， persistent 确保点击外部不关闭
            with ui.dialog().classes('backdrop-blur-sm').props('persistent') as self.dialog, \
                 ui.card().classes('w-[480px] max-w-full p-0 gap-0 rounded-[20px] bg-white shadow-2xl overflow-hidden'):
                
                # Header
                with ui.row().classes('w-full justify-between items-center px-6 py-4 border-b border-slate-100 bg-slate-50/50'):
                    with ui.row().classes('items-center gap-3'):
                        with ui.element('div').classes('w-9 h-9 rounded-xl bg-purple-50 flex items-center justify-center text-purple-600'):
                            ui.icon('settings', size='20px')
                        with ui.column().classes('gap-0'):
                            ui.label(f"{self.config['name']} 设置").classes('text-sm font-bold text-slate-800 leading-tight')
                            ui.label(f"Filter: {self.config['type']}").classes('text-[10px] text-slate-400 font-medium')
                    ui.icon('close', size='20px').classes('cursor-pointer text-slate-300 hover:text-red-500 transition-colors').on('click', self.dialog.close)

                # Body (参数调节区)
                with ui.column().classes('w-full px-8 py-6 gap-6'):
                    for p_def in self.config.get('params', []):
                        with ui.column().classes('w-full gap-2'):
                            # 参数标题 + 当前值 + 单位
                            with ui.row().classes('w-full justify-between items-end'):
                                ui.label(p_def['display_name']).classes('text-xs font-bold text-slate-600')
                                with ui.row().classes('items-center gap-1'):
                                    # 数字输入框同步数值
                                    val_input = ui.number(value=self.params[p_def['internal_name']], precision=2) \
                                        .props('outlined dense hide-bottom-space').classes('w-16 bg-white text-xs')
                                    val_input.bind_value(self.params, p_def['internal_name'])
                                    ui.label(p_def['unit']).classes('text-[10px] text-slate-400 font-bold ml-0.5')

                            # 滑动条主控件
                            slider = ui.slider(min=p_def['min'], max=p_def['max'], step=0.1 if isinstance(p_def['min'], float) else 1) \
                                .props('label').classes('w-full')
                            slider.bind_value(self.params, p_def['internal_name'])
                            # 设置滑动条颜色以符合紫色滤镜主题
                            slider.props('color="purple-5" selection-color="purple-5"')

                # Footer
                with ui.row().classes('w-full justify-between items-center px-6 py-4 bg-slate-50/50 border-t border-slate-100 mt-auto'):
                    # 默认按钮
                    ui.button('恢复默认', on_click=self.reset_to_default).props('flat no-caps dense').classes('text-purple-600 font-bold text-xs hover:bg-purple-50 rounded-lg px-3')
                    
                    with ui.row().classes('items-center gap-3'):
                        ui.button('取消', on_click=self.dialog.close).props('flat no-caps dense').classes('text-slate-500 font-medium text-xs hover:bg-slate-100 rounded-lg px-4')
                        with ui.button(on_click=self.handle_save).props('unelevated no-caps dense').classes('bg-purple-500 hover:bg-purple-600 text-white rounded-lg px-5 py-1.5 shadow-md shadow-purple-200 transition-all'):
                            ui.icon('check', size='14px').classes('mr-1.5')
                            ui.label('保存参数').classes('text-xs font-bold')

            self.dialog.open()

    def reset_to_default(self):
        """恢复所有参数为默认值"""
        defaults = self._get_default_params()
        self.params.update(defaults)
        ui.notify(f"参数已重置为默认", type='info')

    def handle_save(self):
        if self.on_save:
            self.on_save(self.params)
        self.dialog.close()

def apply_regex_steps(text, steps):
    """
    统一的正则处理引擎
    支持 JS 风格的 $1 捕获组语法
    """
    import re
    result = text
    for step in steps:
        find_pattern = step.get('find', '')
        replace_pattern = step.get('replace', '')
        if not find_pattern:
            continue
        try:
            # 兼容性处理：将 JS 风格的 $1, $2 转换为 Python 风格的 \1, \2
            py_replace = re.sub(r'\$(\d+)', r'\\\1', replace_pattern)
            result = re.sub(find_pattern, py_replace, result)
        except Exception as e:
            print(f"正则应用失败: {e}")
    return result

class RegexEditor:
    """
    正则规则编辑器弹窗组件 (UI 修复版)
    """
    def __init__(self, initial_data=None, on_save=None, container=None):
        # 记录原始状态以便检测重名
        self._is_edit_mode = initial_data is not None
        self._original_name = initial_data.get('name') if initial_data else None
        
        # 必须使用 .copy() 避免在未保存时直接修改原始对象
        self.data = initial_data.copy() if initial_data else {
            "name": "",
            "find": "(.)(?=.)",
            "replace": "$1|",
            "test_input": "ice美化助手",
            "can_remove": True
        }
        self.on_save = on_save
        self.container = container
        self.dialog = None

    def open(self):
        # 确保在正确的容器上下文中打开
        context = self.container if self.container else ui.element('div').classes('hidden')
        
        with context:
            # 添加 .props('persistent') 确保点击外部不会关闭窗口
            with ui.dialog().classes('backdrop-blur-sm').props('persistent') as self.dialog, \
                 ui.card().classes('w-[640px] max-w-full p-0 gap-0 rounded-[20px] bg-white shadow-2xl'):
                
                # ---------------- HEADER ----------------
                with ui.row().classes('w-full justify-between items-center px-6 py-4 border-b border-slate-100'):
                    with ui.row().classes('items-center gap-3'):
                        # Logo / Icon 区域
                        with ui.element('div').classes('w-9 h-9 rounded-xl bg-cyan-50 flex items-center justify-center text-cyan-600'):
                            ui.icon('code', size='20px')
                        with ui.column().classes('gap-0'):
                            ui.label('添加正则规则').classes('text-sm font-bold text-slate-800 leading-tight')
                            ui.label('Regular Expression').classes('text-[10px] text-slate-400 font-medium')
                    
                    ui.icon('close', size='20px').classes('cursor-pointer text-slate-300 hover:text-red-500 transition-colors').on('click', self.dialog.close)
                
                # ---------------- BODY ----------------
                with ui.column().classes('w-full px-6 py-5 gap-5'):
                    
                    # 1. 规则名称
                    with ui.column().classes('w-full gap-1.5'):
                        ui.label('规则名称').classes('text-[11px] font-bold text-slate-600')
                        # hide-bottom-space 防止输入框底部留白撑开高度
                        ui.input(placeholder='例如：用|隔开每个字符').bind_value(self.data, 'name') \
                            .props('outlined dense hide-bottom-space').classes('w-full bg-slate-50 text-sm')

                    # 2. 查找与替换 (Grid 布局)
                    with ui.grid().classes('w-full grid-cols-2 gap-4'):
                        # 左：查找
                        with ui.column().classes('gap-1.5'):
                            ui.label('查找模式').classes('text-[11px] font-bold text-slate-600')
                            # 模拟代码输入框外观
                            with ui.row().classes('w-full items-center border border-slate-300 rounded-md bg-slate-50 px-2 h-[40px] focus-within:border-cyan-500 focus-within:ring-1 focus-within:ring-cyan-500 transition-all'):
                                ui.label('/').classes('text-slate-400 font-mono font-bold text-sm select-none mr-1')
                                # input-class="font-mono" 关键：设置等宽字体
                                self.find_input = ui.input().props('borderless dense hide-bottom-space input-class="font-mono text-cyan-700 text-sm"').classes('flex-grow')
                                self.find_input.bind_value(self.data, 'find')
                                self.find_input.on_value_change(self.update_test)
                                ui.label('/g').classes('text-violet-500 font-mono font-bold text-xs bg-violet-50 px-1.5 py-0.5 rounded select-none ml-1')
                        
                        # 右：替换
                        with ui.column().classes('gap-1.5'):
                            ui.label('替换为').classes('text-[11px] font-bold text-slate-600')
                            self.replace_input = ui.input(placeholder='留空则删除').bind_value(self.data, 'replace') \
                                .props('outlined dense hide-bottom-space input-class="font-mono text-sm"').classes('w-full bg-slate-50 h-[40px]')
                            self.replace_input.on_value_change(self.update_test)

                    # 3. 测试区域
                    with ui.column().classes('w-full gap-2'):
                        ui.label('测试区域').classes('text-[11px] font-bold text-slate-600')
                        
                        # 重新使用背景容器包裹整个测试区域
                        with ui.element('div').classes('w-full bg-slate-50/50 border border-slate-100 rounded-xl p-3'):
                            with ui.grid().classes('grid-cols-[1fr_auto_1fr] gap-3 items-center w-full'):
                                    
                                # 左侧：INPUT
                                with ui.column().classes('w-full gap-1'):
                                    ui.label('INPUT').classes('text-[9px] font-bold text-slate-400 tracking-wider')
                                    # 移除了 shadow-sm 以解决底部多余阴影容器的问题
                                    self.test_input_area = ui.textarea().bind_value(self.data, 'test_input') \
                                        .props('outlined dense hide-bottom-space bg-color="white" input-class="font-mono text-xs" input-style="resize: none;"') \
                                        .classes('w-full h-32 rounded-lg bg-white')
                                    self.test_input_area.on_value_change(self.update_test)
                                    
                                # 中间：箭头
                                with ui.element('div').classes('flex justify-center pt-4'):
                                    with ui.element('div').classes('w-8 h-8 rounded-full bg-cyan-50 flex items-center justify-center text-cyan-500'):
                                        ui.icon('arrow_forward', size='16px')
    
                                # 右侧：OUTPUT
                                with ui.column().classes('w-full gap-1'):
                                    ui.label('OUTPUT').classes('text-[9px] font-bold text-slate-400 tracking-wider')
                                    self.output_area = ui.textarea() \
                                        .props('outlined dense readonly hide-bottom-space bg-color="white" input-class="font-mono text-xs text-slate-600" input-style="resize: none;"') \
                                        .classes('w-full h-32 rounded-lg border-dashed bg-white')

                    # 4. AI 助手 (紫色主题)
                    with ui.row().classes('w-full bg-violet-50/50 border border-violet-100 rounded-xl p-3 items-center gap-3 mt-2'):
                        with ui.element('div').classes('w-9 h-9 rounded-lg bg-white border border-violet-100 flex items-center justify-center text-violet-500 shadow-sm shrink-0'):
                            ui.icon('auto_awesome', size='18px')
                        with ui.column().classes('gap-0.5 flex-grow'):
                            ui.label('不确定怎么写？').classes('text-xs font-bold text-violet-800')
                            ui.label('请在规则名称内填入你想实现的效果，然后点击按钮复制指令，将指令提交给豆包、DeepSeek等AI助手。').classes('text-[10px] text-violet-800')
                        
                        # 紫色按钮：确保图标和文字均为紫色
                        with ui.button(on_click=self.copy_ai_prompt).props('dense outline no-caps').classes('bg-white border-violet-300 hover:bg-violet-50 hover:border-violet-400 rounded-lg px-3'):
                            with ui.row().classes('items-center gap-1.5 text-violet-600'):
                                ui.icon('content_copy', size='14px')
                                ui.label('复制指令').classes('text-xs font-bold')

                # ---------------- FOOTER ----------------
                with ui.row().classes('w-full justify-end items-center px-6 py-4 bg-white border-t border-slate-100 gap-3 rounded-b-[20px]'):
                    ui.button('取消', on_click=self.dialog.close).props('flat no-caps dense').classes('text-slate-500 font-medium text-xs hover:bg-slate-100 rounded-lg px-4')
                    with ui.button(on_click=self.handle_save).props('unelevated no-caps dense').classes('bg-cyan-500 hover:bg-cyan-600 text-white rounded-lg px-5 py-1.5 shadow-md shadow-cyan-200 transition-all'):
                        ui.icon('check', size='14px').classes('mr-1.5')
                        ui.label('保存规则').classes('text-xs font-bold')

            self.dialog.open()
            self.update_test()
    def update_test(self):
        """实时正则匹配测试"""
        test_input = self.data.get('test_input', '')
        
        # 使用统一引擎进行测试
        # 为了实时预览单步效果，这里构造一个临时的 steps 列表
        current_step = {
            'find': self.data.get('find', ''),
            'replace': self.data.get('replace', '')
        }
        
        try:
            result = apply_regex_steps(test_input, [current_step])
            self.output_area.set_value(result)
        except Exception as e:
            self.output_area.set_value(f"正则错误: {str(e)}")

    def copy_ai_prompt(self):
        """生成并复制 AI Prompt"""
        prompt = prompt = f"""
# Role
你是一位精通正则表达式（Regex）的专家。

# Task Context
用户期望实现的效果："{self.data.get('name') or 'null'}"

# Rules & Logic (优先级最高)
1. **严格判别准则**：
   - **合法输入**：必须包含明确的“处理指令”或“逻辑动作”。例如：“删除数字”、“在...后面加横线”、“匹配手机号”、“将A替换为B”。
   - **非法输入**（必须拒绝）：
     * 纯名词/专有名词：如“热血兄弟连”、“苹果”、“张三”。
     * 无处理意图的短语：如“今天天气不错”、“你好”。
     * 无意义字符：如“asdfgh”、“null”。
   
2. **强制响应**：
   - 如果输入属于“非法输入”，你**禁止**尝试生成任何正则表达式，必须**原封不动**输出：
     请先在软件的规则名称内填入你想实现的效果，然后再复制指令。

3. **输出要求**（仅在输入合法时执行）：
   - 必须严格遵守下方指定的格式。

# Output Format
查找模式：[在此填入正则表达式]
替换为：[在此填入替换字符串]
输入示例：[提供一个合理的原始文本例子]
期望效果：[展示应用正则后的最终结果]
"""
        # 优先走 Windows 原生剪贴板（更稳定），再兜底走浏览器剪贴板
        try:
            if win32clipboard:
                win32clipboard.OpenClipboard()
                try:
                    win32clipboard.EmptyClipboard()
                    win32clipboard.SetClipboardText(prompt, win32clipboard.CF_UNICODETEXT)
                finally:
                    win32clipboard.CloseClipboard()
                ui.notify('求助指令已复制到剪贴板', type='positive')
                return
        except Exception:
            pass

        ui.run_javascript(f'navigator.clipboard.writeText({json.dumps(prompt)})')
        ui.notify('求助指令已复制到剪贴板', type='positive')

    def handle_save(self):
        """保存回调"""
        if not self.data['name']:
            ui.notify('请输入规则名称', type='warning')
            return
        if self.on_save:
            self.on_save(self.data)
        self.dialog.close()

class RenderPresetEditor:
    def __init__(self, initial_data, on_save, container, layer_tree, workbench=None):
        self.data = copy.deepcopy(initial_data)
        self.on_save = on_save
        self.container = container
        self.layer_tree = layer_tree
        self.workbench = workbench
        self.dialog = None

    def open(self):
        """打开编辑器弹窗"""
        with self.container:
            with ui.dialog().classes('backdrop-blur-sm').props('persistent') as self.dialog:
                with ui.card().classes('w-[760px] max-w-full p-0 gap-0 rounded-[20px] bg-white shadow-2xl overflow-hidden'):
                    # ---------------- HEADER ----------------
                    with ui.row().classes('w-full justify-between items-center px-6 py-4 border-b border-slate-100 bg-slate-50/50'):
                        with ui.row().classes('items-center gap-3'):
                            with ui.element('div').classes('w-9 h-9 rounded-xl bg-indigo-50 flex items-center justify-center text-indigo-600'):
                                ui.icon('settings_suggest', size='20px')
                            with ui.column().classes('gap-0'):
                                ui.label('配置渲染方案').classes('text-sm font-bold text-slate-800 leading-tight')
                                ui.label('Render Preset Configuration').classes('text-[10px] text-slate-400 font-medium')
                        ui.icon('close', size='20px').classes('cursor-pointer text-slate-300 hover:text-red-500 transition-colors').on('click', self.dialog.close)

                    # ---------------- BODY (左右分栏) ----------------
                    with ui.row().classes('w-full gap-0 flex-nowrap'):
                        # 左侧：图层选择器 (w-5/12) - 参考资产树设计
                        with ui.column().classes('w-5/12 p-6 border-r border-slate-100 gap-3 bg-slate-50/20'):
                            with ui.row().classes('w-full items-center justify-between mb-1'):
                                with ui.row().classes('items-center gap-2'):
                                    ui.icon('layers', size='16px', color='indigo-400')
                                    ui.label('选择输出图层').classes('text-xs font-bold text-slate-500 uppercase tracking-wider')
                                
                                # 全选/反选按钮组
                                with ui.row().classes('gap-1'):
                                    ui.button(icon='check_box', on_click=self.select_all_layers).props('flat dense size="sm"').classes('text-indigo-400 hover:text-indigo-600').tooltip('全选')
                                    ui.button(icon='indeterminate_check_box', on_click=self.deselect_all_layers).props('flat dense size="sm"').classes('text-slate-400 hover:text-slate-600').tooltip('全不选')
                            
                            with ui.scroll_area().classes('w-full h-[360px] rounded-xl border border-slate-200 bg-white shadow-inner').style('scrollbar-width: none; -ms-overflow-style: none;') as self.layers_scroll:
                                # 强制隐藏滚动条 (NiceGUI 推荐方式)
                                ui.add_head_html('<style>.q-scrollarea__thumb { display: none !important; }</style>')
                                
                                self.layers_list_container = ui.element('div').classes('p-1 w-full')
                                self._render_layers_list()

                        # 右侧：参数设置区 (w-7/12)
                        with ui.column().classes('w-7/12 p-6 gap-6'):
                            # 1. 方案名称模板
                            with ui.column().classes('w-full gap-2'):
                                with ui.row().classes('items-center gap-2'):
                                    ui.icon('edit_note', size='16px', color='indigo-400')
                                    ui.label('方案名称模板').classes('text-xs font-bold text-slate-500 uppercase tracking-wider')
                                
                                self.name_input = ui.input(placeholder='例如: {文字组 1}_{模板名}').bind_value(self.data, 'filename') \
                                    .props('outlined dense hide-bottom-space').classes('w-full bg-slate-50 text-sm font-mono rounded-lg')
                                self.name_input.on_value_change(self.update_name_preview)
                                
                                # 占位符指示器
                                with ui.element('div').classes('w-full p-3 bg-indigo-50/30 rounded-xl border border-indigo-100/50'):
                                    with ui.row().classes('flex-wrap gap-2 mb-2'):
                                        # 动态生成文字组占位符
                                        group_placeholders = [f'{{文字组 {i+1}}}' for i in range(len(template_state.text_groups))]
                                        fixed_placeholders = ['{模板名}', '{时间}']
                                        
                                        for p in group_placeholders + fixed_placeholders:
                                            ui.badge(p, color='indigo-100').classes('text-indigo-600 px-2 py-1 rounded-md cursor-pointer hover:bg-indigo-200 transition-colors') \
                                                .on('click', lambda p=p: self.add_placeholder(p))
                                    
                                    self.preview_label = ui.html(sanitize=False).classes('text-[10px] text-slate-400 leading-relaxed break-all font-mono italic')
                                    self.update_name_preview()

                            # 2. 格式与平铺 (Grid)
                            with ui.grid().classes('w-full grid-cols-2 gap-4'):
                                with ui.column().classes('gap-2'):
                                    ui.label('输出格式').classes('text-[11px] font-bold text-slate-500')
                                    # 自定义下拉菜单
                                    with ui.button().props('flat no-caps dense').classes('w-full h-[40px] px-3 bg-slate-50 rounded-xl border border-slate-200 flex justify-between items-center') as format_btn:
                                        with ui.row().classes('items-center gap-2'):
                                            ui.icon('image', size='16px', color='slate-400')
                                            self.format_label = ui.label(self.data.get('format', 'jpg').upper()).classes('text-xs font-bold text-slate-600')
                                        ui.icon('expand_more', size='14px', color='slate-300')
                                        
                                        with ui.menu().props('anchor="bottom left" self="top left"').classes('ice-dropdown-menu') as format_menu:
                                            if self.workbench:
                                                self.workbench._render_custom_menu(
                                                    format_menu,
                                                    ['psd', 'jpg', 'png', 'webp', 'bmp', 'gif'],
                                                    current_value=self.data.get('format', 'jpg'),
                                                    on_change=lambda v: (self.data.__setitem__('format', v), self.format_label.set_text(v.upper()))
                                                )
                                    format_btn.on('click', format_menu.open)

                                with ui.column().classes('gap-2'):
                                    ui.label('平铺画布').classes('text-[11px] font-bold text-slate-500')
                                    with ui.row().classes('w-full h-[40px] items-center px-3 bg-slate-50 rounded-xl border border-slate-200'):
                                        tiling_switch = ui.switch().bind_value(self.data['tiling'], 'enabled').props('color="indigo-5" size="sm"')
                                        ui.label('启用平铺').classes('text-[11px] text-slate-600 font-medium ml-1')

                            # 3. 平铺尺寸 (平滑动画)
                            with ui.element('div').classes('w-full transition-all duration-300 ease-in-out').style('overflow: hidden') as tiling_area:
                                def update_tiling_style(enabled):
                                    tiling_area.style(
                                        f'max-height: {"100px" if enabled else "0px"}; '
                                        f'opacity: {"1" if enabled else "0"}; '
                                        f'padding-top: {"12px" if enabled else "0px"}; '
                                        f'visibility: {"visible" if enabled else "hidden"};'
                                    )
                                
                                tiling_switch.on_value_change(lambda e: update_tiling_style(e.value))
                                # 初始状态
                                ui.timer(0.1, lambda: update_tiling_style(self.data['tiling']['enabled']), once=True)
                                
                                with ui.grid().classes('w-full grid-cols-2 gap-4'):
                                    with ui.column().classes('gap-1.5'):
                                        ui.label('宽度 (px)').classes('text-[10px] font-bold text-slate-400 uppercase')
                                        # 移除 props 中的 rounded，使用 classes('rounded-xl') 配合自定义样式解决胶囊形问题
                                        ui.number(value=self.data['tiling'].get('width', 1920), precision=0).classes('w-full bg-white shadow-sm ice-rounded-input').props('outlined dense') \
                                            .bind_value(self.data['tiling'], 'width')
                                    
                                    with ui.column().classes('gap-1.5'):
                                        ui.label('高度 (px)').classes('text-[10px] font-bold text-slate-400 uppercase')
                                        ui.number(value=self.data['tiling'].get('height', 1080), precision=0).classes('w-full bg-white shadow-sm ice-rounded-input').props('outlined dense') \
                                            .bind_value(self.data['tiling'], 'height')

                    # ---------------- FOOTER ----------------
                    with ui.row().classes('w-full justify-end items-center px-6 py-4 bg-slate-50/50 border-t border-slate-100 gap-3 mt-auto'):
                        ui.button('取消', on_click=self.dialog.close).props('flat no-caps dense').classes('text-slate-500 font-medium text-xs hover:bg-slate-100 rounded-lg px-4')
                        with ui.button(on_click=self.handle_save).props('unelevated no-caps dense').classes('bg-indigo-500 hover:bg-indigo-600 text-white rounded-lg px-6 py-2 shadow-md shadow-indigo-100 transition-all'):
                            ui.icon('check', size='14px').classes('mr-1.5')
                            ui.label('保存渲染方案').classes('text-xs font-bold')
            
            self.dialog.open()

    def _render_layers_list(self):
        """渲染图层列表项"""
        with self.layers_list_container:
            self.layers_list_container.clear()
            root_nodes = [node for node in self.layer_tree]
            if not root_nodes:
                ui.label('暂无根图层数据').classes('text-xs text-slate-400 p-4 italic text-center w-full')
                return

            for node in root_nodes:
                kind = node.get('kind')
                # 允许所有根节点显示，包括 PIXEL 图层
                
                name = node.get('name', '未命名')
                path = node.get('path', f"主文档 > {name}")
                
                with ui.row().classes('w-full items-center px-1 py-1 hover:bg-indigo-50/50 rounded-lg transition-colors group cursor-pointer') as row:
                    # 使用全路径进行勾选判断，以保持数据一致性
                    checked = path in self.data.get('root_layers', [])
                    cb = ui.checkbox('', value=checked).props('dense size="xs"').classes('scale-90 -ml-1')
                    cb.on_value_change(lambda e, p=path: self.toggle_root_layer(p, e.value))
                    
                    # 容器点击也能切换勾选
                    row.on('click', lambda c=cb: c.set_value(not c.value))
                    
                    # 图标 - 匹配资产树
                    if kind == 'TEXT':
                        ui.icon('text_fields', color='blue-400', size='14px')
                    elif kind == 'SMARTOBJECT':
                        ui.icon('filter_frames', color='purple-400', size='14px')
                    elif kind == 'GROUP':
                        ui.icon('folder', color='slate-400', size='14px')
                    else:
                        # 默认为像素图层图标
                        ui.icon('image', color='green-400', size='14px')
                    
                    ui.label(name).classes('text-[11px] text-slate-600 font-medium truncate ml-1 flex-grow')

    def select_all_layers(self):
        """全选所有根图层"""
        all_paths = [node.get('path', f"主文档 > {node['name']}") for node in self.layer_tree]
        self.data['root_layers'] = all_paths
        self._render_layers_list()
        ui.notify('已勾选所有根图层', type='info')

    def deselect_all_layers(self):
        """取消所有勾选"""
        self.data['root_layers'] = []
        self._render_layers_list()
        ui.notify('已清空所有勾选', type='info')

    def add_placeholder(self, p):
        current = self.name_input.value or ""
        self.name_input.value = current + p
        self.update_name_preview()

    def update_name_preview(self):
        text = self.name_input.value or ""
        
        # 1. 模板名处理：去除后缀
        doc_name = template_state.current_doc or '未定文档'
        if '.' in doc_name:
            doc_name = os.path.splitext(doc_name)[0]
            
        # 2. 实时高亮占位符
        preview_html = text
        
        # 获取当前所有合法的占位符
        placeholders = [f'{{文字组 {i+1}}}' for i in range(len(template_state.text_groups))] + ['{模板名}', '{时间}']
        
        # 高亮解析
        for p in placeholders:
            # 使用深蓝色高亮
            preview_html = preview_html.replace(p, f'<span class="text-indigo-600 font-bold bg-indigo-100/50 px-1 rounded mx-0.5 not-italic">{p}</span>')
        
        # 模拟解析预览
        parsed_preview = text
        # 匹配并替换所有文字组
        for i in range(len(template_state.text_groups)):
            parsed_preview = parsed_preview.replace(f'{{文字组 {i+1}}}', f'文本{i+1}')
            
        parsed_preview = parsed_preview.replace('{模板名}', doc_name)
        parsed_preview = parsed_preview.replace('{时间}', datetime.datetime.now().strftime('%m%d%H%M'))
        
        self.preview_label.set_content(f'预览: <span class="text-slate-400 font-mono">{parsed_preview}</span>')

    def toggle_root_layer(self, path, checked):
        if 'root_layers' not in self.data:
            self.data['root_layers'] = []
        if checked:
            if path not in self.data['root_layers']:
                self.data['root_layers'].append(path)
        else:
            if path in self.data['root_layers']:
                self.data['root_layers'].remove(path)

    def handle_save(self):
        filename = self.data.get('filename', '')
        if not filename:
            ui.notify('请填写方案名称模板', type='warning')
            return
            
        # 3. 校验文字组合法性
        import re
        groups_in_template = re.findall(r'\{文字组 (\d+)\}', filename)
        max_group_index = len(template_state.text_groups)
        
        for g_idx_str in groups_in_template:
            g_idx = int(g_idx_str)
            if g_idx < 1 or g_idx > max_group_index:
                ui.notify(f'方案中包含不存在的占位符: {{文字组 {g_idx}}} (当前仅支持到 1-{max_group_index})', type='error')
                return

        if not self.data.get('root_layers'):
            ui.notify('请至少选择一个输出图层', type='warning')
            return
            
        # --- 持久化用户偏好到 config.json ---
        settings = local_config.data.setdefault('settings', {})
        render_defaults = settings.setdefault('render_defaults', {})
        render_defaults['filename_template'] = filename
        render_defaults['format'] = self.data.get('format', 'jpg')
        render_defaults['tiling'] = copy.deepcopy(self.data.get('tiling', {"enabled": False}))
        local_config.save_to_disk()
        
        self.dialog.close()
        self.on_save(self.data)

# --- 极速出图侧滑面板组件 ---

class RapidExportPanel:
    """
    极速出图面板：提供剪贴板监听、批量处理 UI 和进度追踪
    """
    def __init__(self, parent):
        self.parent = parent
        self.container = None
        self.handle = None
        self.mapping_container = None
        self.collapsed = True
        
        # 任务队列引擎状态
        self.queue = []  # 等待执行的原始数据列表
        self.task_history = []  # 所有任务的完整记录（含状态和元数据）
        self.is_running = False
        self.strategy_snapshot = None
        self.processed_count = 0
        self.total_count = 0
        self.abort_requested = False
        
        # 持久化配置缓存
        settings = local_config.data.get('settings', {}).get('rapid_export', {})
        self.copy_after_render = settings.get('copy_after_render', False)
        self.export_path = settings.get('export_path', './output')
        self.ignore_header = settings.get('ignore_header', True)
        # 注意：剪贴板监听在“未登录/未连接”时不应启动，否则会在启动阶段误触发提示
        self.clipboard_monitor_active = False
        self._clipboard_monitor_saved = bool(settings.get('clipboard_monitor', False))
        self.last_clipboard_text = ""
        
        # UI 元素引用
        self.progress_label = None
        self.progress_bar = None
        self.render_list_container = None
        self.clipboard_switch = None
        self.ignore_header_switch = None
        self.copy_after_render_switch = None
        self.export_path_label = None
        self.manual_textarea = None
        self.overlay = None
        self.terminate_btn = None
        self.file_uploader = None
        
        self._build_ui()
        
        # 剪贴板监听定时器
        self.clipboard_timer = ui.timer(0.8, self._poll_clipboard, active=self.clipboard_monitor_active)
        # 初始化终止按钮状态
        self._update_terminate_btn_state()

    def on_logged_in(self):
        """登录成功后调用：恢复剪贴板监听开关（若用户之前开启过）。"""
        if self._clipboard_monitor_saved:
            self.clipboard_monitor_active = True
            try:
                self.clipboard_timer.active = True
            except Exception:
                pass
            if self.clipboard_switch:
                self.clipboard_switch.set_value(True)

    def _save_settings(self):
        """保存配置到磁盘"""
        settings = local_config.data.setdefault('settings', {})
        rapid_export = settings.setdefault('rapid_export', {})
        rapid_export['copy_after_render'] = self.copy_after_render
        rapid_export['export_path'] = self.export_path
        rapid_export['ignore_header'] = self.ignore_header
        rapid_export['clipboard_monitor'] = self.clipboard_monitor_active
        local_config.save_to_disk()

    def _build_ui(self):
        with self.parent:
            # 1. 全局任务锁定遮罩 (挂载在主卡片上，z-index 45)
            self.overlay = ui.element('div').classes('fixed inset-0 bg-transparent z-[45] hidden')
            
            # 2. 侧滑容器
            self.container = ui.element('div').classes('ice-rapid-panel ice-rapid-collapsed')
            with self.container:
                # 把手 handle
                self.handle = ui.element('div').classes('ice-rapid-handle').on('click', self.toggle)
                with self.handle:
                    self.handle_icon = ui.icon('chevron_left', size='20px')

                # 头部
                with ui.row().classes('w-full items-center justify-between mb-6'):
                    with ui.row().classes('items-center gap-2'):
                        ui.icon('bolt', color='amber-500').classes('text-xl')
                        ui.label('极速出图').classes('text-lg font-bold text-slate-800')
                        ui.label('Ready').classes('text-[10px] bg-cyan-50 text-cyan-600 px-1.5 py-0.5 rounded border border-cyan-100 font-bold')
                    ui.icon('close').classes('text-slate-300 cursor-pointer hover:text-slate-500').on('click', self.hide)

                with ui.scroll_area().classes('flex-grow w-full'):
                    with ui.column().classes('w-full gap-6 pr-2'):
                        # 1. 剪贴板监听配置
                        with ui.column().classes('w-full gap-3'):
                            with ui.row().classes('w-full items-center justify-between'):
                                with ui.row().classes('items-center gap-2'):
                                    ui.icon('content_paste').classes('text-sm text-slate-400')
                                    ui.label('剪贴板监听').classes('text-xs font-bold text-slate-600')
                                self.clipboard_switch = ui.switch(value=self.clipboard_monitor_active) \
                                    .props('dense').on_value_change(self._on_clipboard_switch_change)
                            
                            # 凹槽映射区
                            self.mapping_container = ui.element('div').classes('ice-green-groove w-full')
                            # 初始提示
                            with self.mapping_container:
                                ui.label('等待载入策略...').classes('text-[10px] text-slate-400 italic m-auto')

                        # 2. 手动粘贴区域
                        with ui.column().classes('w-full gap-2'):
                            with ui.row().classes('w-full items-center justify-between'):
                                ui.label('手动粘贴').classes('text-[10px] font-bold text-slate-400 uppercase tracking-widest')
                                ui.button('入队处理', on_click=self._process_manual_paste) \
                                    .props('flat dense').classes('text-[10px] text-primary lowercase tracking-tight')
                            self.manual_textarea = ui.textarea(placeholder='在此粘贴文本... 支持制表符、逗号、空格分隔，每行一条数据').props('outlined dense hide-bottom-space') \
                                .classes('w-full ice-rapid-textarea bg-white shadow-sm rounded-xl')

                        # 3. 批量解析
                        with ui.column().classes('w-full gap-2'):
                            with ui.row().classes('w-full items-center justify-between'):
                                ui.label('批量解析').classes('text-[10px] font-bold text-slate-400 uppercase tracking-widest')
                                with ui.row().classes('items-center gap-1'):
                                    ui.label('忽略表头').classes('text-[9px] text-slate-400')
                                    self.ignore_header_switch = ui.switch(value=self.ignore_header) \
                                        .props('dense size="xs"').on_value_change(self._on_ignore_header_change)
                            
                            self.file_uploader = ui.upload(
                                label='点击或拖拽表格文件',
                                multiple=False,
                                on_upload=self._handle_file_upload,
                                auto_upload=True
                            ).props('hide-upload-btn accept=".csv,.xlsx,.xlsm,text/csv,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"').classes('w-full ice-batch-uploader')
                            self.file_uploader.on('click', lambda _: self._open_table_file_dialog())
                            ui.label('支持 .csv / .xlsx').classes('ice-batch-uploader-note')

                        # 4. 导出位置
                        with ui.column().classes('w-full gap-2'):
                            ui.label('导出位置').classes('text-[10px] font-bold text-slate-400 uppercase tracking-widest')
                            with ui.row().classes('w-full items-start gap-2 bg-slate-50 rounded-xl p-2 border border-slate-100 overflow-hidden'):
                                # 路径区域：禁止横向溢出，长路径自动换行，占用剩余宽度
                                with ui.element('div').classes('flex-grow overflow-hidden min-w-0'):
                                    self.export_path_label = ui.label(self.export_path).classes('text-[10px] text-slate-500 px-1 font-mono break-all leading-relaxed')
                                ui.icon('folder_open').classes('text-sm text-slate-400 cursor-pointer hover:text-cyan-600 shrink-0 self-center').on('click', self._pick_export_path)

                        # 5. 渲染完成后动作
                        with ui.row().classes('w-full items-center justify-between'):
                            ui.label('渲染完复制到剪贴板').classes('text-xs font-bold text-slate-600')
                            self.copy_after_render_switch = ui.switch(value=self.copy_after_render) \
                                .props('dense').on_value_change(self._on_copy_after_render_change)

                # 6. 进度区域
                with ui.column().classes('w-full gap-4 pt-6 border-t border-slate-100'):
                    with ui.row().classes('w-full items-center justify-between'):
                        self.progress_label = ui.label('就绪').classes('text-xs font-bold text-slate-700')
                        with ui.row().classes('items-center gap-2'):
                            # 终止按钮
                            self.terminate_btn = ui.element('div').classes('ice-btn-terminate').on('click', self.terminate)
                            with self.terminate_btn:
                                ui.icon('stop', size='12px')
                                ui.label('终止')
                            self.percent_label = ui.label('0%').classes('text-xs font-bold text-cyan-600')
                    
                    # 进度条
                    with ui.element('div').classes('w-full h-1.5 bg-slate-100 rounded-full overflow-hidden'):
                        self.progress_bar = ui.element('div').classes('h-full bg-cyan-500 ice-shimmer-bar').style('width: 0%')

                    # 渲染列表
                    self.render_list_container = ui.column().classes('w-full gap-2 max-h-32 overflow-y-auto')

    def _on_clipboard_switch_change(self, e):
        # 未登录时不允许开启，避免启动阶段误触发提示
        if e.value and (not auth_logged_in or not ps_server.is_connected()):
            if self.clipboard_switch:
                self.clipboard_switch.set_value(False)
            ui.notify("请先登录并连接 Photoshop 后再开启剪贴板监听", type='warning')
            return

        self.clipboard_monitor_active = e.value
        self._clipboard_monitor_saved = bool(e.value)
        self.clipboard_timer.active = e.value
        self._save_settings()
        ui.notify(f"剪贴板监听已{'开启' if e.value else '关闭'}")

    def _on_ignore_header_change(self, e):
        self.ignore_header = e.value
        self._save_settings()

    def _on_copy_after_render_change(self, e):
        self.copy_after_render = e.value
        self._save_settings()

    def _extract_upload_filename(self, e):
        valid_exts = ('.csv', '.xlsx', '.xlsm', '.txt')
        generic_names = {'smallfileupload', 'file', 'blob', 'upload', 'object', 'bytesio', 'content'}

        def _pick_valid_name(value):
            if not value:
                return None
            text = str(value).strip().strip('"').strip("'")
            if not text:
                return None
            lowered = text.lower()
            if lowered in ('none', 'null', 'undefined', '未命名文件'):
                return None
            base = os.path.basename(text.replace('\\', '/'))
            if not base:
                return None
            stem = os.path.splitext(base)[0].strip().lower()
            if stem in generic_names:
                return None
            return base

        candidates = [
            getattr(e, 'name', None),
            getattr(e, 'filename', None),
            getattr(getattr(e, 'content', None), 'name', None),
            getattr(getattr(e, 'content', None), 'filename', None),
            getattr(getattr(getattr(e, 'content', None), 'file', None), 'name', None),
        ]
        names = getattr(e, 'names', None)
        if isinstance(names, str) and names.strip():
            candidates.append(names)
        elif isinstance(names, (list, tuple)) and names:
            candidates.extend(names)

        for c in candidates:
            chosen = _pick_valid_name(c)
            if chosen and chosen.lower().endswith(valid_exts):
                return chosen
        for c in candidates:
            chosen = _pick_valid_name(c)
            if chosen:
                return chosen

        # 兜底：从事件公开属性里找“看起来像文件名”的字符串
        for attr in dir(e):
            if attr.startswith('_'):
                continue
            try:
                val = getattr(e, attr)
            except Exception:
                continue
            if isinstance(val, str):
                chosen = _pick_valid_name(val)
                if chosen and chosen.lower().endswith(valid_exts):
                    return chosen
            elif isinstance(val, (list, tuple)):
                for item in val:
                    chosen = _pick_valid_name(item)
                    if chosen and chosen.lower().endswith(valid_exts):
                        return chosen
            elif hasattr(val, '__dict__'):
                for maybe in val.__dict__.values():
                    chosen = _pick_valid_name(maybe)
                    if chosen and chosen.lower().endswith(valid_exts):
                        return chosen
        return '导入文件'

    def _get_active_group_labels(self):
        labels = []
        used_text_keys = {r['mapping_key'] for r in template_state.text_rules}
        for key in template_state.text_groups:
            if key in used_text_keys:
                labels.append(str(key))
        used_img_keys = {r['mapping_key'] for r in template_state.image_rules}
        for key in template_state.image_groups:
            if key in used_img_keys:
                labels.append(str(key))
        return labels

    async def _extract_upload_bytes(self, e):
        content = getattr(e, 'content', None) or getattr(e, 'file', None) or getattr(e, 'data', None)
        if content is None:
            for attr in dir(e):
                if attr.startswith('_'):
                    continue
                try:
                    val = getattr(e, attr)
                except Exception:
                    continue
                if hasattr(val, 'read') or isinstance(val, (bytes, bytearray, memoryview)):
                    content = val
                    break
        if content is None:
            return None
        if hasattr(content, 'file') and hasattr(content.file, 'read'):
            content = content.file
        if isinstance(content, memoryview):
            return content.tobytes()
        if isinstance(content, (bytes, bytearray)):
            return bytes(content)
        if hasattr(content, 'read'):
            import inspect
            try:
                if hasattr(content, 'seek'):
                    content.seek(0)
            except Exception:
                pass
            data = content.read()
            if inspect.isawaitable(data):
                data = await data
            return data if isinstance(data, bytes) else (str(data).encode('utf-8') if data is not None else None)
        return str(content).encode('utf-8')

    def _decode_csv_bytes(self, file_bytes):
        for enc in ('utf-8-sig', 'utf-8', 'gb18030', 'gbk'):
            try:
                return file_bytes.decode(enc)
            except Exception:
                continue
        return file_bytes.decode('utf-8', errors='ignore')

    def _looks_like_xlsx(self, file_bytes):
        if not file_bytes or len(file_bytes) < 4 or file_bytes[:2] != b'PK':
            return False
        try:
            import zipfile
            with zipfile.ZipFile(BytesIO(file_bytes)) as zf:
                names = set(zf.namelist())
                return '[Content_Types].xml' in names and any(name.startswith('xl/') for name in names)
        except Exception:
            return False

    def _looks_like_csv(self, file_bytes):
        if not file_bytes:
            return False
        text = self._decode_csv_bytes(file_bytes[:16384])
        if not text:
            return False
        sample_lines = [ln for ln in text.splitlines() if ln.strip()][:5]
        if not sample_lines and text.strip():
            sample_lines = [text.strip()]
        if not sample_lines:
            return False
        delimiters = [',', '\t', ';', '|']
        for delim in delimiters:
            if any(delim in line for line in sample_lines):
                return True
        return False

    def _detect_table_format(self, filename, file_bytes):
        suffix = os.path.splitext((filename or '').lower())[1]
        if self._looks_like_xlsx(file_bytes):
            return 'xlsx'
        if suffix in ('.xlsx', '.xlsm'):
            return 'xlsx'
        if suffix in ('.csv', '.txt'):
            return 'csv'
        if self._looks_like_csv(file_bytes):
            return 'csv'
        return None

    def _open_table_file_dialog(self):
        if not self.file_uploader:
            return
        ui.run_javascript(f'''
            (() => {{
                const root = document.getElementById('c{self.file_uploader.id}');
                if (!root) return;
                const input = root.querySelector('input[type="file"]');
                if (!input) return;
                input.setAttribute('accept', '.csv,.xlsx,.xlsm,text/csv,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet');
                input.click();
            }})();
        ''')

    async def _handle_file_upload(self, e):
        """处理上传的表格文件：解析 -> 预览 -> 用户确认后入队。"""
        try:
            import io
            filename = self._extract_upload_filename(e)
            file_bytes = await self._extract_upload_bytes(e)
            if not file_bytes:
                ui.notify("读取上传文件失败，请重试拖拽或点击上传", type='negative')
                return

            file_type = self._detect_table_format(filename, file_bytes)
            parsed_rows = []
            if file_type is None:
                # 容错：部分系统上传事件无法给出可靠后缀，尝试按 CSV 直接解析。
                import csv
                try:
                    probe_stream = io.StringIO(self._decode_csv_bytes(file_bytes))
                    probe_reader = csv.reader(probe_stream)
                    probe_rows = [r for r in probe_reader if r]
                    if probe_rows:
                        file_type = 'csv'
                except Exception:
                    pass
            if file_type == 'csv':
                import csv
                text_stream = io.StringIO(self._decode_csv_bytes(file_bytes))
                reader = csv.reader(text_stream)
                for row in reader:
                    if row:
                        normalized = [str(cell).strip() if cell is not None else "" for cell in row]
                        if any(cell != "" for cell in normalized):
                            parsed_rows.append(normalized)
            elif file_type == 'xlsx':
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(filename=io.BytesIO(file_bytes))
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        if any(row):
                            str_row = [str(cell) if cell is not None else "" for cell in row]
                            parsed_rows.append(str_row)
                except ImportError:
                    ui.notify("未安装 openpyxl，无法解析 Excel。请使用 CSV 或运行 'pip install openpyxl'", type='warning')
                    return
            else:
                ui.notify("无法识别文件格式，请上传 CSV 或 XLSX", type='warning')
                return

            if not parsed_rows:
                ui.notify("文件为空", type='warning')
                return

            text_vars, img_vars = self._get_active_group_counts()
            total_vars = text_vars + img_vars
            group_labels = self._get_active_group_labels()
            if len(group_labels) < total_vars:
                group_labels.extend([f'变量 {i + 1}' for i in range(len(group_labels), total_vars)])
            if total_vars <= 0:
                ui.notify("当前策略没有分配可变量（文本/图片组），无法进行预览", type='warning')
                return

            def build_preview_candidates(ignore_header_value: bool, filter_empty_value: bool):
                start_idx = 1 if ignore_header_value and len(parsed_rows) > 1 else 0
                candidates = []
                invalid_count = 0
                for source_idx in range(start_idx, len(parsed_rows)):
                    row = parsed_rows[source_idx]
                    normalized = [str(cell).strip() if cell is not None else "" for cell in row]
                    cells = normalized[:total_vars]
                    if len(cells) < total_vars:
                        cells.extend([''] * (total_vars - len(cells)))
                    # 表格场景允许空值和列数不匹配（自动补齐/截断），仅按“整行为空”可选过滤
                    if filter_empty_value and not any(cells):
                        invalid_count += 1
                        continue
                    candidates.append({
                        'source_idx': source_idx,
                        'cells': cells,
                        'line': '\t'.join(cells),
                    })
                return candidates, invalid_count

            all_candidates, _ = build_preview_candidates(False, False)
            if not all_candidates:
                ui.notify("没有识别到可导入的数据，请检查文件内容", type='warning')
                return

            # 弹出预览对话框：列表选择、默认全选、手动勾选、一键入队或放弃
            async def show_preview_dialog():
                selected_rows = {}

                with ui.dialog().classes('backdrop-blur-sm').props('persistent') as dialog, \
                     ui.card().classes('w-[min(860px,92vw)] max-w-[92vw] p-0 gap-0 rounded-[28px] ice-card bg-white shadow-2xl overflow-hidden'):

                    with ui.row().classes('w-full justify-between items-center px-6 py-4 border-b border-slate-100 gap-3'):
                        with ui.row().classes('items-center gap-3'):
                            ui.icon('table_rows').classes('text-cyan-500 text-xl')
                            ui.label('解析预览').classes('text-lg font-bold text-slate-800')
                            ui.label(f'{filename}').classes('text-[11px] text-slate-400')
                        ui.icon('close', size='20px').classes('cursor-pointer text-slate-300 hover:text-red-500 transition-colors').on('click', dialog.close)

                    with ui.column().classes('w-full p-6 gap-4'):
                        with ui.row().classes('w-full items-center justify-between'):
                            with ui.row().classes('items-center gap-4'):
                                ui.label('跳过首行（通常是表头）').classes('text-xs font-bold text-slate-600')
                                preview_ignore_header = ui.switch(value=self.ignore_header).props('dense size="xs"')
                                ui.label('自动忽略空行').classes('text-xs font-bold text-slate-600')
                                preview_filter_empty = ui.switch(value=True).props('dense size="xs"')
                                status_label = ui.label('').classes('text-[10px] text-slate-400')

                            with ui.row().classes('items-center gap-2'):
                                def select_all():
                                    candidates, _ = build_preview_candidates(preview_ignore_header.value, preview_filter_empty.value)
                                    for item in candidates:
                                        selected_rows[item['source_idx']] = True
                                    render_list()
                                def deselect_all():
                                    candidates, _ = build_preview_candidates(preview_ignore_header.value, preview_filter_empty.value)
                                    for item in candidates:
                                        selected_rows[item['source_idx']] = False
                                    render_list()
                                ui.button('全选').props('flat dense no-caps').classes('text-xs text-cyan-600 hover:bg-cyan-50 rounded px-3').on('click', select_all)
                                ui.button('取消全选').props('flat dense no-caps').classes('text-xs text-slate-500 hover:bg-slate-100 rounded px-3').on('click', deselect_all)

                        grid_template = f'grid-template-columns: 64px 74px repeat({total_vars}, minmax(140px, 1fr));'
                        with ui.element('div').classes('w-full ice-preview-table-wrap'):
                            list_container = ui.scroll_area().classes('w-full max-h-[46vh] bg-white ice-preview-scroll')
                            with list_container:
                                list_inner = ui.column().classes('w-full gap-0')

                        def render_list():
                            list_inner.clear()
                            candidates, invalid_count = build_preview_candidates(preview_ignore_header.value, preview_filter_empty.value)
                            status_label.set_text(f'已识别 {len(candidates)} 条，忽略空行 {invalid_count} 条')

                            for item in candidates:
                                if item['source_idx'] not in selected_rows:
                                    selected_rows[item['source_idx']] = True

                            with list_inner:
                                with ui.row().classes('w-full ice-preview-grid-row ice-preview-table-header').style(grid_template):
                                    with ui.element('div').classes('ice-preview-cell ice-preview-cell-head ice-preview-col-check-head justify-center'):
                                        ui.label('选中')
                                    with ui.element('div').classes('ice-preview-cell ice-preview-cell-head ice-preview-col-serial-head justify-center'):
                                        ui.label('序列')
                                    for col_idx in range(total_vars):
                                        with ui.element('div').classes('ice-preview-cell ice-preview-cell-head ice-preview-col-var-head'):
                                            ui.label(group_labels[col_idx])

                                if not candidates:
                                    with ui.row().classes('w-full justify-center py-10 text-slate-400 text-xs'):
                                        ui.label('当前没有可预览的数据')
                                    return

                                for display_i, item in enumerate(candidates, start=1):
                                    source_idx = item['source_idx']
                                    cells = item['cells']
                                    with ui.row().classes('w-full ice-preview-grid-row ice-preview-table-row').style(grid_template):
                                        def on_check(e, oi=source_idx):
                                            selected_rows[oi] = e.value
                                        with ui.element('div').classes('ice-preview-cell ice-preview-col-check justify-center'):
                                            ui.checkbox(value=selected_rows.get(source_idx, True)).props('dense size="xs"').on_value_change(on_check)
                                        with ui.element('div').classes('ice-preview-cell ice-preview-col-serial justify-center'):
                                            ui.label(str(display_i)).classes('text-[10px] text-slate-500 font-semibold')
                                        for col_idx in range(total_vars):
                                            value = cells[col_idx] if col_idx < len(cells) else ''
                                            ui.element('div').classes('ice-preview-cell ice-preview-col-var').add_slot(
                                                'default',
                                                f'<span style="white-space: nowrap; overflow: hidden; text-overflow: ellipsis; width: 100%; display: inline-block;">{html.escape(str(value))}</span>'
                                            )

                        render_list()
                        preview_ignore_header.on_value_change(lambda _: render_list())
                        preview_filter_empty.on_value_change(lambda _: render_list())

                    with ui.row().classes('w-full justify-end items-center px-6 py-4 border-t border-slate-100 gap-3'):
                        ui.button('放弃').props('flat no-caps dense').classes('text-slate-500 font-medium text-xs hover:bg-slate-100 rounded-lg px-4').on('click', dialog.close)
                        async def confirm_and_add():
                            lines_to_add = []
                            candidates, _ = build_preview_candidates(preview_ignore_header.value, preview_filter_empty.value)
                            for item in candidates:
                                if selected_rows.get(item['source_idx'], False):
                                    lines_to_add.append(item['line'])
                            if not lines_to_add:
                                ui.notify("请至少选择一条数据", type='warning')
                                return
                            dialog.close()
                            self.add_to_queue(lines_to_add, source=f"文件 {filename}")
                        ui.button('添加到队列').props('unelevated no-caps dense').classes('bg-cyan-500 hover:bg-cyan-600 text-white rounded-lg px-6 py-2 shadow-md shadow-cyan-200 transition-all').on('click', confirm_and_add)

                await dialog

            await show_preview_dialog()

        except Exception as ex:
            ui.notify(f"解析文件失败: {ex}", type='negative')
        finally:
            if self.file_uploader:
                try:
                    self.file_uploader.reset()
                except Exception:
                    pass

    async def _pick_export_path(self):
        """选择导出路径"""
        # 优先使用系统原生对话框（tkinter），避免依赖 pywebview 注入 API
        try:
            import tkinter as tk
            from tkinter import filedialog

            def pick_dir_sync():
                root = tk.Tk()
                # 隐藏图标（创建透明1x1图标）
                try:
                    root.iconbitmap(default='')
                except Exception:
                    pass
                root.withdraw()
                root.attributes('-topmost', True)
                initial = os.path.abspath(self.export_path) if self.export_path else os.getcwd()
                path = filedialog.askdirectory(initialdir=initial, title="选择导出文件夹")
                try:
                    root.destroy()
                except Exception:
                    pass
                return path

            path = await asyncio.to_thread(pick_dir_sync)
            if path:
                self.export_path = path
                self.export_path_label.set_text(path)
                self._save_settings()
                ui.notify(f"导出路径已更新: {path}", type='positive')
                return
        except Exception:
            pass

        # 兜底：尝试通过 pywebview JS API
        try:
            path = await ui.run_javascript("window.pywebview.api.pick_folder()")
            if path:
                self.export_path = path
                self.export_path_label.set_text(path)
                self._save_settings()
                ui.notify(f"导出路径已更新: {path}", type='positive')
                return
        except Exception:
            ui.notify("当前环境不支持路径选择器，请手动在 config.json 中修改 rapid_export.export_path", type='warning')

    def _pick_table_file(self):
        """打开文件选择器解析表格"""
        ui.notify("正在打开文件选择器...")
        # 后期对接 pywebview 的文件选择接口

    def _process_manual_paste(self):
        text = self.manual_textarea.value
        if not text.strip():
            ui.notify("请输入要解析的文本", type='warning')
            return
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        self.add_to_queue(lines)
        self.manual_textarea.set_value("")

    def _poll_clipboard(self):
        """轮询剪贴板"""
        # 未登录/未连接时不嗅探，避免启动阶段误提示
        if not auth_logged_in or not ps_server.is_connected():
            return
        if not self.clipboard_monitor_active or self.is_running or not win32clipboard:
            return
        
        try:
            win32clipboard.OpenClipboard()
            # 仅在包含文本格式时尝试嗅探，防止与图片复制冲突
            if win32clipboard.IsClipboardFormatAvailable(win32clipboard.CF_UNICODETEXT):
                data = win32clipboard.GetClipboardData(win32clipboard.CF_UNICODETEXT)
                # 严格对比去除首尾空格后的内容，防止重复入队
                clean_data = data.strip() if data else ""
                if clean_data and clean_data != self.last_clipboard_text:
                    self.last_clipboard_text = clean_data
                    lines = [line.strip() for line in clean_data.split('\n') if line.strip()]
                    if lines:
                        self.add_to_queue(lines, source="剪贴板嗅探")
            win32clipboard.CloseClipboard()
        except Exception:
            pass

    def _get_active_group_counts(self):
        """计算策略中实际使用的变量组数量 (与 StrategyParser 逻辑保持一致)"""
        text_vars = 0
        img_vars = 0
        
        # 统计文字组 (按 template_state.text_groups 顺序，且必须被规则使用)
        used_text_keys = {r['mapping_key'] for r in template_state.text_rules}
        for key in template_state.text_groups:
            if key in used_text_keys:
                text_vars += 1
                
        # 统计图片组 (同理)
        used_img_keys = {r['mapping_key'] for r in template_state.image_rules}
        for key in template_state.image_groups:
            if key in used_img_keys:
                img_vars += 1
                
        return text_vars, img_vars

    def _split_input_line(self, line: str, total_vars: int):
        text = (line or '').strip()
        if not text:
            return []
        # 严格模式：仅支持英文逗号或空格分隔
        if ',' in text:
            parts = [p.strip() for p in text.split(',')]
        else:
            parts = [p.strip() for p in re.split(r'\s+', text)]
        return [p for p in parts if p != ""]

    def add_to_queue(self, data_lines, source="手动"):
        """将数据行解析并加入队列，同时记录任务历史"""
        if not template_state.current_doc:
            ui.notify("请先打开 PSD 文档并载入策略", type='warning')
            return

        # 获取当前实际使用的变量总数
        text_vars, img_vars = self._get_active_group_counts()
        total_vars = text_vars + img_vars
        
        if total_vars == 0:
            ui.notify("当前策略没有分配可变量（文本/图片组），无法执行快速出图", type='warning')
            return

        # 防误触：单参数场景下，剪贴板输入永不自动入队
        if total_vars == 1 and source.startswith("剪贴板"):
            return

        valid_tasks = []
        invalid_rows = []
        non_empty_lines = [ln for ln in data_lines if str(ln).strip()]
        for idx, line in enumerate(non_empty_lines, start=1):
            if source.startswith("文件 "):
                file_parts = [p.strip() for p in str(line).split('\t')]
                if len(file_parts) < total_vars:
                    file_parts.extend([''] * (total_vars - len(file_parts)))
                elif len(file_parts) > total_vars:
                    file_parts = file_parts[:total_vars]
                valid_tasks.append(file_parts)
                continue
            parts = self._split_input_line(str(line), total_vars)
            if len(parts) != total_vars:
                invalid_rows.append(idx)
                continue
            valid_tasks.append(parts)

        if not valid_tasks:
            ui.notify(f"未检测到符合变量数 ({total_vars}列) 的合法数据", type='warning')
            return

        # 批量严格一致性：任意一行不合法则整批不入队，避免任务半成功半失败
        if invalid_rows:
            ui.notify(
                f"存在 {len(invalid_rows)} 行数据参数量不匹配（必须严格为 {total_vars} 列），本批次未入队",
                type='warning'
            )
            return

        # 如果队列为空且当前没在运行，则准备启动
        first_time = (not self.queue and not self.is_running)

        # 获取当前策略的渲染预设文件名模板，用于生成预期输出文件名
        filename_template = "output_{index}"
        if self.strategy_snapshot and self.strategy_snapshot.get('renders', []):
            filename_template = self.strategy_snapshot['renders'][0].get('filename', "output_{index}")
        
        start_idx = len(self.task_history)
        for i, task_data in enumerate(valid_tasks):
            # 为每个任务生成可读的“文件名”/标识（用前两个字段拼接）
            display_name = ' '.join(str(x) for x in task_data[:2])
            # 用模板生成预期输出文件名
            output_name = filename_template
            for j, val in enumerate(task_data):
                output_name = output_name.replace(f'{{文字组 {j+1}}}', str(val))
            doc_name = os.path.splitext(template_state.current_doc or "template")[0]
            output_name = output_name.replace('{模板名}', doc_name)
            output_name = output_name.replace('{时间}', datetime.datetime.now().strftime('%m%d%H%M'))
            output_name = re.sub(r'[\/:*?"<>|]', '_', output_name)
            self.task_history.append({
                'index': start_idx + i + 1,
                'data': task_data,
                'display_name': display_name,
                'output_name': output_name,
                'status': 'waiting',  # waiting / running / success / failed
                'error': None
            })

        self.queue.extend(valid_tasks)
        self.total_count = len(self.queue) + self.processed_count
        ui.notify(f"已从{source}添加 {len(valid_tasks)} 条任务到队列", type='positive')
        self._update_terminate_btn_state()
        self._update_render_list_ui()
        
        if first_time:
            asyncio.create_task(self.process_queue())

    async def process_queue(self):
        """处理任务队列的核心循环"""
        if self.is_running or not self.queue:
            return
            
        self.is_running = True
        self.abort_requested = False
        self.overlay.classes(remove='hidden') # 锁定 UI
        self._update_terminate_btn_state()
        
        # 1. 获取当前策略快照，作为本次运行的唯一基准
        self.strategy_snapshot, _ = StrategyParser.serialize(template_state, template_state.layer_tree)
        if not self.strategy_snapshot:
            ui.notify("序列化策略失败，中止运行", type='negative')
            self._cleanup_after_run()
            return

        self.processed_count = 0
        self.total_count = len(self.queue) + self.processed_count
        
        try:
            while self.queue and not self.abort_requested:
                # 找到当前要处理的任务在 task_history 中的索引
                current_task_data = self.queue[0]
                current_task_idx = None
                for idx, t in enumerate(self.task_history):
                    if t['data'] == current_task_data and t['status'] == 'waiting':
                        current_task_idx = idx
                        break
                if current_task_idx is not None:
                    self.task_history[current_task_idx]['status'] = 'running'
                
                # 取出任务
                task_data = self.queue.pop(0)
                self.processed_count += 1
                
                # 更新 UI 进度
                percent = int((self.processed_count / self.total_count) * 100) if self.total_count > 0 else 0
                self.progress_label.set_text(f"正在处理 {self.processed_count} / {self.total_count}")
                self.percent_label.set_text(f"{percent}%")
                self.progress_bar.style(f"width: {percent}%")
                
                # 更新列表 UI
                self._update_render_list_ui()
                
                # 2. 构造操作包
                operations = self._prepare_operations(task_data)
                
                # 3. 构造渲染包 (使用持久化的导出路径)
                timestamp = datetime.datetime.now().strftime('%H%M%S')
                filename = f"rapid_{self.processed_count:03d}_{timestamp}"
                
                renders = []
                for r_preset in self.strategy_snapshot.get('renders', []):
                    # 克隆渲染预设并覆盖必要字段
                    render_item = copy.deepcopy(r_preset)
                    render_item['folder'] = os.path.abspath(self.export_path)
                    # 简单文件名替换逻辑 (占位符支持)
                    final_name = self._parse_filename(render_item['filename'], task_data)
                    render_item['file_name'] = final_name
                    renders.append(render_item)

                # 4. 执行原子化请求
                fut = asyncio.get_event_loop().create_future()
                def on_done(success, err):
                    if not fut.done(): fut.set_result((success, err))
                
                await ps_server.execute_strategy_atomic(
                    operations=operations,
                    renders=renders,
                    debug=False,
                    target_document=template_state.current_doc,
                    callback=on_done
                )
                
                success, err = await fut
                rendered_files = success.get('rendered_files', []) if isinstance(success, dict) else []
                file_errors = [r.get('error') for r in rendered_files if r.get('status') != 'success']
                atomic_failed = bool(err) or (not success) or bool(file_errors)
                
                # 更新 task_history 中的任务状态
                if current_task_idx is not None:
                    if atomic_failed:
                        self.task_history[current_task_idx]['status'] = 'failed'
                        first_file_error = next((str(e) for e in file_errors if e), None)
                        self.task_history[current_task_idx]['error'] = str(err) if err else (first_file_error or "渲染失败")
                    else:
                        self.task_history[current_task_idx]['status'] = 'success'
                        first_success_name = next((r.get('name') for r in rendered_files if r.get('status') == 'success' and r.get('name')), None)
                        if first_success_name:
                            self.task_history[current_task_idx]['output_name'] = str(first_success_name)
                
                # 更新列表 UI 状态
                self._update_render_list_ui()
                
                # 如果开启了复制到剪贴板，寻找第一个成功的渲染文件并复制其内容
                if (not atomic_failed) and self.copy_after_render:
                    for r_res in rendered_files:
                        if r_res.get('status') == 'success':
                            file_path = os.path.abspath(os.path.join(self.export_path, r_res['name']))
                            self._copy_image_to_clipboard(file_path)
                            break

        except Exception as e:
            ui.notify(f"任务循环异常: {e}", type='negative')
            # 如果有正在运行的任务，标记为失败
            for t in reversed(self.task_history):
                if t['status'] == 'running':
                    t['status'] = 'failed'
                    t['error'] = str(e)
                    break
            self._update_render_list_ui()
        finally:
            self._cleanup_after_run()

    def _prepare_operations(self, task_data):
        """根据快照和输入数据构造操作列表"""
        ops = []
        # 分离文字数据和图片数据 (使用统一逻辑获取活跃组计数)
        text_vars, img_vars = self._get_active_group_counts()
        text_data = task_data[:text_vars]
        img_data = task_data[text_vars:]
        
        # 映射
        for op in self.strategy_snapshot.get('operations', []):
            new_op = copy.deepcopy(op)
            g_idx = new_op.get('group', 1)
            if new_op['type'] == 'update_text_layer':
                if 1 <= g_idx <= len(text_data):
                    val = text_data[g_idx - 1]
                    new_op['text'] = apply_regex_steps(val, new_op.get('regex_steps', []))
            elif new_op['type'] == 'replace_image':
                if 1 <= g_idx <= len(img_data):
                    new_op['image_path'] = os.path.abspath(img_data[g_idx - 1])
            ops.append(new_op)
        return ops

    def _parse_filename(self, template, task_data):
        """解析文件名模板"""
        # 简单替换 {文字组 X}
        res = template
        for i, val in enumerate(task_data):
            res = res.replace(f'{{文字组 {i+1}}}', str(val))
        
        doc_name = os.path.splitext(template_state.current_doc or "template")[0]
        res = res.replace('{模板名}', doc_name)
        res = res.replace('{时间}', datetime.datetime.now().strftime('%m%d%H%M'))
        # 移除非法字符
        res = re.sub(r'[\/:*?"<>|]', '_', res)
        return res

    def _copy_image_to_clipboard(self, file_path):
        """将生成的图片文件内容复制到剪贴板 (仅限 Windows)"""
        if not win32clipboard:
            return

        try:
            import time
            # 等待文件写入完成，避免复制到半截文件
            for _ in range(5):
                if os.path.exists(file_path):
                    try:
                        # 尝试用只读打开确认文件可用
                        with open(file_path, 'rb') as f:
                            f.read(1)
                        break
                    except Exception:
                        pass
                time.sleep(0.1)

            if not Image or not os.path.exists(file_path):
                return

            # 使用 PIL 读取并转换为 Windows 剪贴板识别的 DIB 格式
            dib_data = None
            with Image.open(file_path) as img:
                output = BytesIO()
                img.convert("RGB").save(output, "BMP")
                dib_data = output.getvalue()[14:]  # DIB = BMP 去掉 14 字节文件头
                output.close()

            if not dib_data:
                return

            # Windows剪贴板经常被占用，重试3次
            for attempt in range(3):
                try:
                    win32clipboard.OpenClipboard()
                    try:
                        win32clipboard.EmptyClipboard()
                        win32clipboard.SetClipboardData(win32clipboard.CF_DIB, dib_data)
                        ui.notify("已将渲染结果复制到剪贴板", type='positive')
                        break
                    finally:
                        win32clipboard.CloseClipboard()
                except Exception as e:
                    if attempt < 2:
                        time.sleep(0.1)
                    else:
                        print(f"复制图片到剪贴板失败（已重试3次）: {e}")

        except Exception as e:
            print(f"复制图片到剪贴板失败: {e}")

    def _copy_to_clipboard(self, text):
        if not win32clipboard: return
        try:
            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardText(text, win32clipboard.CF_UNICODETEXT)
            win32clipboard.CloseClipboard()
        except Exception:
            pass

    def _cleanup_after_run(self):
        self.is_running = False
        self.overlay.classes(add='hidden')
        self._update_terminate_btn_state()
        # 任务结束逻辑：如果还有队列说明是被中止的，否则是自然结束
        if not self.queue:
            # 全部完成：UI 延迟回到就绪状态
            async def reset_ui():
                await asyncio.sleep(2.0)
                if not self.is_running and not self.queue:
                    self.progress_label.set_text("就绪")
                    self.percent_label.set_text("0%")
                    self.progress_bar.style("width: 0%")
                    # 只清空列表，保留状态显示
                    # self.render_list_container.clear() 
            asyncio.create_task(reset_ui())
        else:
            self.progress_label.set_text(f"已中止 (剩余 {len(self.queue)} 条)")

    def _update_render_list_ui(self, current_task_info=None):
        """保持渲染列表只显示最近三条：1条已完成, 1条进行中, 1条等待中"""
        self.render_list_container.clear()
        with self.render_list_container:
            # 1. 从 task_history 中找出上一张、正在执行、下一个的任务
            last_completed_task = None
            running_task = None
            next_waiting_task = None

            # 从后往前找最近一个已完成的
            for t in reversed(self.task_history):
                if t['status'] == 'success' or t['status'] == 'failed':
                    last_completed_task = t
                    break

            # 找正在执行的
            for t in self.task_history:
                if t['status'] == 'running':
                    running_task = t
                    break

            # 找第一个等待的
            for t in self.task_history:
                if t['status'] == 'waiting':
                    next_waiting_task = t
                    break

            # --- 第一行：上一张（已完成/失败）---
            if last_completed_task:
                status = last_completed_task['status']
                if status == 'success':
                    icon_html = '''<svg width="14" height="14" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg"><path d="M5 13l4 4L19 7" stroke="#10b981" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"/></svg>'''
                    text_color = 'text-emerald-600'
                    row_classes = 'w-full items-center gap-2 opacity-70'
                else:
                    icon_html = '''<svg width="14" height="14" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg"><path d="M6 18L18 6M6 6l12 12" stroke="#ef4444" stroke-width="3" stroke-linecap="round"/></svg>'''
                    text_color = 'text-red-500'
                    row_classes = 'w-full items-center gap-2 opacity-70'
                with ui.row().classes(row_classes):
                    ui.html(icon_html, sanitize=False)
                    display_name = last_completed_task.get('output_name', last_completed_task['display_name'])
                    ui.label(f"[{last_completed_task['index']:02d}] {display_name}").classes(f'text-[10px] {text_color} truncate flex-grow')
            else:
                # 没有上一张时留空占位
                with ui.row().classes('w-full items-center gap-2 opacity-0'):
                    ui.icon('schedule', color='slate-300').classes('text-[14px]')
                    ui.label('').classes('text-[10px] text-slate-400 truncate flex-grow')
            
            # --- 第二行：正在执行 ---
            if running_task:
                icon_html = '''<svg class="status-spin" width="14" height="14" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg"><path d="M12 22c5.523 0 10-4.477 10-10S17.523 2 12 2" stroke="#06b6d4" stroke-width="3" stroke-linecap="round"/></svg>'''
                text_color = 'text-cyan-600'
                with ui.row().classes('w-full items-center gap-2'):
                    ui.html(icon_html, sanitize=False)
                    display_name = running_task.get('output_name', running_task['display_name'])
                    ui.label(f"[{running_task['index']:02d}] {display_name}").classes(f'text-[10px] {text_color} truncate flex-grow')
            elif not self.is_running and not self.queue and last_completed_task:
                # 全部完成的情况：第二行显示“全部完成”
                with ui.row().classes('w-full items-center gap-2'):
                    ui.icon('check_circle', color='emerald-500').classes('text-[14px]')
                    ui.label(f'全部完成').classes('text-[10px] text-emerald-600 truncate flex-grow')
            else:
                # 没有正在执行时留空占位
                with ui.row().classes('w-full items-center gap-2 opacity-0'):
                    ui.icon('schedule', color='slate-300').classes('text-[14px]')
                    ui.label('').classes('text-[10px] text-slate-400 truncate flex-grow')

            # --- 第三行：下一个（等待中）---
            if next_waiting_task:
                icon_html = '''<svg width="14" height="14" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg"><path d="M12 8v4l3 3m6-3a9 9 0 11-18 0 9 9 0 0118 0z" stroke="#94a3b8" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/></svg>'''
                text_color = 'text-slate-400'
                with ui.row().classes('w-full items-center gap-2 opacity-50'):
                    ui.html(icon_html, sanitize=False)
                    display_name = next_waiting_task.get('output_name', next_waiting_task['display_name'])
                    ui.label(f"[{next_waiting_task['index']:02d}] {display_name}").classes(f'text-[10px] {text_color} truncate flex-grow')
            else:
                # 没有下一个时留空占位
                with ui.row().classes('w-full items-center gap-2 opacity-0'):
                    ui.icon('schedule', color='slate-300').classes('text-[14px]')
                    ui.label('').classes('text-[10px] text-slate-400 truncate flex-grow')

    def _update_terminate_btn_state(self):
        """终止按钮仅在队列仍有剩余任务时可用。"""
        if not self.terminate_btn:
            return
        enabled = bool(self.is_running and len(self.queue) > 0)
        if enabled:
            self.terminate_btn.classes(remove='opacity-40 pointer-events-none')
        else:
            self.terminate_btn.classes(add='opacity-40 pointer-events-none')

    def terminate(self):
        """终止当前任务队列（仅清除后续排队，不强杀当前正在渲染的任务）"""
        # 只有“还有排队任务”时才允许点击
        if not self.is_running or len(self.queue) == 0:
            return
        
        async def do_terminate():
            confirmed = await show_confirm_dialog(
                title='终止任务？',
                message='将清除后续排队任务，当前正在渲染的任务会尝试完成。确定继续吗？',
                confirm_text='清除排队',
                cancel_text='继续渲染',
                icon_type='warning'
            )
            if confirmed:
                self.queue.clear()
                # 调整总数为“已处理到当前”为止，让进度更合理
                if self.processed_count > 0:
                    self.total_count = self.processed_count
                    self.percent_label.set_text("100%")
                    self.progress_bar.style("width: 100%")
                self.progress_label.set_text("已清除后续排队任务")
                self._update_render_list_ui()
                self._update_terminate_btn_state()
                ui.notify("已清除后续排队任务", type='info')
        
        asyncio.create_task(do_terminate())

    def toggle(self):
        if self.is_running:
            ui.notify("任务正在运行，请先终止或等待完成再收起面板", type='warning')
            return
            
        if self.collapsed:
            self.show()
        else:
            self.hide()

    def show(self):
        self.collapsed = False
        self.container.classes(remove='ice-rapid-collapsed')
        self.handle_icon.set_name('chevron_right')
        # 每次拉出时刷新提示
        if not self.queue and not self.is_running:
            self.progress_label.set_text("就绪")
            self.percent_label.set_text("0%")
            self.progress_bar.style("width: 0%")

    def hide(self):
        if self.is_running:
            ui.notify("任务正在运行，收起面板将无法停止任务，建议先终止", type='warning')
            # 不强制拦截，但提示
        
        self.collapsed = True
        self.container.classes(add='ice-rapid-collapsed')
        self.handle_icon.set_name('chevron_left')
        # 退出面板时清空队列
        if not self.is_running:
            self.queue.clear()
            self.render_list_container.clear()

    def update_mapping(self, strategy_data=None):
        """根据策略刷新胶囊映射 UI"""
        if not self.mapping_container:
            return
            
        self.mapping_container.clear()
        
        # 获取文字组和图片组 (优先从传入的 strategy_data 提取)
        text_count = 0
        image_count = 0
        
        if strategy_data and isinstance(strategy_data, dict):
            # 策略库中的 operations
            ops = strategy_data.get('operations', [])
            text_groups = set()
            image_groups = set()
            for op in ops:
                if op.get('type') == 'update_text_layer':
                    g = op.get('group')
                    if g: text_groups.add(g)
                elif op.get('type') == 'replace_image':
                    g = op.get('group')
                    if g: image_groups.add(g)
            text_count = len(text_groups)
            image_count = len(image_groups)
        else:
            # 回退到 template_state，使用统一逻辑计算活跃组
            text_count, image_count = self._get_active_group_counts()
        
        if text_count == 0 and image_count == 0:
            with self.mapping_container:
                ui.label('策略中没有可变量').classes('text-[10px] text-slate-400 italic m-auto')
            return

        with self.mapping_container:
            # 渲染文字胶囊
            for i in range(1, text_count + 1):
                with ui.element('div').classes('ice-capsule ice-capsule-text'):
                    ui.label(f'文本{i}')
                
                # 如果后面还有元素，加个空格符号
                if i < text_count or image_count > 0:
                    with ui.element('div').classes('ice-rapid-space'):
                        ui.label('␣')

            # 渲染图片胶囊
            for i in range(1, image_count + 1):
                with ui.element('div').classes('ice-capsule ice-capsule-image'):
                    ui.label(f'图片路径{i}')
                
                # 最后一个不要空格
                if i < image_count:
                    with ui.element('div').classes('ice-rapid-space'):
                        ui.label('␣')

# --- 制作模板工作台组件 ---

class TemplateWorkbench:
    def __init__(self, parent_container):
        self.container = parent_container
        self.doc_selector_label = None
        self.assets_tree_container = None
        self.strategy_cards_container = None
        self.active_doc_name = template_state.current_doc or "选择文档"
        self.search_filter = ""

    def render(self):
        """渲染完整制作模板界面"""
        with self.container:
            # 1. 顶栏
            with ui.element('div').classes('ice-top-bar'):
                with ui.element('div').classes('ice-doc-selector').on('click', self.sync_open_docs):
                    ui.icon('image', color='primary').classes('text-lg')
                    self.doc_selector_label = ui.label(self.active_doc_name).classes('ice-doc-name')
                    ui.icon('expand_more', color='grey-4').classes('text-xs')
                    
                    with ui.menu() as self.doc_menu:
                        # 动态填充
                        pass
            
            # 如果已有记录的文档，尝试自动恢复并验证
            if template_state.current_doc:
                ui.timer(0.1, self.validate_and_restore_doc, once=True)

            # 2. 工作区
            with ui.element('div').classes('ice-workspace-inner'):
                # 资产树面板
                with ui.element('div').classes('ice-left-panel'):
                    with ui.element('div').classes('ice-search-box'):
                        ui.icon('search', color='grey-4').classes('text-sm')
                        search_input = ui.input(placeholder='搜索图层...').props('borderless dense').classes('flex-grow text-xs')
                        search_input.on_value_change(lambda e: self.update_search(e.value))
                        # 恢复刷新按钮，并确保其受到遮罩保护
                        ui.icon('refresh', color='primary').classes('text-sm cursor-pointer hover:rotate-180 transition-transform').on('click', self.refresh_assets)
                    
                    # 只看文本切换
                    with ui.row().classes('w-full items-center justify-between px-1 mb-2'):
                        with ui.row().classes('items-center gap-1'):
                            ui.icon('text_fields', size='14px', color='primary').classes('opacity-70')
                            ui.label('只看文本图层').classes('text-[10px] font-bold text-slate-500 uppercase tracking-wider')
                        
                        only_text_switch = ui.switch().props('size="sm" color="cyan-5"')
                        # 绑定配置
                        settings = local_config.data.setdefault('settings', {})
                        only_text_switch.value = settings.get('only_show_text_assets', False)
                        
                        def handle_only_text_change(e):
                            local_config.data['settings']['only_show_text_assets'] = e.value
                            local_config.save_to_disk()
                            self.render_assets_tree()
                            
                        only_text_switch.on_value_change(handle_only_text_change)

                    self.assets_tree_container = ui.element('div').classes('flex-grow overflow-y-auto scroll-hide')
                    # 一键添加按钮容器
                    self.batch_add_container = ui.element('div').classes('w-full pt-2 border-t border-slate-100 mt-auto')
                    
                    # 移除初始自动拉取资产，待用户手动触发或切换文档
                    # self.render_assets_tree() 
                    self.render_assets_tree() # 仅渲染空状态

                # 策略配置面板
                with ui.element('div').classes('ice-right-panel scroll-hide'):
                    self.strategy_cards_container = ui.element('div').classes('flex-grow')
                    self.render_strategy_cards()

            # 3. 悬浮操作岛
            with ui.element('div').classes('ice-float-dock'):
                with ui.element('div').classes('ice-float-btn').on('click', self.reset_strategy):
                    ui.icon('replay').classes('text-lg')
                    ui.tooltip('重置当前规则')
                
                ui.element('div').classes('ice-dock-sep')

                def activate_global_fx():
                    template_state.global_filter_active = True
                    self.render_strategy_cards()
                    ui.notify("已开启全局方案区块")

                with ui.element('div').classes('ice-float-btn').on('click', activate_global_fx) as global_fx_btn:
                    ui.icon('auto_fix_high').classes('text-lg')
                    ui.tooltip('开启全局方案')

                with ui.element('div').classes('ice-float-btn').on('click', self.add_render_preset) as render_cfg_btn:
                    ui.icon('movie_filter').classes('text-lg')
                    ui.tooltip('配置渲染方案')

                with ui.element('div').classes('ice-float-btn primary').on('click', self.save_strategy):
                    ui.icon('save').classes('text-lg')
                    ui.tooltip('保存模板')

    def _render_custom_menu(self, menu_obj, options, current_value, on_change, add_label=None, on_add=None, item_icon=None):
        """通用美化下拉菜单渲染器"""
        menu_obj.clear()
        with menu_obj:
            menu_obj.classes('ice-dropdown-menu')
            for opt in options:
                is_active = (opt == current_value)
                classes = 'ice-dropdown-item' + (' active' if is_active else '')
                # 这里使用 lambda 闭包来处理值变更和菜单关闭
                def make_click_handler(o):
                    return lambda: (on_change(o), menu_obj.close())
                
                with ui.menu_item(on_click=make_click_handler(opt)).classes(classes):
                    if item_icon:
                        ui.icon(item_icon, size='16px').classes('mr-2' + (' text-white' if is_active else ' text-gray-400'))
                    ui.label(opt).classes('whitespace-nowrap')
            
            if add_label:
                ui.element('div').classes('ice-dropdown-sep')
                with ui.element('div').classes('ice-dropdown-add-btn').on('click', lambda: (on_add() if on_add else None, menu_obj.close())):
                    ui.icon('add_circle_outline', size='16px')
                    ui.label(add_label).classes('whitespace-nowrap')

    def sync_open_docs(self):
        """同步打开的文档列表"""
        connection_overlay.show("正在同步文档列表...")
        async def on_docs(docs, err):
            with self.container:
                connection_overlay.close()
                if err:
                    ui.notify(f"获取文档失败: {err}", type='negative')
                    return
                
                with self.doc_menu:
                    doc_names = [d.get('name', '未命名') for d in docs]
                    self._render_custom_menu(
                        self.doc_menu, 
                        doc_names, 
                        self.active_doc_name, 
                        on_change=self.switch_document
                    )
                self.doc_menu.open()

        asyncio.create_task(ps_server.list_open_documents(callback=on_docs))

    def switch_document(self, name):
        """切换当前处理的文档"""
        self.active_doc_name = name
        template_state.current_doc = name
        if self.doc_selector_label:
            self.doc_selector_label.set_text(name)
        
        # 切换文档时先收起出图面板
        if rapid_export_panel:
            rapid_export_panel.hide()
            
        # 内部包含了重置 UI、验证存在、转移焦点、刷新资产和读取策略的完整链条
        self.refresh_assets(is_switch=True)

    def update_search(self, val):
        self.search_filter = val.lower()
        with self.assets_tree_container:
            self.render_assets_tree()

    def refresh_assets(self, is_switch=False):
        """
        从 PS 刷新资产树，并强制同步当前焦点文档以防止写入偏移
        is_switch: 是否为切换文档操作，如果是则强制清空当前 UI 状态
        """
        if not template_state.current_doc:
            return

        if is_switch:
            template_state.reset()
            self.render_strategy_cards()

        connection_overlay.show("正在拉取 Photoshop 状态...")
        
        async def sync_and_refresh():
            # 1. 首先验证文档是否还存在于 PS 中
            async def on_docs(docs, err):
                if err:
                    with self.container:
                        connection_overlay.close()
                        ui.notify(f"拉取文档列表失败: {err}", type='negative')
                    return
                
                doc_names = [d.get('name') for d in docs]
                if template_state.current_doc not in doc_names:
                    with self.container:
                        connection_overlay.close()
                        ui.notify(f"错误: 文档 {template_state.current_doc} 已在 PS 中关闭", type='negative')
                        # 文档消失，重置工作台状态
                        template_state.current_doc = None
                        template_state.layer_tree = []
                        self.active_doc_name = "选择文档"
                        if self.doc_selector_label: self.doc_selector_label.set_text("选择文档")
                        template_state.reset()
                        self.render_assets_tree()
                        self.render_strategy_cards()
                    return

                # 2. 强制转移 PS 焦点到该文档，确保资产抓取和后续保存不偏移
                async def on_activated(success, act_err):
                    if not success:
                        with self.container:
                            connection_overlay.close()
                            ui.notify(f"同步文档焦点失败: {act_err}", type='negative')
                        return
                    
                    # 3. 焦点就绪后，抓取图层树
                    async def on_layers(tree, layer_err):
                        if layer_err:
                            with self.container:
                                connection_overlay.close()
                                ui.notify(f"加载图层资产失败: {layer_err}", type='negative')
                            return
                        
                        # 全量注入路径
                        def _inject_paths(nodes, current_path="主文档"):
                            for node in nodes:
                                # 严格去除首尾空格并同步回 node['name']，确保后续所有路径构造都一致
                                node_name = str(node.get('name', '')).strip()
                                node['name'] = node_name 
                                node_path = f"{current_path} > {node_name}"
                                node['path'] = node_path
                                if 'children' in node:
                                    _inject_paths(node['children'], node_path)
                        _inject_paths(tree)
                        
                        template_state.layer_tree = tree
                        with self.assets_tree_container:
                            self.render_assets_tree()
                        
                        # 4. 如果是切换文档，顺便尝试读取 XMP 策略
                        if is_switch:
                            async def on_strategy(strategy_data, s_err):
                                with self.container:
                                    connection_overlay.close()
                                    # 无论是否有 strategy_data，都调用 deserialize 处理 (内部会 handle 空数据并 ensure 方案)
                                    StrategyLoader.deserialize(strategy_data or {}, template_state, tree)
                                    
                                    if strategy_data:
                                        ui.notify("已从 PSD 载入保存的策略", type='positive')
                                        if rapid_export_panel:
                                            rapid_export_panel.show()
                                            rapid_export_panel.update_mapping(strategy_data)
                                    else:
                                        if rapid_export_panel:
                                            rapid_export_panel.hide()
                                    
                                    self.render_strategy_cards()
                                    ui.notify("图层资产已载入", type='positive')
                            
                            await ps_server.read_strategy(callback=on_strategy)
                        else:
                            # 仅刷新：更新卡片校验状态（路径可能因为图层改名变红）
                            with self.container:
                                connection_overlay.close()
                                template_state.ensure_render_preset() # 确保刷新后也有方案
                                self.render_strategy_cards()
                                ui.notify("图层资产已刷新", type='positive')

                    await ps_server.request_layers(callback=on_layers)

                await ps_server.activate_psd(name=template_state.current_doc, callback=on_activated)

            await ps_server.list_open_documents(callback=on_docs)

        asyncio.create_task(sync_and_refresh())

    def validate_and_restore_doc(self):
        """验证之前的文档是否还在 PS 中，如果在则自动恢复"""
        if not template_state.current_doc:
            return
            
        async def on_docs(docs, err):
            if err: return
            
            doc_names = [d.get('name') for d in docs]
            if template_state.current_doc in doc_names:
                # 文档仍在线，重新执行带焦点切换的流程以恢复状态
                with self.container:
                    self.switch_document(template_state.current_doc)
            else:
                # 文档已关闭，重置状态
                template_state.current_doc = None
                template_state.layer_tree = []
                self.active_doc_name = "未连接"
                if self.doc_selector_label:
                    self.doc_selector_label.set_text("未连接")
                with self.assets_tree_container:
                    self.render_assets_tree()

        asyncio.create_task(ps_server.list_open_documents(callback=on_docs))

    def render_assets_tree(self):
        """渲染左侧资产树"""
        self.assets_tree_container.clear()
        if not template_state.layer_tree:
            return

        only_text = local_config.data.get('settings', {}).get('only_show_text_assets', False)

        def _has_renderable_content(node):
            """递归判断该节点或其子节点是否包含可配置项"""
            kind = node.get('kind')
            if kind == 'PIXEL': return False
            if kind in ['TEXT', 'SMARTOBJECT']: return True
            if kind == 'GROUP':
                return any(_has_renderable_content(c) for c in node.get('children', []))
            return False

        def _has_text_content(node):
            """递归判断该节点或其子节点是否包含文本图层"""
            kind = node.get('kind')
            if kind == 'TEXT': return True
            if kind in ['GROUP', 'SMARTOBJECT']:
                return any(_has_text_content(c) for c in node.get('children', []))
            return False

        def _matches_search(node):
            """判断节点是否应该在搜索结果中展示"""
            if not self.search_filter:
                return True
            
            name = node.get('name', '').lower()
            kind = node.get('kind', '')
            
            if kind in ['TEXT', 'SMARTOBJECT'] and self.search_filter in name:
                return True
            
            if kind == 'GROUP' or 'children' in node:
                return any(_matches_search(c) for c in node.get('children', []))
            
            return False

        def _draw_node(nodes, level=0, current_path=None):
            if current_path is None: current_path = []
            
            for node in nodes:
                # 1. 基础过滤：不渲染像素图层或仅包含像素图层的组
                if not _has_renderable_content(node):
                    continue
                
                # 2. 只看文本过滤
                if only_text and not _has_text_content(node):
                    continue

                # 3. 搜索过滤
                if not _matches_search(node):
                    continue
                
                name = node.get('name', '未命名')
                kind = node.get('kind', '')
                node_id = node.get('id')
                is_expanded = node_id in template_state.expanded_nodes or self.search_filter # 搜索时强制展开
                has_children = kind in ['GROUP', 'SMARTOBJECT'] and any(_has_renderable_content(c) for c in node.get('children', []))

                # 4. 层级颜色提示
                if level == 0:
                    if has_children:
                        bg_style = 'background-color: rgba(6, 182, 212, 0.03); border-left: 3px solid rgba(6, 182, 212, 0.3);'
                    else:
                        bg_style = ''
                else:
                    bg_opacity = max(0.12 - (level - 1) * 0.03, 0.02)
                    bg_style = f'background-color: rgba(6, 182, 212, {bg_opacity}); border-left: 1px solid rgba(6, 182, 212, 0.1);'
                
                with ui.element('div').classes('ice-layer-item w-full mb-1').style(bg_style):
                    if kind in ['TEXT', 'SMARTOBJECT']:
                        full_path = " / ".join(current_path + [name])
                        ui.tooltip(full_path).classes('bg-slate-800 text-white text-[10px] px-2 py-1 rounded shadow-xl')

                    with ui.element('div').classes('ice-layer-info flex-grow'):
                        if has_children:
                            icon_name = 'expand_more' if is_expanded else 'chevron_right'
                            ui.icon(icon_name).classes('text-[10px] cursor-pointer mr-1 opacity-60 hover:opacity-100') \
                                .on('click', lambda n=node_id: self.toggle_node(n))
                        else:
                            ui.element('div').classes('w-4')
                            
                        if kind == 'TEXT':
                            ui.icon('text_fields', color='blue-5').classes('text-sm')
                        elif kind == 'SMARTOBJECT':
                            ui.icon('filter_frames', color='purple-5').classes('text-sm')
                        elif kind == 'GROUP':
                            ui.icon('folder', color='grey-5').classes('text-sm')
                        
                        ui.label(name).classes('truncate text-[11px] ml-1 font-medium')
                    
                    with ui.element('div').classes('flex items-center gap-1'):
                        if kind in ['TEXT', 'SMARTOBJECT']:
                            btn_add = ui.element('div').classes('ice-btn-add-layer')
                            with btn_add:
                                ui.icon('add').classes('text-[10px]')
                            btn_add.on('click', lambda n=node: self.add_to_strategy(n))
                            
                        if kind == 'SMARTOBJECT':
                            btn_filter = ui.element('div').classes('ice-btn-add-layer !bg-purple-500')
                            with btn_filter:
                                ui.icon('auto_fix_high').classes('text-[10px]')
                            btn_filter.on('click', lambda n=node: self.add_filter_to_strategy(n))

                if has_children and is_expanded:
                    _draw_node(node.get('children', []), level + 1, current_path + [name])

        with self.assets_tree_container:
            if only_text:
                # 扁平化渲染所有文本图层
                text_nodes = []
                def _collect_flattened_text(nodes, current_path):
                    for n in nodes:
                        name = str(n.get('name', '未命名')).strip()
                        path = current_path + [name]
                        if n.get('kind') == 'TEXT' and _matches_search(n):
                            # 为节点注入完整路径名用于显示
                            n['_display_path'] = " / ".join(path)
                            # 重新同步干净的路径到 node 本身
                            n['path'] = " > ".join(["主文档"] + path)
                            text_nodes.append(n)
                        if 'children' in n:
                            _collect_flattened_text(n['children'], path)
                
                _collect_flattened_text(template_state.layer_tree, [])
                
                if not text_nodes:
                    ui.label('未找到文本资产').classes('text-xs text-slate-400 p-4 italic text-center w-full')
                else:
                    for node in text_nodes:
                        name = node.get('name', '未命名')
                        node_id = node.get('id')
                        display_path = node.get('_display_path', name)
                        
                        with ui.element('div').classes('ice-layer-item w-full mb-1').style('background-color: rgba(6, 182, 212, 0.05);'):
                            ui.tooltip(display_path).classes('bg-slate-800 text-white text-[10px] px-2 py-1 rounded shadow-xl')
                            
                            with ui.element('div').classes('ice-layer-info flex-grow'):
                                ui.element('div').classes('w-4') # 占位
                                ui.icon('text_fields', color='blue-5').classes('text-sm')
                                ui.label(name).classes('truncate text-[11px] ml-1 font-medium')
                            
                            with ui.element('div').classes('flex items-center gap-1'):
                                btn_add = ui.element('div').classes('ice-btn-add-layer')
                                with btn_add:
                                    ui.icon('add').classes('text-[10px]')
                                btn_add.on('click', lambda n=node: self.add_to_strategy(n))
            else:
                # 正常递归渲染树结构
                _draw_node(template_state.layer_tree)

        # 更新一键添加按钮
        self.batch_add_container.clear()
        if only_text:
            # 统计当前可见的文本图层
            text_nodes = []
            def _collect_text(nodes):
                for n in nodes:
                    if n.get('kind') == 'TEXT' and _matches_search(n):
                        text_nodes.append(n)
                    if 'children' in n:
                        _collect_text(n['children'])
            _collect_text(template_state.layer_tree)

            if text_nodes:
                with self.batch_add_container:
                    with ui.button(on_click=self.add_all_text_to_strategy).props('unelevated no-caps dense').classes('w-full bg-cyan-500 hover:bg-cyan-600 text-white rounded-lg py-2 transition-all shadow-sm'):
                        with ui.row().classes('items-center gap-2'):
                            ui.icon('add_circle', size='16px')
                            ui.label(f'一键添加 {len(text_nodes)} 个文本').classes('text-xs font-bold')

    def toggle_node(self, node_id):
        """切换节点的展开/收起状态"""
        if node_id in template_state.expanded_nodes:
            template_state.expanded_nodes.remove(node_id)
        else:
            template_state.expanded_nodes.add(node_id)
        with self.assets_tree_container:
            self.render_assets_tree()

    def add_to_strategy(self, node):
        """添加到右侧策略配置"""
        kind = node.get('kind')
        success = False
        if kind == 'TEXT':
            success = template_state.add_text_rule(node)
        elif kind == 'SMARTOBJECT':
            success = template_state.add_image_rule(node)
        
        if success:
            self.render_strategy_cards()
            ui.notify(f"已添加图层: {node['name']}")

    def add_filter_to_strategy(self, node):
        """添加滤镜规则到右侧策略配置"""
        if template_state.add_filter_rule(node):
            self.render_strategy_cards()
            ui.notify(f"已添加滤镜配置: {node['name']}")

    def add_all_text_to_strategy(self):
        """一键添加所有符合条件的文本图层到策略"""
        text_nodes = []
        
        def _matches_search(node):
            if not self.search_filter: return True
            name = node.get('name', '').lower()
            kind = node.get('kind', '')
            if kind == 'TEXT' and self.search_filter in name: return True
            if 'children' in node:
                return any(_matches_search(c) for c in node['children'])
            return False

        def _collect_text(nodes):
            for n in nodes:
                if n.get('kind') == 'TEXT' and _matches_search(n):
                    text_nodes.append(n)
                if 'children' in n:
                    _collect_text(n['children'])
        
        _collect_text(template_state.layer_tree)
        
        if not text_nodes:
            ui.notify("没有找到可添加的文本图层", type='warning')
            return
            
        count = 0
        for node in text_nodes:
            if template_state.add_text_rule(node):
                count += 1
        
        if count > 0:
            self.render_strategy_cards()
            ui.notify(f"成功一键添加 {count} 个文本图层", type='positive')
        else:
            ui.notify("所选文本图层已全部存在于策略中", type='info')

    def render_strategy_cards(self):
        """渲染右侧配置卡片列表"""
        self.strategy_cards_container.clear()
        
        # 1. 文本规则
        if template_state.text_rules:
            with self.strategy_cards_container:
                with ui.element('div').classes('ice-section-header'):
                    ui.html('<i class="q-icon material-icons">edit</i> 文字修改', sanitize=False)
                for rule in template_state.text_rules:
                    self._create_text_card(rule)

        # 2. 图片规则
        if template_state.image_rules:
            with self.strategy_cards_container:
                with ui.element('div').classes('ice-section-header'):
                    ui.html('<i class="q-icon material-icons">image</i> 图片替换', sanitize=False)
                for rule in template_state.image_rules:
                    self._create_image_card(rule)
        
        # 3. 滤镜规则
        if template_state.filter_rules:
            with self.strategy_cards_container:
                with ui.element('div').classes('ice-section-header'):
                    ui.html('<i class="q-icon material-icons">auto_fix_high</i> 滤镜配置', sanitize=False)
                for rule in template_state.filter_rules:
                    self._create_filter_card(rule)
        
        # 4. 全局滤镜
        if template_state.global_filter_active:
            with self.strategy_cards_container:
                with ui.element('div').classes('ice-section-header'):
                    ui.html('<i class="q-icon material-icons">auto_awesome</i> 全局方案', sanitize=False)
                self._create_global_filter_card()
        
        # 5. 渲染方案
        if template_state.render_presets:
            with self.strategy_cards_container:
                with ui.element('div').classes('ice-section-header'):
                    ui.html('<i class="q-icon material-icons">movie_filter</i> 渲染方案', sanitize=False)
                for preset in template_state.render_presets:
                    self._create_render_preset_card(preset)

        if not template_state.text_rules and not template_state.image_rules and not template_state.filter_rules and not template_state.global_filter_active and not template_state.render_presets:
            with self.strategy_cards_container:
                ui.label('请从左侧点击 + 号添加图层配置').classes('text-center text-gray-400 mt-20 italic')

    def _create_text_card(self, rule):
        # 实时路径校验
        is_invalid = not StrategyParser._resolve_path_exists(template_state.layer_tree, rule['path'])
        
        card_classes = 'ice-config-card'
        if is_invalid:
            card_classes += ' border-red-500 bg-red-50/30'
            
        with ui.element('div').classes(card_classes):
            with ui.element('div').classes('ice-card-header'):
                with ui.element('div').classes('flex items-center gap-2'):
                    ui.label('T').classes('ice-tag-type ice-tag-text')
                    ui.label(rule['name']).classes('ice-card-title')
                    # 路径悬浮气泡：保持与资产树一致的 / 分隔格式 (并移除 "主文档 /" 前缀)
                    display_path = rule['path'].replace('主文档 > ', '').replace(' > ', ' / ')
                    ui.tooltip(display_path).classes('bg-slate-800 text-white text-[10px] px-2 py-1 rounded shadow-xl')
                    if is_invalid:
                        ui.icon('warning', color='red').classes('text-xs').tooltip('路径已失效')
                
                with ui.element('div').classes('flex items-center gap-2'):
                    if not is_invalid:
                        # Rx+ 按钮组
                        with ui.label('Rx+').classes('text-xs text-primary font-bold cursor-pointer hover:opacity-70') as rx_label:
                            with ui.menu() as rx_menu:
                                self._render_regex_menu(rule, rx_menu)
                            rx_label.on('click', rx_menu.open)
                        
                        # 模拟 Select 样式的按钮 (文本规则)
                        with ui.button(rule['mapping_key'] or '文字组 1').props('flat no-caps dense').classes('ice-doc-selector px-3 scale-90 origin-right text-cyan-700 font-bold') as select_btn:
                            ui.icon('expand_more', color='grey-4').classes('text-xs ml-1')
                            with ui.menu() as mapping_menu:
                                self._render_text_mapping_menu(rule, select_btn, mapping_menu)
                    
                    # 删除
                    ui.icon('delete').classes('ice-btn-delete').on('click', lambda: self.remove_rule('text', rule['path']))

            # 正则列表 (仅在有内容时渲染容器，以保持初始大小一致)
            if rule['regex_steps']:
                regex_container = ui.element('div').classes('w-full flex flex-col gap-1')
                with regex_container:
                    for idx, step in enumerate(rule['regex_steps']):
                        with ui.element('div').classes('ice-regex-item px-2 py-1 bg-white/50 border-dashed'):
                            # 上移按钮: 圆角矩形 + 小箭头
                            with ui.element('div').classes('flex items-center justify-center bg-gray-100/50 hover:bg-primary/10 rounded-md p-0.5 mr-1.5 cursor-pointer transition-colors group').on('click', lambda i=idx, r=rule: self.move_regex_up(r, i)):
                                ui.icon('north', size='10px').classes('text-gray-400 group-hover:text-primary')

                            ui.icon('functions', color='primary').classes('text-xs opacity-50')
                            ui.label(step.get('name', '未命名正则')).classes('flex-grow text-[11px] text-gray-600')
                            
                            # 编辑按钮 (根据 can_remove 决定是否可点击)
                            can_edit = step.get('can_remove', True)
                            edit_btn = ui.icon('edit').classes('text-xs transition-colors')
                            if can_edit:
                                edit_btn.classes('cursor-pointer hover:text-primary text-gray-400')
                                # 修改编辑逻辑：调用 update_regex_custom 以便同步到 config
                                edit_btn.on('click', lambda s=step, r=rule: RegexEditor(
                                    initial_data=s,
                                    on_save=lambda updated: self.update_regex_custom(s, updated),
                                    container=self.container
                                ).open())
                            else:
                                edit_btn.classes('text-gray-200 cursor-not-allowed')
                            
                            # 删除按钮
                            ui.icon('delete_outline').classes('text-xs cursor-pointer hover:text-red-500 text-gray-300').on('click', lambda i=idx, r=rule: self.remove_regex(r, i))

    def _create_image_card(self, rule):
        is_invalid = not StrategyParser._resolve_path_exists(template_state.layer_tree, rule['path'])
        card_classes = 'ice-config-card'
        if is_invalid:
            card_classes += ' border-red-500 bg-red-50/30'
            
        with ui.element('div').classes(card_classes):
            with ui.element('div').classes('ice-card-header'):
                with ui.element('div').classes('flex items-center gap-2'):
                    ui.label('IMG').classes('ice-tag-type ice-tag-img')
                    ui.label(rule['name']).classes('ice-card-title')
                    # 路径悬浮气泡：保持与资产树一致的 / 分隔格式 (并移除 "主文档 /" 前缀)
                    display_path = rule['path'].replace('主文档 > ', '').replace(' > ', ' / ')
                    ui.tooltip(display_path).classes('bg-slate-800 text-white text-[10px] px-2 py-1 rounded shadow-xl')
                    if is_invalid:
                        ui.icon('warning', color='red').classes('text-xs').tooltip('路径已失效')
                
                with ui.element('div').classes('flex items-center gap-2'):
                    if not is_invalid:
                        # 模拟 Select 样式的按钮 (图片规则)
                        with ui.button(rule['mapping_key'] or '图片组 1').props('flat no-caps dense').classes('ice-doc-selector px-3 scale-90 origin-right text-cyan-700 font-bold') as select_btn:
                            ui.icon('expand_more', color='grey-4').classes('text-xs ml-1')
                            with ui.menu() as mapping_menu:
                                self._render_image_mapping_menu(rule, select_btn, mapping_menu)
                    
                    ui.icon('delete').classes('ice-btn-delete').on('click', lambda: self.remove_rule('image', rule['path']))

    def _create_filter_card(self, rule):
        is_invalid = not StrategyParser._resolve_path_exists(template_state.layer_tree, rule['path'])
        card_classes = 'ice-config-card'
        if is_invalid:
            card_classes += ' border-red-500 bg-red-50/30'

        with ui.element('div').classes(card_classes):
            with ui.element('div').classes('ice-card-header'):
                with ui.element('div').classes('flex items-center gap-2'):
                    ui.label('FX').classes('ice-tag-type !bg-purple-50 !text-purple-500 !border-purple-100')
                    ui.label(rule['name']).classes('ice-card-title')
                    # 路径悬浮气泡：保持与资产树一致的 / 分隔格式 (并移除 "主文档 /" 前缀)
                    display_path = rule['path'].replace('主文档 > ', '').replace(' > ', ' / ')
                    ui.tooltip(display_path).classes('bg-slate-800 text-white text-[10px] px-2 py-1 rounded shadow-xl')
                    if is_invalid:
                        ui.icon('warning', color='red').classes('text-xs').tooltip('路径已失效')
                
                with ui.element('div').classes('flex items-center gap-2'):
                    if not is_invalid:
                        # 滤镜选择按钮 + 菜单 (使用右对齐 props 和自定义类以左移菜单)
                        with ui.label().classes('text-sm text-purple-400 cursor-pointer hover:text-purple-600') as add_fx_label:
                            ui.icon('add_circle_outline')
                            with ui.menu().props('anchor="bottom right" self="top right"').classes('ice-dropdown-menu ice-filter-menu') as fx_menu:
                                self._render_filter_type_menu(rule, fx_menu)
                            add_fx_label.on('click', fx_menu.open)
                    
                    ui.icon('delete').classes('ice-btn-delete').on('click', lambda: self.remove_rule('filter', rule['path']))

            # 滤镜列表
            if rule.get('filter_steps'):
                steps_container = ui.element('div').classes('w-full flex flex-col gap-1')
                with steps_container:
                    for idx, step in enumerate(rule['filter_steps']):
                        with ui.element('div').classes('ice-regex-item px-2 py-1 bg-white/50 border-dashed'):
                            # 排序按钮
                            with ui.element('div').classes('flex items-center justify-center bg-gray-100/50 hover:bg-purple-100 rounded-md p-0.5 mr-1.5 cursor-pointer transition-colors group').on('click', lambda i=idx, r=rule: self.move_filter_up(r, i)):
                                ui.icon('north', size='10px').classes('text-gray-400 group-hover:text-purple-500')

                            ui.icon('auto_fix_high', color='purple-4').classes('text-xs opacity-50')
                            # 显示滤镜显示名
                            f_def = next((f for f in template_state.system_filters if f['type'] == step['type']), None)
                            ui.label(f_def['name'] if f_def else step['type']).classes('flex-grow text-[11px] text-gray-600')
                            
                            # 设置按钮 (齿轮)
                            ui.icon('settings').classes('text-xs cursor-pointer hover:text-purple-500 text-gray-400 transition-colors') \
                                .on('click', lambda s=step, r=rule, fd=f_def: FilterEditor(
                                    filter_type_config=fd,
                                    initial_params=s['params'],
                                    on_save=lambda updated: (s['params'].update(updated), self.render_strategy_cards()),
                                    container=self.container
                                ).open())
                            
                            # 删除按钮
                            ui.icon('delete_outline').classes('text-xs cursor-pointer hover:text-red-500 text-gray-300').on('click', lambda i=idx, r=rule: self.remove_filter_step(r, i))

    def _render_filter_type_menu(self, rule, menu):
        """渲染滤镜类型选择菜单"""
        filter_names = [f['name'] for f in template_state.system_filters]
        self._render_custom_menu(
            menu,
            filter_names,
            current_value=None,
            on_change=lambda name: self.add_filter_step(rule, next(f for f in template_state.system_filters if f['name'] == name)),
            item_icon='auto_fix_high'
        )

    def add_filter_step(self, rule, filter_def):
        """添加一个具体的滤镜步骤并打开编辑器"""
        with self.container:
            # 构造符合规范的数据结构
            new_step = {
                "type": filter_def['type'],
                "params": {p['internal_name']: p['default'] for p in filter_def['params']}
            }
            # 立即打开参数编辑器
            FilterEditor(
                filter_type_config=filter_def,
                initial_params=new_step['params'],
                on_save=lambda params: (
                    new_step.update({"params": params}),
                    rule.setdefault('filter_steps', []).append(new_step),
                    self.render_strategy_cards(),
                    ui.notify(f"已添加滤镜: {filter_def['name']}")
                ),
                container=self.container
            ).open()

    def remove_filter_step(self, rule, index):
        with self.container:
            rule['filter_steps'].pop(index)
            self.render_strategy_cards()
            ui.notify("已移除滤镜步骤")

    def move_filter_up(self, rule, index):
        if index > 0:
            with self.container:
                steps = rule['filter_steps']
                steps[index], steps[index-1] = steps[index-1], steps[index]
                self.render_strategy_cards()
                ui.notify("已调整滤镜顺序")

    def _create_global_filter_card(self):
        """渲染全局滤镜卡片"""
        # 构造伪规则对象以重用渲染逻辑
        rule = {'name': '全局滤镜方案', 'filter_steps': template_state.global_filter_steps}
        with ui.element('div').classes('ice-config-card'):
            with ui.element('div').classes('ice-card-header'):
                with ui.element('div').classes('flex items-center gap-2'):
                    ui.label('GL').classes('ice-tag-type !bg-amber-50 !text-amber-600 !border-amber-100')
                    ui.label(rule['name']).classes('ice-card-title')
                
                with ui.element('div').classes('flex items-center gap-2'):
                    # 滤镜选择按钮 + 菜单 (使用右对齐 props 和自定义类以左移菜单)
                    with ui.label().classes('text-sm text-amber-500 cursor-pointer hover:text-amber-700') as add_fx_label:
                        ui.icon('add_circle_outline')
                        with ui.menu().props('anchor="bottom right" self="top right"').classes('ice-dropdown-menu ice-filter-menu') as fx_menu:
                            self._render_filter_type_menu(rule, fx_menu)
                        add_fx_label.on('click', fx_menu.open)
                    
                    ui.icon('delete').classes('ice-btn-delete').on('click', lambda: (
                        ui.notify("已重置全局滤镜"),
                        template_state.global_filter_steps.clear(), 
                        template_state.__setattr__('global_filter_active', False), 
                        self.render_strategy_cards()
                    ))

            # 复用滤镜列表展示逻辑
            if rule.get('filter_steps'):
                steps_container = ui.element('div').classes('w-full flex flex-col gap-1')
                with steps_container:
                    for idx, step in enumerate(rule['filter_steps']):
                        with ui.element('div').classes('ice-regex-item px-2 py-1 bg-white/50 border-dashed'):
                            # 排序按钮 (使用 amber 风格)
                            with ui.element('div').classes('flex items-center justify-center bg-gray-100/50 hover:bg-amber-100 rounded-md p-0.5 mr-1.5 cursor-pointer transition-colors group').on('click', lambda i=idx, r=rule: self.move_filter_up(r, i)):
                                ui.icon('north', size='10px').classes('text-gray-400 group-hover:text-amber-600')

                            ui.icon('auto_awesome', color='amber-6').classes('text-xs opacity-50')
                            f_def = next((f for f in template_state.system_filters if f['type'] == step['type']), None)
                            ui.label(f_def['name'] if f_def else step['type']).classes('flex-grow text-[11px] text-gray-600')
                            
                            # 设置按钮 (齿轮)
                            ui.icon('settings').classes('text-xs cursor-pointer hover:text-amber-600 text-gray-400 transition-colors') \
                                .on('click', lambda s=step, r=rule, fd=f_def: FilterEditor(
                                    filter_type_config=fd,
                                    initial_params=s['params'],
                                    on_save=lambda updated: (s['params'].update(updated), self.render_strategy_cards()),
                                    container=self.container
                                ).open())
                            
                            # 删除按钮
                            ui.icon('delete_outline').classes('text-xs cursor-pointer hover:text-red-500 text-gray-300').on('click', lambda i=idx, r=rule: self.remove_filter_step(r, i))

    def add_global_filter_step(self, filter_def):
        """添加一个全局滤镜步骤"""
        with self.container:
            new_step = {
                "type": filter_def['type'],
                "params": {p['internal_name']: p['default'] for p in filter_def['params']}
            }
            # 打开编辑器
            FilterEditor(
                filter_type_config=filter_def,
                initial_params=new_step['params'],
                on_save=lambda params: (
                    new_step.update({"params": params}),
                    template_state.global_filter_steps.append(new_step),
                    self.render_strategy_cards(),
                    ui.notify(f"已添加全局滤镜: {filter_def['name']}")
                ),
                container=self.container
            ).open()

    def add_regex(self, rule, preset_obj):
        """挂载正则预设到规则中"""
        with self.container:
            # 复制一份 preset 对象，防止修改原始配置
            new_step = preset_obj.copy()
            rule['regex_steps'].append(new_step)
            self.render_strategy_cards()
            ui.notify(f"已挂载正则: {new_step.get('name', '未命名')}")

    def remove_regex(self, rule, index):
        with self.container:
            ui.notify("已删除正则规则")
            rule['regex_steps'].pop(index)
            self.render_strategy_cards()

    def remove_rule(self, rule_type, path):
        with self.container:
            ui.notify("已移除图层规则")
            template_state.remove_rule(rule_type, path)
            self.render_strategy_cards()

    def move_regex_up(self, rule, index):
        """将正则项上移一位"""
        if index > 0:
            with self.container:
                steps = rule['regex_steps']
                steps[index], steps[index-1] = steps[index-1], steps[index]
                self.render_strategy_cards()
                ui.notify("已上移正则顺序")

    def _create_render_preset_card(self, preset):
        """渲染渲染方案卡片 (移除阻塞式路径校验)"""
        with ui.element('div').classes('ice-config-card'):
            with ui.element('div').classes('ice-card-header'):
                with ui.element('div').classes('flex items-center gap-2'):
                    ui.label('RD').classes('ice-tag-type !bg-indigo-50 !text-indigo-600 !border-indigo-100')
                    ui.label(preset['filename']).classes('ice-card-title truncate max-w-[150px]')
                
                with ui.element('div').classes('flex items-center gap-2'):
                    # 预览按钮 (眼睛)
                    ui.icon('visibility').classes('text-xs cursor-pointer hover:text-indigo-600 text-gray-400 transition-colors') \
                        .on('click', lambda p=preset: self.preview_render(p))
                    
                    # 编辑按钮 (齿轮)
                    ui.icon('settings').classes('text-xs cursor-pointer hover:text-indigo-600 text-gray-400 transition-colors') \
                        .on('click', lambda p=preset: RenderPresetEditor(
                            initial_data=p,
                            on_save=lambda updated: (p.update(updated), self.render_strategy_cards()),
                            container=self.container,
                            layer_tree=template_state.layer_tree,
                            workbench=self
                        ).open())
                    
                    # 删除按钮
                    ui.icon('delete').classes('ice-btn-delete').on('click', lambda p=preset: (
                        ui.notify("已移除渲染方案"),
                        template_state.render_presets.remove(p),
                        self.render_strategy_cards()
                    ))

            # 方案详情摘要
            with ui.element('div').classes('w-full flex flex-col gap-1 px-1'):
                with ui.element('div').classes('flex items-center justify-between text-[10px] text-slate-400'):
                    ui.label(f"格式: {preset['format'].upper()}")
                    if preset['tiling']['enabled']:
                        ui.label(f"平铺: {int(preset['tiling']['width'])}x{int(preset['tiling']['height'])}px")
                
                # 图层摘要
                layers_str = ", ".join([p.split(' > ')[-1] for p in preset.get('root_layers', [])])
                ui.label(f"图层: {layers_str}").classes('text-[10px] text-slate-500 truncate italic')

    def add_render_preset(self):
        """打开编辑器添加新的渲染方案"""
        # 自动生成一个简单的内部 ID 名
        preset_idx = len(template_state.render_presets) + 1
        
        # 加载持久化默认设置
        defaults = local_config.data.get('settings', {}).get('render_defaults', {})
        
        new_preset = {
            "name": f"方案 {preset_idx}",
            "description": "",
            "output_path": "./output",
            "filename": defaults.get('filename_template', "{文字组 1}_{模板名}_{时间}"),
            "format": defaults.get('format', "jpg"),
            "quality": 85,
            "root_layers": [],
            "tiling": copy.deepcopy(defaults.get('tiling', {
                "enabled": False,
                "width": 1920,
                "height": 1080,
                "ppi": 300
            }))
        }
        # 默认根据 PS 中的可见性勾选根图层
        if template_state.layer_tree:
            visible_paths = [
                node.get('path', f"主文档 > {node['name']}") 
                for node in template_state.layer_tree 
                if node.get('visible', True)
            ]
            new_preset['root_layers'] = visible_paths

        RenderPresetEditor(
            initial_data=new_preset,
            on_save=lambda data: (
                template_state.render_presets.append(data),
                self.render_strategy_cards(),
                ui.notify("已添加渲染方案")
            ),
            container=self.container,
            layer_tree=template_state.layer_tree,
            workbench=self
        ).open()

    def _render_regex_menu(self, rule, menu):
        """渲染正则预设下拉菜单"""
        presets = local_config.data.get('text_replacement_presets', [])
        preset_names = [p['name'] for p in presets]
        
        self._render_custom_menu(
            menu,
            preset_names,
            current_value=None, # 正则菜单不需要选中高亮
            on_change=lambda v: self.add_regex(rule, next(p for p in presets if p['name'] == v)),
            add_label='自定义正则',
            on_add=lambda r=rule: RegexEditor(
                on_save=lambda data: self.add_regex_custom(r, data),
                container=self.container
            ).open()
        )

    def add_regex_custom(self, rule, data):
        """添加自定义正则并同步到本地配置"""
        with self.container:
            new_name = data.get('name', '').strip()
            presets = local_config.data.setdefault('text_replacement_presets', [])
            
            # --- 检测锁：防止重名冲突 ---
            for p in presets:
                if p['name'] == new_name:
                    if not p.get('can_remove', True):
                        ui.notify(f"保存失败：'{new_name}' 是系统预设名称，不可占用", type='negative')
                        return
                    else:
                        ui.notify(f"保存失败：已存在同名正则 '{new_name}'", type='warning')
                        return

            # 确保标记为可移除（非系统预设）
            data['can_remove'] = True
            presets.append(data.copy())
            local_config.save_to_disk()
            
            # 添加到当前文字规则
            rule['regex_steps'].append(data)
            self.render_strategy_cards()
            ui.notify(f"已保存并添加正则: {data['name']}")

    def update_regex_custom(self, old_step, new_data):
        """更新正则逻辑并同步到本地配置"""
        with self.container:
            old_name = old_step.get('name')
            new_name = new_data.get('name', '').strip()
            presets = local_config.data.setdefault('text_replacement_presets', [])

            # --- 检测锁：修改名称时的冲突检查 ---
            if old_name != new_name:
                for p in presets:
                    if p['name'] == new_name:
                        ui.notify(f"修改失败：名称 '{new_name}' 已被占用", type='negative')
                        return

            old_step.update(new_data)
            
            # 如果该规则是用户定义的，则同步更新配置
            if old_step.get('can_remove', True):
                for i, p in enumerate(presets):
                    if p['name'] == old_name:
                        presets[i] = old_step.copy()
                        local_config.save_to_disk()
                        break
            
            self.render_strategy_cards()
            ui.notify(f"已更新正则配置: {new_data['name']}")

    def _render_text_mapping_menu(self, rule, btn, menu):
        """渲染文字映射菜单"""
        self._render_custom_menu(
            menu,
            template_state.text_groups,
            rule['mapping_key'],
            on_change=lambda v: (rule.__setitem__('mapping_key', v), btn.set_text(v)),
            add_label='新建',
            on_add=lambda: self.add_new_text_group(rule, btn, menu),
            item_icon='title'
        )

    def _render_image_mapping_menu(self, rule, btn, menu):
        """渲染图片映射菜单"""
        self._render_custom_menu(
            menu,
            template_state.image_groups,
            rule['mapping_key'],
            on_change=lambda v: (rule.__setitem__('mapping_key', v), btn.set_text(v)),
            add_label='新建',
            on_add=lambda: self.add_new_image_group(rule, btn, menu),
            item_icon='image'
        )

    def add_new_text_group(self, rule, btn, menu):
        """新增文字变量组"""
        with self.container:
            new_num = len(template_state.text_groups) + 1
            new_name = f"文字组 {new_num}"
            template_state.text_groups.append(new_name)
            # 自动选中并刷新菜单
            rule['mapping_key'] = new_name
            btn.set_text(new_name)
            self._render_text_mapping_menu(rule, btn, menu)
            ui.notify(f"已新增变量列: {new_name}")

    def add_new_image_group(self, rule, btn, menu):
        """新增图片变量组"""
        with self.container:
            new_num = len(template_state.image_groups) + 1
            new_name = f"图片组 {new_num}"
            template_state.image_groups.append(new_name)
            rule['mapping_key'] = new_name
            btn.set_text(new_name)
            self._render_image_mapping_menu(rule, btn, menu)
            ui.notify(f"已新增图片变量: {new_name}")

    def reset_strategy(self):
        with self.container:
            template_state.reset()
            self.render_strategy_cards()
            ui.notify("配置已重置")

    async def preview_render(self, preset):
        """执行预览渲染 (针对特定渲染方案)"""
        # 准备数据包 - 仅流通路径，由插件端负责实时解析
        operations = []
        
        # 0. 计算变量组标号映射 (与 StrategyParser 保持一致)
        group_mapping = {}
        next_group_idx = 1
        for key in template_state.text_groups:
            if any(r['mapping_key'] == key for r in template_state.text_rules):
                group_mapping[key] = next_group_idx
                next_group_idx += 1
        
        # 辅助解析函数
        def resolve(path):
            return ps_server._resolve_layer_path(template_state.layer_tree, path)

        # 辅助函数：校验路径是否合法，不合法的直接跳过预览
        def is_valid(path):
            return StrategyParser._resolve_path_exists(template_state.layer_tree, path)

        for r in template_state.text_rules:
            layer_id, parent_chain, _ = resolve(r['path'])
            if layer_id is None: continue
            
            # 获取该图层所属的变量组标号
            g_idx = group_mapping.get(r['mapping_key'], 1)
            raw_preview_text = f"预览文本{g_idx}"
            
            # 应用统一正则引擎处理预览文本
            preview_text = apply_regex_steps(raw_preview_text, r['regex_steps'])
            operations.append({
                "type": "update_text_layer",
                "layer_id": layer_id,
                "parent_chain": parent_chain,
                "target_path": r['path'],
                "text": preview_text,
                "regex_steps": r['regex_steps']
            })
        
        for r in template_state.image_rules:
            layer_id, parent_chain, _ = resolve(r['path'])
            if layer_id is None: continue
            
            # 预览模式使用指定的占位图
            example_img_path = os.path.abspath(os.path.join(os.path.dirname(__file__), 'config', 'example.png'))
            
            operations.append({
                "type": "replace_image",
                "layer_id": layer_id,
                "parent_chain": parent_chain,
                "target_path": r['path'],
                "image_path": example_img_path
            })

        for r in template_state.filter_rules:
            layer_id, parent_chain, _ = resolve(r['path'])
            if layer_id is None: continue
            
            for step in r.get('filter_steps', []):
                operations.append({
                    "type": "apply_filter",
                    "layer_id": layer_id,
                    "parent_chain": parent_chain,
                    "target_path": r['path'],
                    "filter_type": step['type'],
                    "params": step['params']
                })
        
        # 使用传入的渲染方案
        temp_filename = f"preview_{datetime.datetime.now().strftime('%H%M%S_%f')}"
        
        # 解析 root_layers 路径为 root_ids
        root_ids = []
        for path in preset.get('root_layers', []):
            rid, _, _ = resolve(path)
            if rid: root_ids.append(rid)

        renders = [{
            "file_name": temp_filename,
            "format": preset['format'],
            "folder": os.path.abspath(TEMP_PREVIEW_DIR),
            "root_ids": root_ids,
            "root_layers": [path for path in preset.get('root_layers', []) if is_valid(path)],
            "tiling": preset.get('tiling', {}).get('enabled', False),
            "width": int(preset.get('tiling', {}).get('width', 0)),
            "height": int(preset.get('tiling', {}).get('height', 0)),
            "resolution": int(preset.get('tiling', {}).get('ppi', 300)),
            "filters": template_state.global_filter_steps if template_state.global_filter_active else []
        }]

        connection_overlay.show("正在同步预览焦点...")
        
        async def do_sync_and_preview():
            # 1. 确认文档还在
            async def on_docs(docs, err):
                if err or template_state.current_doc not in [d.get('name') for d in docs]:
                    with self.container:
                        connection_overlay.close()
                        ui.notify("文档已关闭，预览失败", type='negative')
                    return
                
                # 2. 激活
                async def on_activated(success, act_err):
                    if not success:
                        with self.container:
                            connection_overlay.close()
                            ui.notify("无法同步预览焦点", type='negative')
                        return
                    
                    # 3. 渲染
                    with self.container:
                        connection_overlay.update("loading", "正在生成预览渲染...", auto_close_delay=0)
                    async def on_preview(res, err):
                        with self.container:
                            connection_overlay.close()
                            if err:
                                ui.notify(f"预览失败: {err}", type='negative')
                                return
                            
                            # 拼接完整路径并打开
                            ext = preset['format'].lower()
                            file_path = os.path.abspath(os.path.join(TEMP_PREVIEW_DIR, f"{temp_filename}.{ext}"))
                            
                            if os.path.exists(file_path):
                                try:
                                    os.startfile(file_path)
                                    ui.notify("已通过默认软件打开预览", type='positive')
                                except Exception as e:
                                    ui.notify(f"打开预览失败: {e}", type='negative')
                            else:
                                ui.notify("预览文件生成失败", type='negative')
                    
                    await ps_server.execute_strategy_atomic(
                        operations=operations,
                        renders=renders,
                        debug=False,
                        callback=on_preview
                    )

                await ps_server.activate_psd(name=template_state.current_doc, callback=on_activated)

            await ps_server.list_open_documents(callback=on_docs)

        asyncio.create_task(do_sync_and_preview())

    async def save_strategy(self):
        """保存策略到 PSD"""
        if not template_state.current_doc:
            ui.notify("请先选择文档", type='warning')
            return

        # 1. 序列化与校验 (内部会检查是否有渲染方案)
        strategy_json, errors = StrategyParser.serialize(template_state, template_state.layer_tree)
        
        if errors:
            # 特殊处理：系统级校验错误 (如：缺少渲染方案、输出路径冲突)
            for err_type, err_info in errors:
                if err_type == 'system':
                    ui.notify(f"无法保存：{err_info}", type='warning')
                    return

            error_messages = []
            for err_type, err_info in errors:
                error_messages.append(f"路径失效: {err_info}")
            
            self.render_strategy_cards()
            ui.notify(f"检测到 {len(errors)} 条路径失效，保存失败！", type='negative')
            
            with ui.dialog() as diag, ui.card().classes('p-6 gap-4'):
                ui.label('保存失败：图层路径验证未通过').classes('text-lg font-bold text-red-600')
                ui.label('以下图层在 Photoshop 中已不存在，请删除对应规则后再试：').classes('text-sm text-gray-600')
                with ui.column().classes('w-full bg-red-50 p-3 rounded-lg border border-red-100 max-h-40 overflow-y-auto'):
                    for msg in error_messages:
                        ui.label(msg).classes('text-[11px] text-red-800 font-mono')
                ui.button('知道了', on_click=diag.close).classes('w-full bg-red-500 text-white')
            diag.open()
            return

        # 2. 准备执行同步与写入流程
        connection_overlay.show("正在同步文档焦点...")
        
        async def do_sync_and_save():
            # 2.1 再次确认文档还在 PS 里面
            async def on_docs(docs, err):
                if err:
                    with self.container:
                        connection_overlay.close()
                        ui.notify(f"同步文档失败: {err}", type='negative')
                    return
                
                doc_names = [d.get('name') for d in docs]
                if template_state.current_doc not in doc_names:
                    with self.container:
                        connection_overlay.close()
                        ui.notify(f"错误: 文档 {template_state.current_doc} 已在 PS 中关闭，无法保存", type='negative')
                    return

                # 2.2 激活文档，确保写入 XMP 到正确的文件
                async def on_activated(success, act_err):
                    if not success:
                        with self.container:
                            connection_overlay.close()
                            ui.notify(f"激活文档失败: {act_err}", type='negative')
                        return
                    
                    # 2.3 执行写入
                    with self.container:
                        connection_overlay.update("loading", "正在将保存模板到 PSD...", auto_close_delay=0)
                    
                    async def on_save(success, w_err):
                        with self.container:
                                    connection_overlay.close()
                                    if success:
                                        ui.notify("模板策略已保存到 PSD (XMP)", type='positive')
                                        if rapid_export_panel:
                                            rapid_export_panel.show()
                                            # 获取最新序列化后的数据来更新映射
                                            latest_strategy, _ = StrategyParser.serialize(template_state, template_state.layer_tree)
                                            rapid_export_panel.update_mapping(latest_strategy)
                                    else:
                                        ui.notify(f"保存失败: {w_err}", type='negative')

                    await ps_server.write_strategy(strategy_json, callback=on_save)

                await ps_server.activate_psd(name=template_state.current_doc, callback=on_activated)

            await ps_server.list_open_documents(callback=on_docs)

        asyncio.create_task(do_sync_and_save())

# --- 状态显示组件 (Reusable Status Overlay) ---

class ConnectionOverlay:
    """
    连接状态遮罩层：提供加载、成功、失败三种状态的平滑切换
    """
    def __init__(self):
        self.dialog = None
        self.content_wrap = None
        self.text_label = None
        self.state = 'loading' # loading, success, error

    def _get_loading_svg(self):
        return '''
        <svg class="status-spin" width="64" height="64" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
            <path d="M12 22c5.523 0 10-4.477 10-10S17.523 2 12 2" stroke="#06b6d4" stroke-width="2.5" stroke-linecap="round"/>
        </svg>
        '''

    def _get_success_svg(self):
        return '''
        <svg class="scale-in" width="64" height="64" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
            <circle cx="12" cy="12" r="10" fill="#10b981"/>
            <path class="checkmark-draw" d="M8 12l3 3 5-5" stroke="white" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"/>
        </svg>
        '''

    def _get_error_svg(self):
        return '''
        <svg class="scale-in" width="64" height="64" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
            <circle cx="12" cy="12" r="10" fill="#ef4444"/>
            <path d="M15 9l-6 6M9 9l6 6" stroke="white" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"/>
        </svg>
        '''

    def show(self, text: str = "正在连接 Photoshop..."):
        with ui.dialog().props('persistent transition-show=fade transition-hide=fade') as self.dialog:
            with ui.card().classes('p-12 rounded-[32px] ice-card bg-white items-center gap-8').style('min-width: 320px;'):
                self.icon_container = ui.html(self._get_loading_svg(), sanitize=False)
                self.text_label = ui.label(text).classes('text-lg font-bold text-cyan-800 tracking-tight transition-all')
        self.dialog.open()

    def update(self, state: str, text: str, auto_close_delay: float = 1.5):
        """
        更新状态并可选自动关闭
        state: 'success' | 'error' | 'loading'
        """
        self.state = state
        self.text_label.set_text(text)
        
        # 统一先移除所有可能的颜色类，再根据状态添加
        self.text_label.classes(remove='text-cyan-800 text-green-600 text-red-600')
        
        if state == 'success':
            self.icon_container.content = self._get_success_svg()
            self.text_label.classes('text-green-600')
        elif state == 'error':
            self.icon_container.content = self._get_error_svg()
            self.text_label.classes('text-red-600')
        
        if auto_close_delay > 0:
            ui.timer(auto_close_delay, self.close, once=True)

    def close(self):
        if self.dialog:
            self.dialog.close()

# 实例化全局单例
connection_overlay = ConnectionOverlay()

# --- 工作台侧边栏组件 ---

class WorkspaceSidebar:
    """
    工作台侧边栏：包含模式切换按钮、PS连接状态、到期时间显示
    """
    def __init__(self):
        self.current_mode = None  # 初始不选中任何模式
        self.mode_buttons = {}
        self.ps_dot = None
        self.ps_text = None
        self.expiry_icon = None
        self.expiry_tooltip = None
        self.sidebar_container = None
        self.ps_monitor_timer = None  # PS监控定时器
        self.ps_monitor_started = False  # 监控是否已启动
        
    def _get_template_icon(self):
        """制作模板图标"""
        return '''<svg viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
            <path d="M4 5a1 1 0 0 1 1-1h4a1 1 0 0 1 1 1v4a1 1 0 0 1-1 1H5a1 1 0 0 1-1-1V5zM4 15a1 1 0 0 1 1-1h4a1 1 0 0 1 1 1v4a1 1 0 0 1-1 1H5a1 1 0 0 1-1-1v-4zM14 5a1 1 0 0 1 1-1h4a1 1 0 0 1 1 1v4a1 1 0 0 1-1 1h-4a1 1 0 0 1-1-1V5zM14 15a1 1 0 0 1 1-1h4a1 1 0 0 1 1 1v4a1 1 0 0 1-1 1h-4a1 1 0 0 1-1-1v-4z" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
        </svg>'''
    
    def _get_quick_icon(self):
        """快速出图图标（闪电）"""
        return '''<svg viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
            <path d="M13 2L3 14h9l-1 8 10-12h-9l1-8z" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" fill="currentColor"/>
        </svg>'''
    
    def _get_batch_icon(self):
        """批量出图图标（列表）"""
        return '''<svg viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
            <path d="M8 6h13M8 12h13M8 18h13M3 6h.01M3 12h.01M3 18h.01" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
        </svg>'''
    
    def _get_calendar_icon(self):
        """日历图标"""
        return '''<svg viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
            <path d="M8 2v4M16 2v4M3 10h18M5 4h14a2 2 0 0 1 2 2v14a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2V6a2 2 0 0 1 2-2z" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
        </svg>'''
    
    def _get_info_icon(self):
        """关于图标"""
        return '''<svg viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
            <path d="M13 16h-1v-4h-1m1-4h.01M21 12a9 9 0 11-18 0 9 9 0 0118 0z" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
        </svg>'''
    
    def create(self, on_mode_change=None):
        """
        创建侧边栏UI
        
        Args:
            on_mode_change: 模式切换回调函数 (mode: str) -> None
        """
        self.on_mode_change = on_mode_change
        
        with ui.column().classes('ice-sidebar') as self.sidebar_container:
            # 顶部：功能导航组
            with ui.column().classes('ice-dock-top'):
                self._create_mode_button('template', '制作模板', self._get_template_icon())
                self._create_mode_button('quick', '快速出图', self._get_quick_icon())
                self._create_mode_button('batch', '批量出图', self._get_batch_icon())
            
            # 底部：系统与品牌组
            with ui.column().classes('ice-dock-bottom'):
                # 状态点
                self._create_ps_indicator()
                # 到期时间
                self._create_expiry_display()
                # 关于按钮
                self._create_about_button()
                # 品牌 Logo
                self._create_branding_logo()
        
        return self.sidebar_container
    
    def _create_mode_button(self, mode: str, label: str, icon_svg: str):
        """创建模式切换按钮"""
        is_active = (mode == self.current_mode)
        btn_classes = 'ice-mode-btn'
        if is_active:
            btn_classes += ' active'
        
        btn = ui.button().props('flat').classes(btn_classes).on('click', lambda m=mode: self._switch_mode(m))
        with btn:
            ui.html(icon_svg, sanitize=False)
            with ui.element('div').classes('ice-mode-info-bar'):
                ui.label(label).classes('ice-mode-info-text')
        
        self.mode_buttons[mode] = btn
    
    def _switch_mode(self, mode: str):
        """切换模式"""
        if mode == self.current_mode:
            return
        
        # 更新按钮状态
        for m, btn in self.mode_buttons.items():
            if m == mode:
                btn.classes(add='active', remove='')
            else:
                btn.classes(remove='active')
        
        self.current_mode = mode
        
        # 调用回调
        if self.on_mode_change:
            self.on_mode_change(mode)
    
    def _create_ps_indicator(self):
        """创建PS连接状态指示器"""
        with ui.element('div').classes('ice-ps-indicator'):
            # 状态点
            self.ps_dot = ui.element('div').classes('ice-ps-dot disconnected')
            # 底部通用提示
            with ui.element('div').classes('ice-bottom-tooltip'):
                self.ps_text = ui.label('等待连接...')
    
    def _create_expiry_display(self):
        """创建到期时间显示"""
        expiry_time = self._get_expiry_time()
        
        with ui.element('div').classes('ice-bottom-btn') as expiry_container:
            ui.html(self._get_calendar_icon(), sanitize=False)
            with ui.element('div').classes('ice-bottom-tooltip'):
                self.expiry_tooltip = ui.label(f'到期：{expiry_time}')
        
        self.expiry_icon = expiry_container

    def _create_about_button(self):
        """创建关于按钮"""
        with ui.element('div').classes('ice-bottom-btn').on('click', show_about_dialog):
            ui.html(self._get_info_icon(), sanitize=False)
            with ui.element('div').classes('ice-bottom-tooltip'):
                ui.label('关于软件')

    def _create_branding_logo(self):
        """创建品牌 Logo"""
        with ui.element('div').classes('ice-sidebar-logo'):
            ui.html(ABOUT_INFO['logo_svg'], sanitize=False).style('transform: scale(0.85);')
    
    def _get_expiry_time(self) -> str:
        """获取到期时间字符串"""
        if not auth_client.user_data:
            return '未知'
        
        expires_at = auth_client.user_data.get('expires_at')
        if expires_at:
            return str(expires_at)
        return '未知'
    
    def update_ps_status(self, connected: bool, waiting: bool = False):
        """更新PS连接状态"""
        if waiting:
            # 等待连接状态
            self.ps_dot.classes(remove='connected disconnected', add='waiting')
            self.ps_text.classes(remove='connected disconnected', add='waiting')
            self.ps_text.set_text('连接中...')
        elif connected:
            # 已连接
            self.ps_dot.classes(remove='disconnected waiting', add='connected')
            self.ps_text.classes(remove='disconnected waiting', add='connected')
            self.ps_text.set_text('已连接')
        else:
            # 未连接
            self.ps_dot.classes(remove='connected waiting', add='disconnected')
            self.ps_text.classes(remove='connected waiting', add='disconnected')
            self.ps_text.set_text('等待连接...')
    
    def start_ps_monitor(self):
        """启动PS连接状态监控（仅在登录成功后调用）"""
        if self.ps_monitor_started:
            return  # 避免重复启动
        
        self.ps_monitor_started = True
        last_status = None
        
        def check_status():
            nonlocal last_status
            is_connected = ps_server.is_connected()
            
            # 只在状态变化时更新
            if is_connected != last_status:
                last_status = is_connected
                
                if not is_connected:
                    # 断连：触发等待动画
                    self.update_ps_status(False, waiting=True)
                    # 显示等待遮罩（如果还没有显示）
                    try:
                        if not connection_overlay.dialog or not hasattr(connection_overlay.dialog, 'value') or not connection_overlay.dialog.value:
                            connection_overlay.show("正在等待 Photoshop 重新连接...")
                    except:
                        connection_overlay.show("正在等待 Photoshop 重新连接...")
                else:
                    # 已连接：关闭等待遮罩
                    self.update_ps_status(True, waiting=False)
                    try:
                        if connection_overlay.dialog:
                            connection_overlay.close()
                    except:
                        pass
        
        # 每秒检查一次
        self.ps_monitor_timer = ui.timer(1.0, check_status)

# --- 工作台视图 ---

# 全局变量：工作台容器和侧边栏
workspace_view = None
workspace_container = None
workspace_sidebar = None
current_workspace_content = None
login_shell_container = None  # 登录界面容器引用
template_workbench = None # 制作模板工作台实例
rapid_export_panel = None # 极速出图面板全局引用

def create_workspace_view():
    """
    创建工作台主视图 (功能模式容器)
    返回工作台容器
    """
    global workspace_sidebar, current_workspace_content, template_workbench, rapid_export_panel
    
    def on_mode_change(mode: str):
        """模式切换回调"""
        print(f"切换到模式: {mode}")
        
        # 切换内容区域
        content_area.clear()
        if mode == 'template':
            template_workbench = TemplateWorkbench(content_area)
            template_workbench.render()
        else:
            with content_area:
                ui.label(f'{mode} 模式正在开发中...').classes('text-xl text-gray-400 italic m-auto')
    
    with ui.column().classes('ice-workspace-container') as container:
        # 圆角矩形卡片容器
        with ui.row().classes('ice-workspace-card gap-0') as card:
            # 左侧：按钮区域（侧边栏）
            workspace_sidebar = WorkspaceSidebar()
            sidebar = workspace_sidebar.create(on_mode_change=on_mode_change)
            
            # 右侧：操作区域（主内容）
            with ui.column().classes('ice-workspace-content') as content_area:
                current_workspace_content = content_area
                # 初始显示欢迎
                ui.label('请在左侧选择操作模式').classes('text-xl text-gray-300 m-auto')

            # 初始化极速出图面板 (挂载在主窗口卡片上)
            rapid_export_panel = RapidExportPanel(card)
    
    return container

def switch_to_workspace():
    """切换到工作台界面"""
    global workspace_view, workspace_sidebar, login_shell_container
    
    # 执行过渡动画
    if login_shell_container:
        login_shell_container.classes(add='page-fade-out')
    
    async def show_workspace():
        # 延迟隐藏登录界面，等待动画完成
        await asyncio.sleep(0.4)
        if login_shell_container:
            login_shell_container.set_visibility(False)
        
        # 显示工作台并执行进入动画
        if workspace_view:
            with workspace_view: # 进入上下文以避免 Timer 创建错误
                workspace_view.classes(add='show page-fade-in')
                # 初始化PS状态并启动监控
                if workspace_sidebar:
                    is_connected = ps_server.is_connected()
                    workspace_sidebar.update_ps_status(is_connected, waiting=not is_connected)
                    workspace_sidebar.start_ps_monitor()
    
    asyncio.create_task(show_workspace())

# --- 通用对话框组件 ---

async def show_confirm_dialog(
    title: str = '确认操作',
    message: str = '确定要继续吗？',
    confirm_text: str = '确认',
    cancel_text: str = '取消',
    icon_type: str = 'warning'  # 'warning', 'info', 'error', 'success'
) -> bool:
    """
    通用确认对话框
    
    Args:
        title: 对话框标题
        message: 提示消息
        confirm_text: 确认按钮文字
        cancel_text: 取消按钮文字
        icon_type: 图标类型 (warning/info/error/success)
    
    Returns:
        bool: True=确认, False=取消
    """
    # 图标配置
    icon_configs = {
        'warning': {
            'gradient': 'linear-gradient(135deg, #fbbf24 0%, #f59e0b 100%)',
            'shadow': '0 4px 16px rgba(251, 191, 36, 0.15)',  # 降低阴影浓度
            'svg': '''<path d="M12 9v4m0 4h.01M21 12a9 9 0 11-18 0 9 9 0 0118 0z" stroke="#ffffff" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>'''
        },
        'error': {
            'gradient': 'linear-gradient(135deg, #ef4444 0%, #dc2626 100%)',
            'shadow': '0 4px 16px rgba(239, 68, 68, 0.2)',
            'svg': '''<path d="M10 14l2-2m0 0l2-2m-2 2l-2-2m2 2l2 2m7-2a9 9 0 11-18 0 9 9 0 0118 0z" stroke="#ffffff" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>'''
        },
        'info': {
            'gradient': 'linear-gradient(135deg, #3b82f6 0%, #2563eb 100%)',
            'shadow': '0 4px 16px rgba(59, 130, 246, 0.2)',
            'svg': '''<path d="M13 16h-1v-4h-1m1-4h.01M21 12a9 9 0 11-18 0 9 9 0 0118 0z" stroke="#ffffff" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>'''
        },
        'success': {
            'gradient': 'linear-gradient(135deg, #10b981 0%, #059669 100%)',
            'shadow': '0 4px 16px rgba(16, 185, 129, 0.2)',
            'svg': '''<path d="M9 12l2 2 4-4m6 2a9 9 0 11-18 0 9 9 0 0118 0z" stroke="#ffffff" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>'''
        }
    }
    
    config = icon_configs.get(icon_type, icon_configs['warning'])
    
    with ui.dialog() as dialog, ui.card().classes('p-6 rounded-3xl shadow-2xl').style('min-width: 320px; max-width: 400px;'):
        # 图标 - 改为圆角矩形
        ui.html(f'''
        <div style="display: flex; justify-content: center; margin-bottom: 16px;">
          <div style="width: 56px; height: 56px; border-radius: 16px; 
                      background: {config['gradient']}; 
                      display: flex; align-items: center; justify-content: center; 
                      box-shadow: {config['shadow']};">
            <svg width="28" height="28" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
              {config['svg']}
            </svg>
          </div>
        </div>
        ''', sanitize=False)
        
        # 标题和消息
        ui.label(title).classes('text-xl font-bold text-center mb-2 text-gray-800')
        ui.label(message).classes('text-sm text-center mb-6 text-gray-500')
        
        # 按钮 - 调换顺序，高亮按钮在左
        with ui.row().classes('w-full gap-3 justify-center'):
            ui.button(confirm_text, on_click=lambda: dialog.submit(True)).classes(
                'px-6 py-2 rounded-xl text-white'
            ).style(f'''
                min-width: 100px; 
                background: {config['gradient']}; 
                box-shadow: {config['shadow']};
            ''')
            
            ui.button(cancel_text, on_click=lambda: dialog.submit(False)).props('outline').classes(
                'px-6 py-2 rounded-xl text-gray-600 border-gray-300 hover:bg-gray-50'
            ).style('min-width: 100px;')
    
    return await dialog

# --- 窗口控制逻辑 ---

def minimize_window():
    """最小化窗口 - 使用 app.native.main_window"""
    if app.native.main_window:
        app.native.main_window.minimize()

async def close_window():
    """关闭窗口 - 带二次确认"""
    # 如果有快速出图任务正在运行，强制提示
    if rapid_export_panel and rapid_export_panel.is_running:
        confirmed = await show_confirm_dialog(
            title='强制退出？',
            message='当前有正在进行的渲染任务，退出将立即终止所有任务。确定要退出吗？',
            confirm_text='强制退出',
            cancel_text='取消',
            icon_type='error'
        )
        if confirmed:
            app.shutdown()
        return

    confirmed = await show_confirm_dialog(
        title='确认退出？',
        message='退出后将终止目前进行的任务',
        confirm_text='确认退出',
        cancel_text='取消',
        icon_type='warning'
    )
    if confirmed:
        app.shutdown()

# --- 界面构建 ---

with ui.element('div').classes('ice-win-controls'):
    # 最小化按钮
    min_btn = ui.button(on_click=minimize_window).props('flat dense').classes('ice-win-btn-minimal')
    with min_btn:
        ui.html('''<svg viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
          <path d="M5 12h14" stroke="currentColor" stroke-width="2.5" stroke-linecap="round"/>
        </svg>''', sanitize=False)
    
    # 关闭按钮
    close_btn = ui.button(on_click=close_window).props('flat dense').classes('ice-win-btn-minimal')
    with close_btn:
        ui.html('''<svg viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
          <path d="M18 6L6 18M6 6l12 12" stroke="currentColor" stroke-width="2.5" stroke-linecap="round"/>
        </svg>''', sanitize=False)

with ui.row().classes('ice-shell w-full') as login_shell:
    login_shell_container = login_shell
    # 左侧：信息展示卡片
    with ui.card().classes('w-72 p-10 ice-card bg-white/95 flex-shrink-0 h-[560px]'):
        # Logo & Title
        with ui.column().classes('w-full items-center mb-12'):
            ui.html('''
              <div class="ice-logo-wrap" style="margin-bottom: 16px;">
                <div class="ice-logo-badge" aria-hidden="true">
                  <svg width="44" height="44" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
                    <path d="M6.509 16.082a1 1 0 0 1-1.504-1.318L6.549 13H3a1 1 0 0 1-.993-.883L2 12a1 1 0 0 1 .883-.993L3 11h2.785L4.24 9.25a1 1 0 1 1 1.5-1.324L8.451 11h2.548V8.45L7.927 5.739a1 1 0 0 1-.165-1.313l.077-.099a1 1 0 0 1 1.312-.165l.1.077 1.748 1.544L11 3a1 1 0 0 1 .77-.974l.113-.02L12 2a1 1 0 0 1 .993.883L13 3l-.001 2.798 1.77-1.544a1 1 0 0 1 1.316 1.506l-3.086 2.694V11h2.548l2.702-3.076a1 1 0 0 1 1.312-.169l.099.077a1 1 0 0 1 .168 1.312l-.077.1L18.208 11H21a1 1 0 0 1 .993.883L22 12a1 1 0 0 1-.883.993L21 13h-2.793l1.545 1.764a1 1 0 0 1-.004 1.322l-.089.089a1 1 0 0 1-1.322-.004l-.089-.089L15.546 13h-2.547v2.548l3.087 2.693a1 1 0 0 1 .173 1.312l-.077.1a1 1 0 0 1-1.311.172l-.1-.076L13 18.201 13 21a1 1 0 0 1-.883.993L12 22a1 1 0 0 1-.993-.883L11 21l-.001-2.783-1.736 1.533a1 1 0 1 1-1.323-1.5l3.059-2.7V13H9.21l-2.701 3.082Z" fill="currentColor"/>
                  </svg>
                </div>
              </div>
            ''', sanitize=False)
            ui.label('ice美化助手').classes('text-2xl ice-title text-cyan-700 text-center mb-1')
            ui.label('ICE Tools').classes('text-[10px] text-gray-400 tracking-wider uppercase font-medium')
        
        # 信息组 (回归中文)
        with ui.column().classes('w-full items-start gap-8'):
            # 版本
            with ui.column().classes('gap-1'):
                ui.label('软件版本').classes('text-xs text-gray-400 font-bold')
                ui.label(cfg.APP_VERSION).classes('text-sm text-cyan-700 font-mono')
            
            # 网站
            with ui.column().classes('gap-1 cursor-pointer group').on('click', lambda: webbrowser.open(cfg.OFFICIAL_WEBSITE)):
                ui.label('官方网站').classes('text-xs text-gray-400 font-bold')
                ui.label('chestnutfish.chat').classes('text-sm text-gray-600 group-hover:text-cyan-600 transition-colors')
            
            # 支持
            with ui.column().classes('gap-1 cursor-pointer group').on('click', lambda: webbrowser.open(cfg.OFFICIAL_WEBSITE + "/support")):
                ui.label('技术支持').classes('text-xs text-gray-400 font-bold')
                ui.label('获取协助').classes('text-sm text-gray-600 group-hover:text-cyan-600 transition-colors')

        # 底部留白
        ui.element('div').classes('mt-auto')
        ui.label(f'© 2026 ICE').classes('text-[9px] text-gray-300 tracking-widest uppercase')

        error_text = ui.label('').classes('text-[10px] text-gray-500 text-center mt-2')
        error_text.visible = False

    # 右侧：账号卡片
    with ui.card().classes('w-[480px] p-10 ice-auth-card flex-shrink-0 h-[560px] flex flex-col justify-center'):
        with ui.tabs().classes('w-full') as tabs:
            login_tab = ui.tab('登录')
            register_tab = ui.tab('注册')
            renew_tab = ui.tab('续费')
        
        with ui.tab_panels(tabs, value=login_tab).classes('w-full bg-transparent flex-grow overflow-hidden'):
            # 登录面板
            with ui.tab_panel(login_tab).classes('p-4 h-full flex flex-col justify-center gap-6'):
                with ui.column().classes('w-full gap-4'):
                    username_input = ui.input('账号').props('outlined dense').classes('w-full')
                    username_input.value = local_config.get_auth_value('username', '')
                    
                    password_input = ui.input('密码', password=True).props('outlined dense').classes('w-full')
                    password_input.value = local_config.get_auth_value('password', '')
                    
                    with ui.row().classes('w-full items-center justify-between px-2'):
                        remember_checkbox = ui.switch('记住密码')
                       
                        remember_checkbox.props('dense size="xs" color="cyan-6"')
                        remember_checkbox.classes(
                        'text-xs text-gray-500 font-medium '
                        'hover:text-cyan-600 transition-colors select-none'
                        )
                        
                        remember_checkbox.value = local_config.get_auth_value('remember_password', False)
                        
                        #ui.label('忘记密码？').classes('text-xs text-cyan-600 cursor-pointer hover:underline').on('click', lambda: ui.notify('请联系管理员重置', type='info'))
                
                async def handle_login():
                    if not username_input.value or not password_input.value:
                        ui.notify('请输入账号密码', type='warning')
                        return
                    
                    login_btn.set_visibility(False)  # 临时隐藏/禁用防止重复点击
                    spinner = ui.spinner(size='lg').classes('self-center')
                    
                    try:
                        # 1. 预检设备
                        res = await auth_client.precheck(username_input.value, password_input.value)
                        if res.get("status") == "error":
                            ui.notify(res.get("message", "登录失败"), type='negative')
                            return
                        
                        # 2. 如果未绑定，弹出对话框
                        if not res.get("is_bound", True):
                            if res.get("remaining_slots", 0) <= 0:
                                ui.notify('设备绑定数已达上限', type='negative')
                                return
                            
                            confirm = await show_confirm_dialog(
                                title="绑定设备",
                                message=f"该账号尚未绑定此设备，是否消耗 1 个名额进行绑定？(剩余名额: {res.get('remaining_slots')})",
                                confirm_text="确认绑定",
                                icon_type="info"
                            )
                            if confirm:
                                bind_res = await auth_client.confirm_bind(username_input.value, password_input.value)
                                if bind_res.get("status") != "success":
                                    ui.notify(bind_res.get("message", "绑定失败"), type='negative')
                                    return
                            else:
                                return

                        # 3. 登录
                        login_res = await auth_client.login(username_input.value, password_input.value)
                        if login_res.get("status") == "success":
                            ui.notify('登录成功', type='positive')
                            global auth_logged_in
                            auth_logged_in = True
                            
                            # 保存配置
                            local_config.update_auth(
                                username=username_input.value,
                                password=password_input.value if remember_checkbox.value else "",
                                remember_password=remember_checkbox.value
                            )
                            
                            # --- 触发 PS 连接工作流 ---
                            connection_overlay.show("正在等待 Photoshop 连接...")
                            
                            # 启动服务器
                            try:
                                await ps_server.start()
                            except Exception as e:
                                connection_overlay.update("error", f"服务器启动失败: {e}")
                                return

                            # 等待连接 (轮询检查，带超时)
                            start_time = asyncio.get_event_loop().time()
                            timeout = 60.0 # 1分钟超时
                            
                            while not ps_server.is_connected():
                                if asyncio.get_event_loop().time() - start_time > timeout:
                                    connection_overlay.update("error", "连接超时，请确保插件已打开", auto_close_delay=3)
                                    return
                                await asyncio.sleep(0.5)
                            
                            # 连接成功！切换动画
                            connection_overlay.update("success", "Photoshop 连接成功！")
                            
                            # --- 启动心跳检测 ---
                            print(f">>> [DEBUG] 正在启动心跳检测: 频率={cfg.HEARTBEAT_INTERVAL}s, 最大重试={cfg.HEARTBEAT_MAX_RETRIES}")
                            # 1. 立即执行一次心跳
                            asyncio.create_task(do_heartbeat())
                            # 2. 开启定时器
                            heartbeat_timer.activate()
                            
                            # 延迟进入主界面 (等待成功动画播完)
                            await asyncio.sleep(1.5)
                            # 切换到工作台
                            switch_to_workspace()
                            ui.notify("欢迎使用工作台", color='positive')
                            # 登录后：恢复剪贴板监听等依赖连接的功能
                            try:
                                if rapid_export_panel:
                                    rapid_export_panel.on_logged_in()
                            except Exception:
                                pass
                            
                        else:
                            ui.notify(login_res.get("message", "登录失败"), type='negative')
                    finally:
                        spinner.delete()
                        login_btn.set_visibility(True)

                login_btn = ui.button('立即登录', on_click=handle_login).classes('w-full py-3 rounded-xl bg-gradient-to-r from-cyan-500 to-cyan-600 text-white font-bold shadow-lg')
            
            # 注册面板
            with ui.tab_panel(register_tab).classes('p-4 h-full flex flex-col justify-center gap-6'):
                with ui.column().classes('w-full gap-4'):
                    reg_username = ui.input('账号').props('outlined dense').classes('w-full')
                    
                    # 邮箱输入框带后缀下拉 (位置修复)
                    with ui.row().classes('w-full items-center no-wrap'):
                        reg_email = ui.input('邮箱').props('outlined dense').classes('flex-grow')
                        with reg_email:
                            # 更改锚点到右侧，并使用自定义类调整箭头位置
                            with ui.menu().props('anchor="bottom right" self="top right"').classes('ice-dropdown-menu ice-email-menu') as email_menu:
                                for suffix in ['@qq.com','@foxmail.com','@gmail.com', '@163.com', '@outlook.com', '@126.com']:
                                    with ui.menu_item(on_click=lambda s=suffix: (reg_email.set_value((reg_email.value or '').split('@')[0] + s), email_menu.close())).classes('ice-dropdown-item'):
                                        ui.label(suffix).classes('whitespace-nowrap')
                            # 将按钮移至文本框内部尾部
                            ui.button(icon='expand_more').props('flat dense').on('click', email_menu.open).classes('text-gray-400')
                    
                    reg_password = ui.input('密码', password=True).props('outlined dense').classes('w-full')
                    
                    # 验证码行
                    with ui.row().classes('w-full items-center no-wrap gap-3'):
                        reg_code = ui.input('验证码').props('outlined dense').classes('flex-grow')
                        
                        async def get_code():
                            if not reg_username.value or not reg_email.value:
                                ui.notify('请先填写账号和邮箱', type='warning')
                                return
                            res = await auth_client.get_auth_code(reg_username.value, reg_email.value)
                            if res.get("status") == "success":
                                ui.notify('验证码已发送，请检查邮箱', type='positive')
                            else:
                                ui.notify(res.get("message", "发送失败"), type='negative')
                                
                        ui.button('发送验证码', on_click=get_code).props('outline dense').classes('text-cyan-600 px-4 text-xs border-cyan-200 hover:bg-cyan-50 h-10 rounded-xl')
                
                async def handle_register():
                    if not reg_username.value or not reg_password.value or not reg_email.value or not reg_code.value:
                        ui.notify('请完善注册信息', type='warning')
                        return
                    res = await auth_client.register(reg_username.value, reg_email.value, reg_password.value, reg_code.value)
                    if res.get("status") == "success":
                        ui.notify('注册成功', type='positive')
                        tabs.set_value('登录')
                    else:
                        ui.notify(res.get("message", "注册失败"), type='negative')

                ui.button('确认注册', on_click=handle_register).classes('w-full py-3 rounded-xl bg-gradient-to-r from-cyan-500 to-cyan-600 text-white font-bold shadow-lg')
            
            # 续费面板
            with ui.tab_panel(renew_tab).classes('p-4 h-full flex flex-col justify-center gap-10'):
                with ui.column().classes('w-full gap-6'):
                    renew_username = ui.input('账号').props('outlined dense').classes('w-full')
                    renew_username.value = local_config.get_auth_value('username', '')
                    
                    renew_password = ui.input('密码', password=True).props('outlined dense').classes('w-full')
                    renew_password.value = local_config.get_auth_value('password', '')
                    
                    renew_key = ui.input('卡密').props('outlined dense').classes('w-full')
                
                async def handle_renew():
                    if not renew_username.value or not renew_password.value or not renew_key.value:
                        ui.notify('请完整填写续费信息', type='warning')
                        return
                    res = await auth_client.renew(renew_username.value, renew_password.value, renew_key.value)
                    if res.get("status") == "success":
                        ui.notify(f'续费成功！新到期时间: {res.get("new_expiry_time")}', type='positive')
                    else:
                        ui.notify(res.get("message", "核销失败"), type='negative')

                ui.button('核销卡密', on_click=handle_renew).classes('w-full py-3 rounded-xl bg-gradient-to-r from-cyan-500 to-cyan-600 text-white font-bold shadow-lg')

        # 移除底部的 Footer，保持右侧容器纯净

# 创建工作台视图（通过 CSS 控制显示，避免 set_visibility(False) 导致的 display:none 破坏动画）
workspace_view = create_workspace_view()
# workspace_view.set_visibility(False) # 移除这一行，由 CSS 类 'show' 控制

# 启动定时器：每秒检测连接状态
ui.timer(1.0, check_connection_status)

# 在运行前初始化 PS 服务器
ui.timer(0.1, init, once=True)

# 设置窗口属性
app.native.window_args.update({
    'resizable': True,
    'min_size': (960, 720)
})

# 运行
ui.run(title="ice美化助手", port=0, native=True, reload=False, frameless=True, window_size=(960, 720))